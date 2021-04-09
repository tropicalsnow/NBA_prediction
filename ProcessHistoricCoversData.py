import numpy as np

import requests


import openpyxl

from datetime import datetime
from bs4 import BeautifulSoup, NavigableString, Tag

import sys
sys.path.append('Libs')
from featureUtilities import Spread2Prob_Convert

def ProcessCoversData(history_file, starting_column=13, away_only=False):
 
    #Parse history files
    HISTORY_workbook = openpyxl.load_workbook(history_file)
    HISTORY_worksheet = HISTORY_workbook.worksheets[0]

    #Write headers
    if away_only:    
        #Check if header already exists and write header
        col_head = []
        hindx=1
        while True:
            header= HISTORY_worksheet.cell(1, hindx).value
            if header is None:
                break
            col_head.append(header)
            hindx=hindx+1
        #check if expert name is in the columns header
        try:
            starting_column = col_head.index("COVERS")+1                        
        except:
            #expert does not exist already in header so add it
            starting_column=len(col_head)+1


        HISTORY_worksheet.cell(1, starting_column).value= "COVERS" #Cell indices start from 1
    else:
        HISTORY_worksheet.cell(0+1, starting_column+1).value= "Covers consensus" #Cell indices start from 1
        HISTORY_worksheet.cell(0+1, starting_column+2+1).value= "Covers sides"   #Cell indices start from 1
        HISTORY_worksheet.cell(0+1, starting_column+4+1).value= "Covers picks"   #Cell indices start from 1


    history_games=HISTORY_worksheet.max_row-2

    #Loop over history data
    for i in  range(history_games):

        print("Checking game "+str(i))

        #Get game info
        away_team_Short =           HISTORY_worksheet.cell(i+2+1, 0+1).value   #Cell indices start from 1
        home_team_Short =           HISTORY_worksheet.cell(i+2+1, 1+1).value    #Cell indices start from 1
        historic_datetime_object  = datetime.strptime(HISTORY_worksheet.cell(i+2+1, 2+1).value,'%m/%d/%y')  #Cell indices start from 1


        #Construct url
        url = "https://contests.covers.com/Consensus/TopConsensus/NBA/Overall/"+str(historic_datetime_object.year)+"-"+str(historic_datetime_object.month).zfill(2)+"-"+str(historic_datetime_object.day).zfill(2)



        response = requests.get(url)
        HTMLtext = response.text.replace("<br />"," ")  #Remove line breaks and replace them with space
                                        
        soup = BeautifulSoup(HTMLtext, 'lxml')  # Parse the HTML as a string.

        tables = soup.find_all('table')  #grab any available tables. Can be more than one
        
        for table in tables:

            for row in table.find_all('tr'):
                columns = row.find_all('td')


                if columns:  #Make sure that columns are not empty (e.g. header columns)  
                    #Get home and away teams
                    teams_row_list=columns[0].text.split()
                    COVERS_away_team_Short=teams_row_list[1].upper()
                    if COVERS_away_team_Short=="BK" or COVERS_away_team_Short=="NETS":
                        COVERS_away_team_Short="BKN"
                    if COVERS_away_team_Short=="NY":
                        COVERS_away_team_Short="NYK"                
                    if COVERS_away_team_Short=="NO":
                        COVERS_away_team_Short="NOP"      
                    if COVERS_away_team_Short=="GS":
                        COVERS_away_team_Short="GSW"                
                    if COVERS_away_team_Short=="SA":
                        COVERS_away_team_Short="SAS"
                    if COVERS_away_team_Short=="PHO":
                        COVERS_away_team_Short="PHX"                

                    COVERS_home_team_Short=teams_row_list[2].upper()
                    if COVERS_home_team_Short=="BK" or COVERS_home_team_Short=="NETS":
                        COVERS_home_team_Short="BKN"
                    if COVERS_home_team_Short=="NY":
                        COVERS_home_team_Short="NYK"                
                    if COVERS_home_team_Short=="NO":
                        COVERS_home_team_Short="NOP"      
                    if COVERS_home_team_Short=="GS":
                        COVERS_home_team_Short="GSW"                
                    if COVERS_home_team_Short=="SA":
                        COVERS_home_team_Short="SAS"
                    if COVERS_home_team_Short=="PHO":
                        COVERS_home_team_Short="PHX"


                
                    #Get partial date too. Although the web address contains the full date
                    date_row_list=columns[1].text.split()
                    COVERS_datetime_object=datetime.strptime(date_row_list[1]+" "+date_row_list[2][0:2]+" "+str(historic_datetime_object.year), '%b %d %Y').date()  #Note: we add the historic YEAR to avoid problems with leap years
                    #TODO: Check that partial dates match
                
                    #Get consensus
                    cons_list=columns[2].text.split()
                    conensus_AWAY=int(cons_list[0][0:2].zfill(2))
                    conensus_HOME=int(cons_list[1][0:2].zfill(2))
            

                    #Get sides
                    sides_list=columns[3].text.split()
                    sides_AWAY=float(sides_list[0])
                    sides_HOME=float(sides_list[1])
            
                    #Get picks
                    picks_list=columns[4].text.split()
                    picks_AWAY=int(picks_list[0])
                    picks_HOME=int(picks_list[1])


                    #Check if data is valid
                    if COVERS_away_team_Short==away_team_Short and COVERS_home_team_Short == home_team_Short and \
                        COVERS_datetime_object.day==historic_datetime_object.day and COVERS_datetime_object.month==historic_datetime_object.month:

                        print(i, COVERS_away_team_Short, COVERS_home_team_Short, historic_datetime_object )

                        #Write to excel
                        if away_only:
                            probs_AWAY= Spread2Prob_Convert(sides_AWAY, [-0.04927,  0.1403, 0.5472, -0.04684])

                            HISTORY_worksheet.cell(3+i, starting_column).value=probs_AWAY  # Cell indices start from 1
                        else:
                            HISTORY_worksheet.cell(2+i+1, starting_column+1).value= conensus_AWAY #Away team consensus. Cell indices start from 1
                            HISTORY_worksheet.cell(2+i+1, starting_column+1+1).value= conensus_HOME #Home team consensus. Cell indices start from 1

                            HISTORY_worksheet.cell(2+i+1, starting_column+2+1).value= sides_AWAY #Away team sides. Cell indices start from 1
                            HISTORY_worksheet.cell(2+i+1, starting_column+3+1).value= sides_HOME #Home team sides. Cell indices start from 1
                            
                            HISTORY_worksheet.cell(2+i+1, starting_column+4+1).value= picks_AWAY #Away team picks. Cell indices start from 1
                            HISTORY_worksheet.cell(2+i+1, starting_column+5+1).value= picks_HOME #Home team picks. Cell indices start from 1



    
    
    HISTORY_workbook.save(history_file)
    


def main():

    history_file='.\Data\Market_data\MarketData_2020-21.xlsx'
    ProcessCoversData(history_file, starting_column=13)

  


if __name__ == '__main__':
    main()   