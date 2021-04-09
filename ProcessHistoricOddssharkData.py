#script to capture Oddsharks data from any historic file (i.e. MarketData or testfile )

import os
import time

from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

from bs4 import BeautifulSoup as soup
import numpy as np
import openpyxl
from datetime import datetime
import pandas as pd
from time import strptime
import math


import sys
sys.path.append('Libs')

from TeamsList import getTeam_by_Short 


def attempt_click(current_button, driver):

    try:
        current_button.click()
               
    except:
        print("\t\tCannot click button. Obscured by popup? Trying to click")
        time.sleep(5)
        accept_cookies_button=WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, 'agree-button')))
        accept_cookies_button.click()
        time.sleep(5)
        current_button.click()

def RotateCalendar(driver, historic_datetime_object):

    try:
        
        #open the popup calendar
        calendar_link=WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, 'dateselector__button-wrapper')))        
        attempt_click(calendar_link,driver)

        #select the correct month & year
        prev_month_button= driver.find_element_by_class_name("DayPicker-NavButton--prev")
        next_month_button= driver.find_element_by_class_name("DayPicker-NavButton--next")

        month_picker=driver.find_element_by_class_name("DayPicker-Caption")
        
        while True:
            [month, year]=month_picker.text.split()

            if int(year) != historic_datetime_object.year:
                if int(year)< historic_datetime_object.year:
                    attempt_click(next_month_button, driver)
                elif int(year) > historic_datetime_object.year:
                    attempt_click(prev_month_button, driver)

            if datetime.strptime(month, '%B').month != historic_datetime_object.month:
                if datetime.strptime(month, '%B').month < historic_datetime_object.month:
                    attempt_click(next_month_button, driver)
                elif datetime.strptime(month, '%B').month > historic_datetime_object.month:
                    attempt_click(prev_month_button, driver)

            if int(year) == historic_datetime_object.year and datetime.strptime(month, '%B').month == historic_datetime_object.month:
                break



        #Get the weekdays picker
        weekdays_picker=driver.find_element_by_class_name("DayPicker-Body")
        weekdays_picker_weeks= weekdays_picker.find_elements_by_class_name("DayPicker-Week")

        #loop through the weeks to find the correct date
        for i in range(len(weekdays_picker_weeks)):
            weekdays_picker_days=weekdays_picker_weeks[i].find_elements_by_class_name("DayPicker-Day")
            for j in range(len(weekdays_picker_days)):
                if weekdays_picker_days[j].text == str(historic_datetime_object.day):
                    attempt_click(weekdays_picker_days[j], driver)
                    
                    #close the popup calendar
                    attempt_click(calendar_link,driver)

                    break





        # #open the popup calendar
        # calendar_link=WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, 'spin-selector')))        
        # attempt_click(calendar_link,driver)
        
        # #Click on reset button
        # reset_button=WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.LINK_TEXT, 'Reset')))
        # attempt_click(reset_button,driver)

        # year_found=False
        # month_found=False
        # day_found=False

 
        # #Get handles for year, month and day(later)
        # dials = driver.find_elements_by_class_name("spin-selector__dial")
        # years= dials[0].find_elements_by_class_name("spin-selector__value")
        # months= dials[1].find_elements_by_class_name("spin-selector__value")

        
        
        # while True:


        #     #YEAR 

        #     cal_year= dials[0].find_element_by_class_name("active").text
        #     if len(cal_year)>4:
        #         cal_year=int(cal_year[5:9])
        #     else:
        #         cal_year=int(cal_year)
            
           
        #     if cal_year==historic_datetime_object.year:
        #         year_found=True
        #     else:
        #         #spin up/down the years
                
        #         #find current active year index
        #         index=0
        #         for i in range(len(years)):
        #             if int(years[i].text or 0) == cal_year:
        #                 index=i
        #                 break
                
        #         if cal_year>historic_datetime_object.year: #spin up
        #             years[index-1].click()
        #         else:  #spin down
        #             years[index+1].click()
                
        #         year_found=False




        #     #MONTH

        #     cal_month=dials[1].find_element_by_class_name("active").text
        #     if len(cal_month)>3:
        #         cal_month=cal_month[4:7]
        #     cal_month=strptime(cal_month,'%b').tm_mon #Convert to number


        #     if cal_month==historic_datetime_object.month:
        #         month_found=True
        #     else:
        #         #spin up/down the months
                
        #         #find current active year index
        #         index=0
        #         for i in range(len(months)):

        #             if months[i].text is '':
        #                 pass
        #             else:        
        #                 if  strptime(months[i].text,'%b').tm_mon == cal_month:
        #                     index=i
        #                     break
                
        #         if cal_month>historic_datetime_object.month: #spin up
        #             months[index-1].click()
        #         else:  #spin down
        #             months[index+1].click()

        #         month_found=False




        #     #DAY
        #     days = dials[2].find_elements_by_class_name("spin-selector__value") #get the days handle after the month has been selected since the number of days change

        #     cal_day=dials[2].find_element_by_class_name("active").text
        #     if len(cal_day)>2:
        #         cal_day=int(cal_day[3:5])
        #     else:
        #         cal_day=int(cal_day)


        #     if cal_day==historic_datetime_object.day:
        #         day_found=True
        #     else:
        #         #spin up/down the days
                
        #         #find current active year index
        #         index=0
        #         for i in range(len(days)):

        #             if int(days[i].text or 0) == cal_day:
        #                 index=i
        #                 break
                
        #         if cal_day>historic_datetime_object.day: #spin up
        #             days[index-1].click()
        #         else:  #spin down
        #             days[index+1].click()

        #         day_found=False


        #     if year_found and month_found and day_found:
        #         break
        #     else:
        #         time.sleep(0.5)


        # #Click on apply button
        # app_button=WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.LINK_TEXT, 'Apply')))
        # attempt_click(app_button,driver)

        return 1
    except:
        return 0








def RotateCalendar_old(driver, historic_datetime_object):

    try:

        #open the popup calendar
        calendar_link=WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, 'os-spin-selector')))        
        attempt_click(calendar_link,driver)


        #Click on reset button
        reset_button=WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.LINK_TEXT, 'Reset')))
        attempt_click(reset_button,driver)
        

    
        #Get the 3+3 (up and down) spin-up buttons (year,month,day)
        buttons_up=WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.CLASS_NAME, 'spin-up')))
        buttons_down=WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.CLASS_NAME, 'spin-down')))


        year_found=False
        month_found=False
        day_found=False


        while True:   

                   
            #Get displayed calendar data (year,month,day)       
            calendar_data=WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.CLASS_NAME, 'value-scroller')))
        

            #Get year
            cal_year=calendar_data[0].text
            if len(cal_year)>4:
                cal_year=int(cal_year[5:9])
            else:
                 cal_year=int(cal_year)


 
            #Get month
            cal_month=calendar_data[1].text
            if len(cal_month)>3:
                cal_month=cal_month[4:7]
            cal_month=strptime(cal_month,'%b').tm_mon #Convert to number

            #Get day
            cal_day=calendar_data[2].text
            if len(cal_day)>2:
                cal_day=int(cal_day[3:5])
            else:
                cal_day=int(cal_day)
            


            #Check year
            if cal_year==historic_datetime_object.year:
                year_found=True
            # elif cal_year>historic_datetime_object.year :
            #     #spin down the years    
            #     buttons_down[0].click()
            #     year_found=False
            else:
                #spin up the years
                attempt_click(buttons_up[0],driver)
                year_found=False

     
            #Check month
            if cal_month==historic_datetime_object.month:
                month_found=True
            # elif cal_month>historic_datetime_object.month:
            #     #spin down the months
            #     buttons_down[1].click()
            #     month_found=False
            else:
                #spin up the months
                attempt_click(buttons_up[1],driver)
                month_found=False

           #Check day
            if cal_day==historic_datetime_object.day:
                day_found=True
            # elif cal_day>historic_datetime_object.day:
            #     #spin down the days
            #     buttons_down[2].click()    
            #     day_found=False        
            else:
                #spin up the days
                attempt_click(buttons_up[2],driver)
                day_found=False

        

            if year_found and month_found and day_found:
                break
            else:
                time.sleep(0.5)



      
        #Click on apply button
        app_button=WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.LINK_TEXT, 'Apply')))
        attempt_click(app_button,driver)

        return 1
    except:
        return 0


def find_game_urls(driver, initial_url, historic_datetime_object):
 
    
 
    #Navigate to main ODDSHARK nba scores page
    driver.get(initial_url)
    time.sleep(5)
     

    #Go to the list of games for specific date
    calendar_done=0
   
    
    while calendar_done==0:
        calendar_done=RotateCalendar(driver, historic_datetime_object)
  

  
    time.sleep(5)
    gameURLlist=[]
    pre_fetch_Stats=[]  #AwaySpread, HomeSpread and ATS

    #Grab header text so as to check correct date (backup check)
    header=WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, 'header-text')))
    header_date_object=strptime(header.text,'%B %d %Y')
 
  
    if historic_datetime_object.year==header_date_object.tm_year and historic_datetime_object.month==header_date_object.tm_mon and historic_datetime_object.day==header_date_object.tm_mday:


        
        # Get the current page source.
        source = driver.page_source

        # Parse into soup() the source of the page after the link is clicked and use "html.parser" as the Parser.
        soupify = soup(source, 'html.parser')

        #loop for all available matches on that date. Differentiate between completed and upcoming matches
        soup_m1= soupify.find_all('div', {'class': 'matchup final'})
        soup_m2= soupify.find_all('div', {'class': 'matchup pre'})
        if len(soup_m1)>0:
            soup_m=soup_m1
        else:
            soup_m=soup_m2
        for match in soup_m:


            #Capture the URL links
            atags=match.find('a', {'class': 'scores-matchup__link'})
            game_url=atags['href']

            if  '/nba/' in game_url and str(historic_datetime_object.year) in game_url and  historic_datetime_object.strftime("%B").lower() in game_url:
                gameURLlist.append('https://www.oddsshark.com'+game_url)


            #Capture the stats
            dtags = match.find('div', {'class': 'scores-matchup__graph-section'})             
            spreads_header = dtags.find('div', {'class': 'graph--duel__header--middle'})
            spreads = spreads_header.find_all('div', {'class': 'graph--duel__header-value'})
            
            
            if spreads[0].text.upper()=="EV" or spreads[1].text.upper()=="EV":
                away_spread=0
                home_spread=0
            else:
                away_spread=   float(spreads[0].text or 0) 
                home_spread=   float(spreads[1].text or 0)

            

            ats_header_away=  dtags.find('div', {'class': 'graph--duel__header--left'})
            ats_away = ats_header_away.find('div', {'class': 'graph--duel__header-value'})

            ats_header_home=  dtags.find('div', {'class': 'graph--duel__header--right'})
            ats_home = ats_header_home.find('div', {'class': 'graph--duel__header-value'})
            
            if len(ats_away.text) == 0 and len(ats_home.text) >0:
                ats=float(ats_home.text.split()[0])
            elif len(ats_away.text) > 0 and len(ats_home.text)==0:
                ats=float(ats_away.text.split()[0])
            elif len(ats_home.text) == 0 and len(ats_home.text) ==0:
                ats=0
            
            
            
            pre_fetch_Stats.append([away_spread,home_spread,ats])


  
    return gameURLlist, pre_fetch_Stats


def scrape_game_data(driver,game_url,pre_fets_stats):

    # Initialisations here
    H2HList = np.zeros([6,2])
    LastNdata = np.zeros([7,2])
    awayMoneyline=0
    homeMoneyline=0
    away_spread=0
    home_spread=0
    ats=0


    # Navigate to specified page.
    while  True: 
        driver.get(game_url)
        try:
            WebDriverWait(driver, 5).until(EC.presence_of_all_elements_located((By.CLASS_NAME, 'os-menu-mainmenu__main-nav')))
            break
        except TimeoutException:
            print("Page did not load correctly. Reloading")

  

    ############## GET MONEYLINES #########################

    # Find the matchup tab element using a css selector and click it.
    try:
        matchup_tab=WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, 'react-tabs-0')))
        attempt_click(matchup_tab,driver)
        old_type=False
    except TimeoutException:
        print("\t\tCannot find Matchup tab. Probably an old ODDSSHARK page")
        old_type=True


 



    #Click to display ALL moneylines (i.e. "See More")
    try:
      

        # Get the current page source.
        source = driver.page_source

        # Parse into soup() the source of the page after the link is clicked and use "html.parser" as the Parser.
        soupify = soup(source, 'html.parser')



        ############## GET MEDIAN MONEYLINES FOR BOTH TEAMS ##############
        MoneyLineList = []


        if old_type:
            #No Matchup tab found so this is probably an old type ODDSSHARK webpage. We need to capture the data differently
            table_block  = soupify.find_all('div', {'class': 'matchup_page_block odds_block'})
            content_block =table_block[0].find_all('div', {'class': 'block_content'})
            for row in content_block[0].find_all('div', {'class': 'books'}):
                columns = row.text.split()    
                if len(columns)==6: #Only consider full columns
                    MoneyLineList.append(int(columns[0]))  #Away moneyline
                    MoneyLineList.append(int(columns[1]))  #Home moneyline
                
        else:
            
            for ultag in soupify.find_all('table', {'class': 'table table--compact-odds table--right-only table--reduced-padding table--react table--double-striped'}):
                for row in ultag.find_all('tr'):
                    columns = row.find_all('td')
                    if len(columns)>0 and len(columns[len(columns)-2].text)>0:
                        MoneyLineList.append(int(columns[len(columns)-2].text))


            #Remove the first two lines (opening lines)
            MoneyLineList.pop(0)
            MoneyLineList.pop(0)



        #Get median Moneyline for both teams
        awayMoneyline=np.median(MoneyLineList[::2])
        if math.isnan(awayMoneyline):
            awayMoneyline=0
        homeMoneyline=np.median(MoneyLineList[1::2])
        if math.isnan(homeMoneyline):
            homeMoneyline=0



        ############## GET H2H (LAST 10 GAMES) DATA #########################    
        if old_type:

            divtag = soupify.find_all('div', {'class': 'matchup_page_block head_summary_block'})
            away_rows=divtag[0].find_all('div', {'class': 'away_summary'})
            home_rows=divtag[0].find_all('div', {'class': 'home_summary'})
            away_text=away_rows[0].text.split()
            home_text=home_rows[0].text.split()

            Record=away_text[1].split("-") 
            H2HList[0,0]=int(Record[0]) #Away record
            H2HList[0,1]=int(Record[1]) #Home record

            H2HList[1,0]=float(away_text[4])  #Away score
            H2HList[1,1]=float(home_text[4])  #Home score

            H2HList[2,0]=float(away_text[6])  #Away FGP
            H2HList[2,1]=float(home_text[6])  #Home FGP

            H2HList[3,0]=float(away_text[7]) #Away Rebounds
            H2HList[3,1]=float(home_text[7]) #Home Rebounds

            H2HList[4,0]=eval(away_text[8])*100 #Away 3PP
            H2HList[4,1]=eval(home_text[8])*100 #Away 3PP

            H2HList[5,0]=float(away_text[9]) #Away steals
            H2HList[5,1]=float(home_text[9]) #Home steals

        else:
    

            ultag = soupify.find_all('table', {'class': 'table--teamstats'})


            rows =ultag[2].find_all('tr')  #The H2Htable is the 3rd table (the other two are the Team records table and the Edge Finder table)


            Record=rows[1].find_all('td')[0].text.split("-") 
            H2HList[0,0]=int(Record[0]) #Away record
            H2HList[0,1]=int(Record[1]) #Home record

            H2HList[1,0]=float(rows[4].find_all('td')[0].text)   #Away score
            H2HList[1,1]=float(rows[4].find_all('td')[1].text)  #Home score

            H2HList[2,0]=float(rows[6].find_all('td')[0].text.replace("%","")) #Away FGP
            H2HList[2,1]=float(rows[6].find_all('td')[1].text.replace("%","")) #Home FGP

            H2HList[3,0]=float(rows[4].find_all('td')[0].text) #Away Rebounds
            H2HList[3,1]=float(rows[4].find_all('td')[1].text)  #Home Rebounds

            H2HList[4,0]=eval(rows[8].find_all('td')[0].text)*100 #Away 3PP
            H2HList[4,1]=eval(rows[8].find_all('td')[1].text)*100 #Away 3PP

            H2HList[5,0]=float(rows[9].find_all('td')[0].text) #Away steals
            H2HList[5,1]=float(rows[9].find_all('td')[1].text) #Home steals



        ############## GET CONSENSUS SPREADS #########################

        if old_type:

            
            away_spread=pre_fets_stats[0]
            home_spread=pre_fets_stats[1]
            ats=pre_fets_stats[2]
        
        else:
            try:
                main_header = soupify.find_all('div', {'class': 'graph--duel'})
                spread_header= main_header[0].find('div', {'class': 'graph--duel__header--middle'})
                spreads=spread_header.find_all('div', {'class': 'graph--duel__header-value'})

                away_spread = float(spreads[0].text)
                home_spread = float(spreads[1].text)
            except:
                print("Spread data does not exist. Setting to zeros")
                away_spread=0
                home_spread=0




        ############## GET LAST N GAMES DATA #########################
        if old_type: #Only last 5 games
            
            divtag = soupify.find('div', {'class': 'matchup_page_block previous_games_block'})
            divtbl = divtag.find_all('div', {'class': 'base-table-wrapper'})
            
            #AWAY Team  
            tbody = divtbl[0].find_all('tbody')

            for row in tbody[1].find_all('tr'):
                columns = row.find_all('td')

                scr= columns[2].text.split("-") #score
                LastNdata[0,0]=LastNdata[0,0]+int(scr[0])
                LastNdata[1,0]=LastNdata[1,0]+int(scr[1])
                
                if int(scr[0])>int(scr[1]):
                    LastNdata[2,0]=LastNdata[2,0]+1 #win

                LastNdata[3,0]=LastNdata[3,0]+float(columns[4].text or 0) #line
                LastNdata[4,0]=LastNdata[4,0]+float(columns[7].text) #FG%
                LastNdata[5,0]=LastNdata[5,0]+float(columns[8].text) #FT%
                LastNdata[6,0]=LastNdata[6,0]+float(columns[9].text) #3PTM


            #HOME Team  
            tbody = divtbl[1].find_all('tbody')
            
            for row in tbody[1].find_all('tr'):
                columns = row.find_all('td')

                scr= columns[2].text.split("-") #score
                LastNdata[0,1]=LastNdata[0,1]+int(scr[0])
                LastNdata[1,1]=LastNdata[1,1]+int(scr[1])
                
                if int(scr[0])>int(scr[1]):
                    LastNdata[2,1]=LastNdata[2,1]+1 #win

                LastNdata[3,1]=LastNdata[3,1]+float(columns[4].text or 0) #line
                LastNdata[4,1]=LastNdata[4,1]+float(columns[7].text) #FG%
                LastNdata[5,1]=LastNdata[5,1]+float(columns[8].text) #FT%
                LastNdata[6,1]=LastNdata[6,1]+float(columns[9].text) #3PTM
        


            #Get the averages
            LastNdata=LastNdata/5


        else:   #10 last games here
            
            #AWAY Team
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, 'react-tabs-22'))).click()
            tbl= soupify.find('table', {'class': 'table table--react table--striped table--reduced-padding'})

            for row in tbl.find_all('tr', {'class': 'last-10-games-row gc-row'}):
                columns = row.find_all('td')
                
                scr= columns[2].text.split("-") #score
                LastNdata[0,0]=LastNdata[0,0]+int(scr[0])
                LastNdata[1,0]=LastNdata[1,0]+int(scr[1])

                if int(scr[0])>int(scr[1]):
                    LastNdata[2,0]=LastNdata[2,0]+1 #win
                                    
                try:
                    LastNdata[3,0]=LastNdata[3,0]+float(columns[4].text or 0) #line
                except:
                    LastNdata[3,0]=0


                LastNdata[4,0]=LastNdata[4,0]+float(columns[7].text) #FG%
                LastNdata[5,0]=LastNdata[5,0]+float(columns[8].text) #FT%
                LastNdata[6,0]=LastNdata[6,0]+float(columns[9].text) #3PTM




            #HOME Team
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, 'react-tabs-24'))).click()
            #we need to reload the source here    
            source = driver.page_source
            soupify = soup(source, 'html.parser')
        
            tbl= soupify.find('table', {'class': 'table table--react table--striped table--reduced-padding'})



            for row in tbl.find_all('tr', {'class': 'last-10-games-row gc-row'}):
                columns = row.find_all('td')
            
                scr= columns[2].text.split("-") #score
                LastNdata[0,1]=LastNdata[0,1]+int(scr[0])
                LastNdata[1,1]=LastNdata[1,1]+int(scr[1])

                if int(scr[0])>int(scr[1]):
                    LastNdata[2,1]=LastNdata[2,1]+1 #win
                
                try:
                    LastNdata[3,1]=LastNdata[3,1]+float(columns[4].text or 0) #line
                except:
                    LastNdata[3,1]=0

                LastNdata[4,1]=LastNdata[4,1]+float(columns[7].text) #FG%
                LastNdata[5,1]=LastNdata[5,1]+float(columns[8].text) #FT%
                LastNdata[6,1]=LastNdata[6,1]+float(columns[9].text) #3PTM

            #Get the averages
            LastNdata=LastNdata/10

    except TimeoutException:
        print("Cannot find ODDS table. Loading took too much time! Returning zeros ")

 

    return awayMoneyline, homeMoneyline ,away_spread, home_spread, ats, H2HList, LastNdata






 
def ProcessOddsharkData(history_file, starting_column=19):


    #Automate data capture from Oddsshark with selenium
    options = FirefoxOptions()
    options.add_argument("--headless")
    driver = webdriver.Firefox(options=options)
    initial_url="https://www.oddsshark.com/nba/scores"


    HISTORY_workbook = openpyxl.load_workbook(history_file)
    w_sheet = HISTORY_workbook.worksheets[0]

   
    #Write headers
    w_sheet.cell(0+1, starting_column+1).value="ODDSSHARK Moneyline"  #Cell indices start from 1
    w_sheet.cell(0+1, starting_column+2+1).value="ODDSSHARK Spreads"  #Cell indices start from 1  
    w_sheet.cell(0+1, starting_column+5+1).value="H2H Record"  #Cell indices start from 1
    w_sheet.cell(0+1, starting_column+7+1).value="H2H Score"  #Cell indices start from 1
    w_sheet.cell(0+1, starting_column+9+1).value="H2H FGP"  #Cell indices start from 1
    w_sheet.cell(0+1, starting_column+11+1).value="H2H Rebounds"  #Cell indices start from 1
    w_sheet.cell(0+1, starting_column+13+1).value="H2H 3PP"  #Cell indices start from 1
    w_sheet.cell(0+1, starting_column+15+1).value="H2H Steals"  #Cell indices start from 1


    #write extra headers
    w_sheet.cell(0+1, starting_column+18+1).value="Aw LastN Win pct"  #Cell indices start from 1
    w_sheet.cell(0+1, starting_column+19+1).value="Aw LastN Score"  #Cell indices start from 1
    w_sheet.cell(0+1, starting_column+21+1).value="Aw LastN Line"  #Cell indices start from 1  
    w_sheet.cell(0+1, starting_column+22+1).value="Aw LastN FG pct"  #Cell indices start from 1
    w_sheet.cell(0+1, starting_column+23+1).value="Aw LastN FT pct"  #Cell indices start from 1
    w_sheet.cell(0+1, starting_column+24+1).value="Aw LastN 3PTM pct"  #Cell indices start from 1
    w_sheet.cell(0+1, starting_column+26+1).value="Hm LastN Win pct"  #Cell indices start from 1
    w_sheet.cell(0+1, starting_column+27+1).value="Hm LastN Score"  #Cell indices start from 1
    w_sheet.cell(0+1, starting_column+29+1).value="Hm LastN Line" #Cell indices start from 1
    w_sheet.cell(0+1, starting_column+30+1).value="Hm LastN FG pct" #Cell indices start from 1
    w_sheet.cell(0+1, starting_column+31+1).value="Hm LastN FT pct" #Cell indices start from 1
    w_sheet.cell(0+1, starting_column+32+1).value="Hm LastN 3PTM pct" #Cell indices start from 1
    



    df = pd.read_excel(history_file,  engine='openpyxl')  #Use pandas to read excel data
    spreadsheet_length=len(df)
    

    #Capture the URLs for all the games in the excel file
    FullURLlist=[]
    FullPreFetchStats=[]
    starting_row=2 # Starting row of first game entry

    date_block_index=starting_row  #Loop over date blocks
    while (date_block_index <= spreadsheet_length):
        date_string=df['Date'][date_block_index-1]
        historic_datetime_object  = datetime.strptime(date_string,'%m/%d/%y') #Convert to python datetime object  = datetime.strptime(df['Date'][date_block_index],'%m/%d/%y') #Convert to python datetime object

        #Find all games in the spreadsheet at that specific date
        games_block=df[df['Date'].str.match(date_string,na=False)] #Ignore missing values
        games_block_size=len(games_block)

        #Get ALL game urls from ODDSSHARK for that specific date. Keep trying until we get something
        gameURLs_size=0
        while gameURLs_size==0:
            print("Trying to find URLs for date: ",historic_datetime_object.year,"/",historic_datetime_object.month,"/",historic_datetime_object.day)
            gameURLlist, pre_fetch_Stats=find_game_urls(driver,initial_url, historic_datetime_object)
            gameURLs_size=len(gameURLlist)


        if games_block_size==gameURLs_size: #These two should match
            for i in range(games_block_size):        

                away_team_Sort=games_block.values[i][0]
                TeamAWAY=getTeam_by_Short(away_team_Sort)
                home_team_Sort=games_block.values[i][1]
                TeamHOME=getTeam_by_Short(home_team_Sort)
                
                for j in range(gameURLs_size):

                    #Partial matching of team names in URL
                    if  TeamAWAY.name[0:5].lower().replace(" ","-")  in gameURLlist[j] and TeamHOME.name[0:5].lower().replace(" ","-")  in gameURLlist[j]:

                        print("URL found for: ",TeamAWAY.name," - ",TeamHOME.name,": ",gameURLlist[j])                    
                        #Match found then append URL to list
                        FullURLlist.append(gameURLlist[j])
                        FullPreFetchStats.append(pre_fetch_Stats[j])
                        break

            #Move to the next date block
            date_block_index=date_block_index+games_block_size
            print(date_block_index)    
    
    





    #Populate game ODDSSHARK stats from captured URLs
    print("URLs captured. Now capturing ODDSSHARK data")
    for i in range(len(FullURLlist)):

        print(i," ",df['Away'][i+starting_row-1]," - ",df['Home'][i+starting_row-1], ": ",FullURLlist[i])  #Remember: pandas ignores the empty rows. So we start from 1 not 2


        #Goto that url and capture data
        awayMoneyline, homeMoneyline ,away_spread, home_spread, ats, H2HList, LastNdata=scrape_game_data(driver,FullURLlist[i],FullPreFetchStats[i])

    
        #Append scraped data to excel
        w_sheet.cell(i+starting_row+1, starting_column+1).value= awayMoneyline  #Moneyline AWAY. Cell indices start from 1
        w_sheet.cell(i+starting_row+1, starting_column+1+1).value= homeMoneyline  #Moneyline HOME. Cell indices start from 1
        w_sheet.cell(i+starting_row+1, starting_column+2+1).value= away_spread  #spread AWAY. Cell indices start from 1
        w_sheet.cell(i+starting_row+1, starting_column+3+1).value= home_spread  #spread HOME. Cell indices start from 1
        w_sheet.cell(i+starting_row+1, starting_column+5+1).value= int(H2HList[0][0])  #H2H record AWAY. Cell indices start from 1
        w_sheet.cell(i+starting_row+1, starting_column+6+1).value= int(H2HList[0][1])  #H2H record HOME. Cell indices start from 1
        w_sheet.cell(i+starting_row+1, starting_column+7+1).value= float(H2HList[1][0])  #H2H Score AWAY. Cell indices start from 1
        w_sheet.cell(i+starting_row+1, starting_column+8+1).value= float(H2HList[1][1])  #H2H Score HOME. Cell indices start from 1       
        w_sheet.cell(i+starting_row+1, starting_column+9+1).value= float(H2HList[2][0])  #H2H FGP AWAY. Cell indices start from 1
        w_sheet.cell(i+starting_row+1, starting_column+10+1).value= float(H2HList[2][1])  #H2H FGP HOME. Cell indices start from 1        
        w_sheet.cell(i+starting_row+1, starting_column+11+1).value= float(H2HList[3][0])  #H2H Rebounds AWAY. Cell indices start from 1
        w_sheet.cell(i+starting_row+1, starting_column+12+1).value= float(H2HList[3][1])  #H2H Rebounds HOME. Cell indices start from 1
        w_sheet.cell(i+starting_row+1, starting_column+13+1).value= float(H2HList[4][0])  #H2H 3PP AWAY. Cell indices start from 1        
        w_sheet.cell(i+starting_row+1, starting_column+14+1).value= float(H2HList[4][1])  #H2H 3PP HOME. Cell indices start from 1        
        w_sheet.cell(i+starting_row+1, starting_column+15+1).value= float(H2HList[5][0])  #H2H Steals AWAY. Cell indices start from 1     
        w_sheet.cell(i+starting_row+1, starting_column+16+1).value= float(H2HList[5][1])  #H2H Steals HOME. Cell indices start from 1     
        
        


        #Append LastN data
        w_sheet.cell(i+starting_row+1, starting_column+18+1).value= LastNdata[2,0] #Aw Last 10GA Win pct. Cell indices start from 1     
        w_sheet.cell(i+starting_row+1, starting_column+19+1).value= LastNdata[0,0] #Aw Last 10GA Score. Cell indices start from 1     
        w_sheet.cell(i+starting_row+1, starting_column+20+1).value= LastNdata[1,0] #                     Cell indices start from 1             
        w_sheet.cell(i+starting_row+1, starting_column+21+1).value= LastNdata[3,0] #Aw Last 10GA Line. Cell indices start from 1     
        w_sheet.cell(i+starting_row+1, starting_column+22+1).value= LastNdata[4,0] #Aw Last 10GA FG pct. Cell indices start from 1     
        w_sheet.cell(i+starting_row+1, starting_column+23+1).value= LastNdata[5,0] #Aw Last 10GA FT pct. Cell indices start from 1     
        w_sheet.cell(i+starting_row+1, starting_column+24+1).value= LastNdata[6,0] #Aw Last 10GA 3PTM pct. Cell indices start from 1       
        w_sheet.cell(i+starting_row+1, starting_column+26+1).value= LastNdata[2,1] #Hm Last 10GA Win pct. Cell indices start from 1     
        w_sheet.cell(i+starting_row+1, starting_column+27+1).value= LastNdata[0,1] #Hm Last 10GA Score. Cell indices start from 1     
        w_sheet.cell(i+starting_row+1, starting_column+28+1).value= LastNdata[1,1] #                    Cell indices start from 1     
        w_sheet.cell(i+starting_row+1, starting_column+29+1).value= LastNdata[3,1]  #Hm Last 10GA Line. Cell indices start from 1     
        w_sheet.cell(i+starting_row+1, starting_column+30+1).value= LastNdata[4,1]  #Hm Last 10GA FG pct. Cell indices start from 1     
        w_sheet.cell(i+starting_row+1, starting_column+31+1).value= LastNdata[5,1]  #Hm Last 10GA FT pct. Cell indices start from 1     
        w_sheet.cell(i+starting_row+1, starting_column+32+1).value= LastNdata[6,1]   #Hm Last 10GA 3PTM pct. Cell indices start from 1     
        
        



    # We are done with the driver so quit.
    driver.quit()

    #write to the excel history file
    HISTORY_workbook.save(history_file)




def main():



    history_file='.\Data\Market_data\MarketData_2020-21.xlsx'
    
    ProcessOddsharkData(history_file,19)



if __name__ == '__main__':
    main()   