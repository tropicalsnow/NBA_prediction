from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver import Firefox

from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import WebDriverException
from pyquery import PyQuery as pyquery


import time
from datetime import datetime
import os.path, sys
sys.path.append('Libs')
import TeamsList
import numpy as np


import openpyxl
from dateutil import tz

class Season(object):
    def __init__(self,name):
        self.name = name
        self.games = list()
        self.urls = list()

    def add_game(self,game):
        self.games.append(game)

    def add_url(self,url):
        self.urls.append(url)



class Crawler(object):
    """
    A class to crawl links from oddsportal.com website.
    Makes use of Selenium and BeautifulSoup modules.
    """
    
    def __init__(self, wait_on_page_load=3):
        """
        Constructor
        """
        self.base_url = 'https://www.oddsportal.com'
        self.wait_on_page_load = wait_on_page_load
        if wait_on_page_load == None:
            self.wait_on_page_load = 3
        self.options = Options()
        self.options.add_argument('--headless')
        self.driver = webdriver.Firefox(options=self.options)


    def go_to_link(self,link):
        """
        returns True if no error
        False when page not found
        """
        self.driver.get(link)
        try:
            # If no Login button, page not found
            self.driver.find_element_by_css_selector('.button-dark')
        except NoSuchElementException:
            print('Problem with link, could not find Login button - %s' %(link))
            return False
        # Workaround for ajax page loading issue
        time.sleep(self.wait_on_page_load)
        return True

    def get_html_source(self):
        return self.driver.page_source


    def get_seasons_for_league(self, main_league_results_url):
        """
        Params:
            (str) main_league_results_url e.g. https://www.oddsportal.com/hockey/usa/nhl/results/

        Returns:
            (list) urls to each season for given league
        """
        seasons = []
        print('Getting all seasons for league via %s' %(main_league_results_url))
        if not self.go_to_link(main_league_results_url):
            print('League results URL loaded unsuccessfully %s' %(main_league_results_url))
            # Going to send back empty list so this is not processed further
            return seasons
        html_source = self.get_html_source()
        html_querying = pyquery(html_source)
        season_links = html_querying.find('div.main-menu2.main-menu-gray > ul.main-filter > li > span > strong > a')
        print('Extracted links to %d seasons' %(len(season_links)))
        for season_link in season_links:
            this_season = Season(season_link.text)
            # Start the Season's list of URLs with just the root one
            this_season_url = self.base_url + season_link.attrib['href']
            this_season.urls.append(this_season_url)
            seasons.append(this_season)
        return seasons

    
    def fill_in_season_pagination_links(self, season):
        """
        Params:
            (Season) object with just one entry in its urls field, to be modified
        """
        first_url_in_season = season.urls[0]
        self.go_to_link(first_url_in_season)
        html_source = self.get_html_source()
        html_querying = pyquery(html_source)
        # Check if the page says "No data available"
        no_data_div = html_querying.find('div.message-info > ul > li > div.cms')
        if no_data_div != None and no_data_div.text() == 'No data available':
            # Yes, found "No data available"
            print('Found "No data available", skipping %s' %(first_url_in_season))
            return
        # Just need to locate the final pagination tag
        pagination_links = html_querying.find('div#pagination > a')
        # It's possible, however, there is no pagination...
        if len(pagination_links) <= 1:
            return
        last_page_number = -1
        last_page_url = None
        for link in reversed(pagination_links):
            span = link.find('span')
            if span != None and span.text != None and '»|' in span.text:
                # This is the last link because it has these two characters in it...
                last_page_number = int(link.attrib['x-page'])
                last_page_url = first_url_in_season + link.attrib['href']
                break
        # If the last page number was set, the page format must've changed - RuntimeError
        if last_page_number == -1:
            print('Could not locate final page URL from %s' %(first_url_in_season))
            raise RuntimeError('Could not locate final page URL from %s', first_url_in_season)
        for i in range(2,last_page_number):
            this_url = last_page_url.replace('page/' + str(last_page_number), 'page/' + str(i))
            season.urls.append(this_url)
        season.urls.append(last_page_url)

    def close_browser(self):
        time.sleep(5)
        try:
            self.driver.quit()
        except WebDriverException:
            print('WebDriverException on closing browser - maybe closed?')

class Scraper(object):
    """
    A class to scrape/parse match results from oddsportal.com website.
    Makes use of Selenium and BeautifulSoup modules.
    """
    
    def __init__(self, wait_on_page_load=3):    
        """
        Constructor
        """
        self.base_url = 'https://www.oddsportal.com'
        self.wait_on_page_load = wait_on_page_load
        if wait_on_page_load == None:
            self.wait_on_page_load = 3
        self.options = Options()
        self.options.add_argument('--headless')
        self.driver = webdriver.Firefox(options=self.options)
        self.attemptLogin()

    def attemptLogin(self):

        try:
            self.go_to_link("https://www.oddsportal.com/login/")
            self.driver.find_element_by_id("login-username1").send_keys("vzografos")
            self.driver.find_element_by_id("login-password1").send_keys("ad689d13")
            self.driver.find_element_by_css_selector(".inline-btn-2[type=submit]").click()
        except:
            print("Could not login succesfully")
        
        

    def close_browser(self):
        time.sleep(5)
        try:
            self.driver.quit()
        except WebDriverException:
            print('WebDriverException on closing browser - maybe closed?')

    def go_to_link(self,link):
        """
        returns True if no error
        False whe page not found
        """
        self.driver.get(link)
        try:
            # if no Login button -> page not found
            self.driver.find_element_by_css_selector('.button-dark')
        except NoSuchElementException:
            try:
                self.driver.find_element_by_id('user-header-logout')
            except:
                print('Problem with link, could not find Login button - %s'%( link))
                return False
        # Workaround for ajax page loading issue
        time.sleep(self.wait_on_page_load)
        return True
        
    def get_html_source(self):
        return self.driver.page_source

    
    def populate_games_into_season(self, season):
        """
        Params:
            season (Season) with urls but not games populated, to modify
        """
        page_counter=0
        for url in season.urls:
            page_counter=page_counter+1
            self.go_to_link(url)
            html_source = self.get_html_source()
            html_querying = pyquery(html_source)
            # Check if the page says "No data available"
            no_data_div = html_querying.find('div.message-info > ul > li > div.cms')
            if no_data_div != None and no_data_div.text() == 'No data available':
                # Yes, found "No data available"
                print('Found "No data available", skipping %s'%(url))
                continue
            if season.name == 'current':
                tournament_table = html_querying.find('table#tournamentTable')
            else:
                tournament_table = html_querying.find('div#tournamentTable > table#tournamentTable')
            table_rows = tournament_table.find('tbody > tr')
            num_table_rows = len(table_rows)
            for i in range(0,num_table_rows):
                # Finding the table cell with game time and assessing if its blank tells us if this is a game data row
                time_cell = tournament_table.find('tbody > tr').eq(i).find('td.table-time')
                if  len(str(time_cell).strip())==0:
                    # This row of the table does not contain game/match data
                    continue
                
                time_cell = time_cell[0]
                for key, value in time_cell.attrib.items():
                    if key == 'class':
                        time_cell_classes = value.split(' ')
                        for time_cell_class in time_cell_classes:
                            if 0 == len(time_cell_class) or time_cell_class[0] != 't':
                                continue
                            if time_cell_class[1] == '0' or time_cell_class[1] == '1' or time_cell_class[2] == '2':
                                unix_time = int(time_cell_class.split('-')[0].replace('t',''))
                                game_datetime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(unix_time))
                                break
                        break
                
                # If time still isn't set at this point, then assume corrupt data and skip the row
                if len(game_datetime) == 0:
                    continue
                # Now get the table cell - the link within it, actually - with participants
                participants_link = tournament_table.find('tbody > tr').eq(i).find('td.table-participant > a')


                for plink in participants_link:
                    if "javascript:void(0)" not in plink.attrib['href']:
                        plink = plink.attrib['href']                        
                        break

                if "inplay-odds" in plink:
                    #trim
                    plink=plink[:-12]

                game_url = self.base_url + plink
                
                participants = participants_link.text().split(' - ')
                team_home = participants[0].strip()
                team_away = participants[1].strip()

                
                if season.name != 'current':

                    #Get score
                    overall_score_cell = tournament_table.find('tbody > tr').eq(i).find('td.table-score')
                    overall_score_string = overall_score_cell.text()
                    overall_score_string = overall_score_string.split()[0]
                    if ':' in overall_score_string:
                        score_home = int(overall_score_string.split(':')[0])
                        score_away = int(overall_score_string.split(':')[1])
                    elif '-' in overall_score_string:
                        score_home = int(overall_score_string.split('-')[0])
                        score_away = int(overall_score_string.split('-')[1])
                    else:
                        print('Could not split score string - delimiter unknown. Skipping game')
                        continue

                    

                    # Based on the score we can infer the outcome, as follows...
                    if score_home > score_away:
                        outcome = 2 #HOME
                    else:    
                        outcome = 1 #AWAY
                else:
                    score_home =0
                    score_away = 0
                    outcome = 0
                    


                    

                #visit the odd details URL
                self.go_to_link(game_url)
                html_source = self.get_html_source()
                html_querying = pyquery(html_source)
                # Check if the page says "No data available"
                no_data_div = html_querying.find('div.message-info > ul > li > div.cms')
                if no_data_div != None and no_data_div.text() == 'No data available':
                    # Yes, found "No data available"
                    print('Found "No data available", skipping %s'%(game_url))
                    continue
                
                #Get epxerts + odds
                experts_dict={}
                odds_table = html_querying.find("div#odds-data-table > div.table-container > table.table-main")
                odds_table_rows = odds_table.find('tbody > tr')
                num_odds_table_rows = len(odds_table_rows)
                
                print("\tProcessing games in page %i of %i:  \t%.2f%%\t" %(page_counter, len(season.urls), (i+1)/num_table_rows*100) , end="\r", flush=True)

                for j in range(num_odds_table_rows):
                
                    expert_name = odds_table.find('tbody > tr').eq(j).find('td>  div.l > a.name').text()

                    if len(expert_name)==0:
                        continue

                    odds = odds_table.find('tbody > tr').eq(j).find('td.right ').text().split()
                    #drop any odds that cannot be cast to float
                    for odd in odds:
                        try:
                            float(odd)
                        except:
                            odds.remove(odd)
                    
                    #add to dictionary but drop any odds that are 
                    if len(odds)==2:
                        experts_dict[expert_name]=[odds[1], odds[0]]  #reverse odds to AWAY, HOME


                #append game info to list
                game =[game_datetime, team_away, team_home, score_away, score_home,  outcome , experts_dict]
                season.add_game(game)





def writeOddsPortalSeason_ToExcel(this_season, data_file, away_only=True) :

    from_zone = tz.gettz('GMT')
    to_zone = tz.gettz('EST')


    if away_only:
        workbook = openpyxl.load_workbook(data_file)
        worksheet = workbook.worksheets[0]
        n_games=worksheet.max_row-2

        #grab file header
        col_head = []
        hindx=1
        while True:
            header= worksheet.cell(1, hindx).value
            if header is None:
                break
            col_head.append(header)
            hindx=hindx+1
            
    else:
        workbook = openpyxl.Workbook()
        worksheet = workbook.worksheets[0]
    


    if not away_only:
        #write header
        worksheet.cell(1, 1).value='Away'
        worksheet.cell(1, 2).value='Home'
        worksheet.cell(1, 3).value='Date'
        worksheet.cell(1, 4).value='Result'
        worksheet.cell(1, 5).value='Score Away'
        worksheet.cell(1, 6).value='Score Home'


    #loop over season captured games
    for i in range(len(this_season.games)):

        #convert from GMT to EST and save the date
        date = datetime.strptime(this_season.games[i][0], "%Y-%m-%d %H:%M:%S")
        date = date.replace(tzinfo=from_zone)
        date = date.astimezone(to_zone)
        date_string = date.strftime("%m/%d/%y")

        try:
            tA=TeamsList.getTeam_by_partial_Name(this_season.games[i][1]).short
        except:
            try:
                tA=TeamsList.getTeam_by_partial_ANY(this_season.games[i][1]).short
            except:
                tA=None

        try:
            tH=TeamsList.getTeam_by_partial_Name(this_season.games[i][2]).short
        except:
            try:
                tH=TeamsList.getTeam_by_partial_ANY(this_season.games[i][2]).short
            except:
                tH=None     


        if tA == None or tH == None:
            continue
        
        
        if away_only:                

            #loop over excel games
            for j in  range(n_games):
                #Get game info
                away_team_Short =           worksheet.cell(j+3, 1).value #cell indices start from 1
                home_team_Short =           worksheet.cell(j+3, 2).value #cell indices start from 1
                dt_string                  = worksheet.cell(j+3, 3).value  #cell indices start from 1

                #find a match between captured game and game in excel file
                if (away_team_Short==tA) and (home_team_Short==tH) and (dt_string == date_string):
                    #grab the dictionary keys (or the expert names)
                    for key in this_season.games[i][6].keys():
                        #check if key is in the columns header
                        try:
                            col_index = col_head.index(key)                        
                        except:
                            #expert does not exist already in header so add it
                            col_index=len(col_head)
                            col_head.append(key)
                            worksheet.cell(1, col_index+1).value = key


                        #covert the away odds to probs and add them under the appropriate column
                        away_odds = float(this_season.games[i][6][key][0] or 0)
                        if away_odds <=0:
                            away_probs = 0.5
                        else:
                            away_probs = 1/away_odds

                        worksheet.cell(j+3, col_index+1).value = away_probs

                        
                        
                    


        else:

            #write Date
            worksheet.cell(i+3,3).value = date_string
            #write AWAY team
            worksheet.cell(i+3,1).value =   TeamsList.getTeam_by_partial_Name(this_season.games[i][1]).short
            #write HOME team
            worksheet.cell(i+3,2).value =   TeamsList.getTeam_by_partial_Name(this_season.games[i][2]).short
            #write result
            worksheet.cell(i+3,4).value = this_season.games[i][5]
            #write scores
            worksheet.cell(i+3,5).value = this_season.games[i][3] #Away
            worksheet.cell(i+3,6).value = this_season.games[i][4] #Home
            
            index=8
            for expert_pred in this_season.games[i][6]:
                
                [odds_away, odds_home] = this_season.games[i][6][expert_pred]

                #write           
                worksheet.cell(i+3,index).value =   expert_pred+"::"+odds_away+","+odds_home
                index=index+1
        



    workbook.save(data_file)  
    



def Capture_OddsPortal_HistData(Seasons):

    wait_on_page_load = 3 # seconds - default wait time for each page to load completely
    main_league_results_url="https://www.oddsportal.com/basketball/usa/nba/results/"


    #Crawler for season links 
    crawler = Crawler(wait_on_page_load=wait_on_page_load)
    working_seasons = crawler.get_seasons_for_league(main_league_results_url)
    crawler.close_browser()


    #Now scrape games for each season of this league's history
    for this_season in working_seasons:


        if int(this_season.name[0:4])  in Seasons:

            crawler = Crawler(wait_on_page_load=wait_on_page_load)
            print('Season "%s" - getting all pagination links \t\t' %(this_season.name))
            crawler.fill_in_season_pagination_links(this_season)
            crawler.close_browser()

            scraper = Scraper(wait_on_page_load=wait_on_page_load)
            print('Season "%s" - populating all game data via pagination links \t\t' %(this_season.name))
            scraper.populate_games_into_season(this_season)
            scraper.close_browser()
        
            #dump to excel file
            out_file ="./Data/Experts_data/Oddsportal_"+this_season.name[0:4]+"-"+this_season.name[-2:]+".xlsx"
            writeOddsPortalSeason_ToExcel(this_season, out_file, away_only=False)
          
    
    print('done')
 

def Capture_OddsPortal_UpcomingData(out_file):

    wait_on_page_load = 3 # seconds - default wait time for each page to load completely
    main_url="https://www.oddsportal.com/basketball/usa/nba/"

    curr_season = Season("current")
    curr_season.add_url(main_url)


    scraper = Scraper(wait_on_page_load=wait_on_page_load)
    scraper.populate_games_into_season(curr_season)

    scraper.close_browser()

    #dump to excel file
    writeOddsPortalSeason_ToExcel(curr_season, out_file, away_only=True)




if __name__ == '__main__':
    
    Capture_OddsPortal_HistData([2020])

    #Capture_OddsPortal_UpcomingData("test_games_EXPERTS.xlsx")

