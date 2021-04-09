import numpy as np
import random
import os, sys, inspect

import joblib
import time
import openpyxl



sys.path.append('../')
sys.path.append('../Betting_strategies')

from runBacktest_modular import getTrueOdds

from Kelly_multibet_constrained_Shorting_Strategy import Kelly_multibet_constrained_Shorting_Strategy


from Pareto_Front_Strategy import ExpectationRisk
from multibet_utils import outcomes_HACK, PoissonBinomialPDF


def LoadTestData(filename):
    
    workbook = openpyxl.load_workbook(filename,data_only = True)
    worksheet = workbook.worksheets[0]
 
    no_games=worksheet.max_row-2

 
    AWAY_teams = []
    HOME_teams=[]
    DATE=[]
    bets=np.zeros(no_games)
    Probs=np.zeros([no_games,2])
    existingStakes=np.zeros(no_games)
    existing_bets=np.zeros(no_games)
    market_odds=np.zeros([no_games,2])
    market_bets=np.zeros(no_games)
    market_probs=np.zeros([no_games,2])

    for i in range(no_games):
        AWAY_teams.append(worksheet.cell(2+i+1, 0+1).value) #cell indices start from 1
        HOME_teams.append(worksheet.cell(2+i+1, 1+1).value) #cell indices start from 1
        DATE.append(worksheet.cell(2+i+1, 2+1).value)  #cell indices start from 1

        bets[i]= worksheet.cell(2+i+1, 7).value               #cell indices start from 1
        Probs[i,0]=worksheet.cell(2+i+1, 8).value/100    #cell indices start from 1
        Probs[i,1]=worksheet.cell(2+i+1, 9).value/100  #cell indices start from 1
        existingStakes[i]=worksheet.cell(2+i+1, 10).value  #cell indices start from 1
        existing_bets[i]=worksheet.cell(2+i+1, 11).value  #cell indices start from 1

        market_odds[i,0]=worksheet.cell(2+i+1, 12).value         #cell indices start from 1
        market_odds[i,1]=worksheet.cell(2+i+1, 13).value   #cell indices start from 1
        market_bets[i]=worksheet.cell(2+i+1, 14).value   #cell indices start from 1
        market_probs[i,0]=worksheet.cell(2+i+1, 15).value #cell indices start from 1
        market_probs[i,1]=worksheet.cell(2+i+1, 16).value  #cell indices start from 1


    return  AWAY_teams, HOME_teams, DATE, no_games, Probs, bets, existing_bets, market_odds, market_bets, market_probs, existingStakes
       


def main():

    #COMMON PARAMETERS
    comission_pcnt=0.02 # 2%  comission
    initial_balance= float(input("Enter current balance: "))  
    min_bet=0 #Minimum wager (e.g. Betfair exchange)
    max_bet=5000 #Maximum bet the market can take? (much higher for Sportsbook)
    Market_file='./test_games.xlsx'


    # LOAD UPCOMING GAME DATA
    AWAY_teams, HOME_teams, DATE, no_games, Probs, bets, existing_bets, market_odds, market_bets, market_probs, existingStakes   =  LoadTestData(Market_file)
   
   

    #Adjust odds given comission percentage
    true_odds=getTrueOdds(market_odds,comission_pcnt)




    #Betting STRATEGY: 
    alpha=0.9
    beta=0.05
    max_exposure=0.5 #percentage of balance for total bets
    max_bet_pcnt=0.1 #percentage of balance of the maximum single bet
    [Stakes,Bets]=Kelly_multibet_constrained_Shorting_Strategy(bets.astype(int), true_odds, Probs, alpha, beta, max_exposure, max_bet_pcnt, initial_balance, existingStakes)




    #Fix stakes. i.e. round, min, max etc         
    Stakes=np.round(Stakes,2)
    for i in range(len(Stakes)):
        if Stakes[i]<min_bet:
             Stakes[i]=0 #Assumes Exchange betting only
        if Stakes[i]>max_bet:
            Stakes[i]=max_bet #Assumes Exchange has some max bet amount 
        if Stakes[i]>initial_balance:
            Stakes[i]=initial_balance





    #ADD RESULTS TO EXCEL  and PRINT RESULTS TO SCREEN
    BET_workbook = openpyxl.load_workbook(Market_file,data_only = False)

    w_sheet = BET_workbook.worksheets[0]

    print("\nBet summary for ", DATE[0])

    
    for i in range(no_games):
        if Bets[i] == 1:
            print("\t Bet on AWAY TEAM (%s):, %.2f  with prob: %.2f" % (AWAY_teams[i], Stakes[i], Probs[i][0]))
        else:
            print("\t Bet on HOME TEAM (%s):, %.2f  with prob: %.2f" % (HOME_teams[i], Stakes[i], Probs[i][1]))

        
        w_sheet.cell(2+i+1, 10).value = Stakes[i] #cell indices start from 1
        w_sheet.cell(2+i+1, 11).value = int(Bets[i]) #cell indices start from 1
        
 



    #Calculate Expecation-Variance of calculated bet stakes
    outc=outcomes_HACK(no_games)
    k=no_games
    num_events=int(pow(2,no_games))
    Event_Probs=np.zeros(num_events)
    for i in range(num_events):
        individual_probs=np.diag(Probs[:,outc[i,:]-1])
        Event_Probs[i]= PoissonBinomialPDF(k,no_games,individual_probs)

    EV, _ = ExpectationRisk(Stakes,no_games,market_odds,Event_Probs,bets.astype(int),outc)
    if EV[0]==0:      
        ratio=0
    else:
        ratio = -EV[0]/EV[1]
    print("\nExpectation: ", -EV[0], ", Risk: ", EV[1], ", E/V ratio: ",  ratio )

    
    BET_workbook.save(Market_file) 


if __name__ == '__main__':
    main()    