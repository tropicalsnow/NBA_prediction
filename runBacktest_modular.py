#V2.0. Modular, uses simultaneous games and passes the same data to each strategy. Ported from MATLAB.
#V2.1  Meta-classifier (and associated features) have been removed since probabilities from test_games.xlsx & MarketData_20XX-20XX (may) already include meta-probs


import numpy as np
import random
import os, sys, inspect
import openpyxl

import joblib
import matplotlib.pyplot as plt
from operator import add

from datetime import datetime

sys.path.append('Libs')
sys.path.append('./Betting_strategies')



import multiprocessing 
import multiprocessing.pool



#import different betting strategies
from Fixed_Bet_Strategy import Fixed_Bet_Strategy
from Fixed_Percentage_Strategy import Fixed_Percentage_Strategy
from Kelly_Strategy import Kelly_Strategy
from Kelly_multibet_constrained_Strategy import Kelly_multibet_constrained_Strategy
from Expectation_based_Strategy import Expectation_based_Strategy
from Edge_based_Strategy import Edge_based_Strategy
from With_Market_Strategy import With_Market_Strategy
from Probability_threshold_Strategy import Probability_threshold_Strategy
from Edge_based_Strategy_with_threshold import Edge_based_Strategy_with_threshold
from Market_Predictions_Strategy import Market_Predictions_Strategy
from Fixed_Variance_Strategy import Fixed_Variance_Strategy
from Fixed_EVRatio_Strategy import Fixed_EVRatio_Strategy
from Pareto_Front_Strategy import Pareto_Front_Strategy
from Experimental_Strategy import Experimental_Strategy
from Kelly_shorting_Strategy import Kelly_shorting_Strategy
from Manual_Strategy import Manual_Strategy
from Supervised_Learning_Strategy import Supervised_Learning_Strategy
from Kelly_multibet_constrained_Shorting_Strategy import Kelly_multibet_constrained_Shorting_Strategy
from Fixed_Short_Strategy import Fixed_Short_Strategy
from Percentage_Short_Strategy import Percentage_Short_Strategy
from Reinforcement_Learning_Strategy import Reinforcement_Learning_Strategy


from strategy_Stats import calculate_Strategy_Stats, average_Strategy_Stats 
import customCostFunction



def StrategiesRun(Strategies_to_test,min_bet,max_bet,results,f,Fixed_bet_amount,FULL_our_probs, FULL_our_prediction, FULL_market_odds, FULL_market_prediction, \
    initial_balance, no_games, randomize,verbose,comission_pcnt,ax,ax2,saving):
    


    num_of_strategies=len(Strategies_to_test)
        
    strategy_balances=initial_balance*np.ones([num_of_strategies])
    strategy_saving_balances=np.zeros([num_of_strategies])

    running_balances=[]
    running_saving_balances=[]
    running_mean_balances=[]
    running_stakes=[]
    running_accuracies=[]
    running_bets_with_market=[]
    stats=[]
    name=[]
    loading_pcnt=[]
    for strats in range(num_of_strategies):
        running_balances.append([initial_balance])
        running_saving_balances.append([0])
        running_stakes.append([])
        loading_pcnt.append([])
        running_accuracies.append([0])
        running_bets_with_market.append([])
        running_mean_balances.append([initial_balance])
        stats.append([])
        name.append([])  

    N=np.r_[0:no_games]

    
    if not randomize:
        rng = np.random.default_rng(99999) #fix random number of games per day per run. But NOT the same number of games per day.  
    else:
        rng = np.random.default_rng() #new entropy
    random.seed(a=rng)
    
    
    while len(N)>0:    

            
        #System parameters
        
        #random number of simultaneous games. Based on historical examples
        mu=7.454
        sigma=2.751
        

        simultaneous_games=int(rng.normal(mu,sigma,1)[0])
        simultaneous_games = np.max([1,simultaneous_games])
        simultaneous_games = np.min([15,simultaneous_games])
 

        if len(N) > simultaneous_games:
            if randomize:
                #Draw random simultaneous games 
                K=random.sample(list(N),simultaneous_games)
                

            else:
                #Draw simultaneous games in sequence
                K=N[0:simultaneous_games]
                
            #remove those from history list
            N=np.setdiff1d(N,K)
            
        else:
            K=N
            N=[]
 
        if verbose:
            print("Remaining games: ",len(N))    
        

        #Extract the data for the sequence of bets
        Probs=FULL_our_probs[K] 
        bets=FULL_our_prediction[K].astype(int)
        market_odds=FULL_market_odds[K]
        market_probs=np.zeros(market_odds.shape)
        for mi in range(len(K)):
            if all(market_odds[mi]) > 0:
                market_probs[mi] = 1/market_odds[mi]
        
        market_bets=FULL_market_prediction[K]
      
 
        #Adjust odds given comission percentage. 
        true_market_odds=getTrueOdds(market_odds,comission_pcnt)
        #TODO: Calculate true_market_probs as well


        for strats in range(num_of_strategies):


            if Strategies_to_test[strats]==1:
                #STRATEGY 1. Equal (fixed) amounts to each event (NAIVE strategy)
                name[strats] = 'Fixed'
                (Stakes,Bets)=Fixed_Bet_Strategy(bets,Fixed_bet_amount)
            
            elif Strategies_to_test[strats]==2:
                #STRATEGY 2. Equal (percentile of balance) amounts to each event (NAIVE strategy)    
                name[strats] = 'Fixed percent'
                (Stakes,Bets)=Fixed_Percentage_Strategy(bets,f,strategy_balances[strats])

            elif Strategies_to_test[strats]==3:
                #STRATEGY 3. Fractional Kelly criterion betting strategy
                name[strats] = 'Kelly'
                kf=0.3
                cutoff=0.05 #max percentage of balance to bet on single bet 
                [Stakes,Bets]=Kelly_Strategy(bets,Probs,true_market_odds,kf,cutoff,strategy_balances[strats])
                        
            elif Strategies_to_test[strats]==4:
                #STRATEGY 4.  Expectation-based betting strategy
                name[strats] = 'Expectation based'
                [Stakes,Bets]=Expectation_based_Strategy(bets,Probs,true_market_odds,f,min_bet,strategy_balances[strats])

            elif Strategies_to_test[strats]==5:
                #STRATEGY 5. Edge-based  betting strategy
                name[strats] = 'Edge based'
                [Stakes,Bets]=Edge_based_Strategy(bets,Probs,market_probs,f,min_bet,strategy_balances[strats])

            elif Strategies_to_test[strats]==6:
                #STRATEGY 6. With Market fixed-betting strategy
                name[strats] = 'With Market'
                [Stakes,Bets]=With_Market_Strategy(bets,market_bets,Fixed_bet_amount)

            elif Strategies_to_test[strats]==7:
                #STRATEGY 7. Probability based certainty fixed-betting strategy    
                name[strats] = 'Probability threshold'
                prob_threshold=0.65
                [Stakes,Bets]=Probability_threshold_Strategy(bets,Fixed_bet_amount,Probs,prob_threshold)
    
            elif Strategies_to_test[strats]==8:    
                #STRATEGY 8. Edge-based  with threshold betting strategy           
                name[strats] = 'Edge-based with threshold'
                prob_threshold=0.85
                [Stakes,Bets]=Edge_based_Strategy_with_threshold(bets,Probs,market_probs,f,prob_threshold,min_bet,strategy_balances[strats])

            elif Strategies_to_test[strats]==9: 
                #STRATEGY 9. Bet using market predictions
                name[strats] = 'Market predictions'
                [Stakes,Bets]=Market_Predictions_Strategy(market_bets,Fixed_bet_amount)

            elif Strategies_to_test[strats]==10: 
                #STRATEGY 10. Bet using fixed Variance threshold
                name[strats] = 'Fixed variance'
                var_threshold=0.35
                [Stakes,Bets]=Fixed_Variance_Strategy(bets,Probs,true_market_odds,Fixed_bet_amount,var_threshold)

            elif Strategies_to_test[strats]==11:
                #STRATEGY 11. Bet using fixed  Exp/var ratio threshold
                name[strats] = 'Fixed EV ratio'
                ratio_threshold = 0.1
                [Stakes,Bets]=Fixed_EVRatio_Strategy(bets,Probs,true_market_odds,f,ratio_threshold,min_bet,strategy_balances[strats])

            elif Strategies_to_test[strats]==12:
                #STRATEGY 12. Always bet against the market (Fixed)
                name[strats] = 'Short fixed'
                [Stakes,Bets]=Fixed_Short_Strategy(market_bets,Fixed_bet_amount)

            elif Strategies_to_test[strats]==13:                
                #STRATEGY 13. Pareto front
                name[strats] = 'Pareto'
                Max_individual_bet=strategy_balances[strats] * 0.05 #no more than 5% max per bet
                Total_BetMax= Max_individual_bet * len(K)
                Min_individual_bet=0  
                [Stakes,Bets,_,_]=Pareto_Front_Strategy(Probs,true_market_odds,Total_BetMax,Max_individual_bet,Min_individual_bet)

            elif Strategies_to_test[strats]==14:
                #STRATEGY 14. Risk-constrained Kelly. Multi-bet version
                name[strats] = 'RC Kelly Multibet'
                alpha=0.8
                beta=0.05
                max_exposure=0.5 #percentage of balance for total bets
                max_bet_pcnt=0.1 #percentage of balance of the maximum single bet
                [Stakes,Bets]=Kelly_multibet_constrained_Strategy(bets, true_market_odds, Probs, alpha, beta, max_exposure, max_bet_pcnt, strategy_balances[strats])

            elif Strategies_to_test[strats]==15: 
                #STRATEGY 15. RC Kelly Multibet Shorting                
                name[strats] = 'RC Kelly Multibet Shorting'
                alpha=0.9
                beta=0.05
                max_exposure=0.5 #percentage of balance for total bets
                max_bet_pcnt=0.1 #percentage of balance of the maximum single bet
                [Stakes,Bets]=Kelly_multibet_constrained_Shorting_Strategy(bets, true_market_odds, Probs, alpha, beta, max_exposure, max_bet_pcnt, strategy_balances[strats])


            elif Strategies_to_test[strats]==16:
                #STRATEGY 16. Always bet against the market (percentile)
                name[strats] = 'Short percent'
                (Stakes,Bets)=Percentage_Short_Strategy(market_bets,f,strategy_balances[strats])


            elif Strategies_to_test[strats]==100: 
                #STRATEGY 100. Experimental                
                name[strats] = 'Experimental'
                alpha=0.9
                beta=0.05
                max_exposure=0.25 #percentage of balance for total bets
                max_bet_pcnt=0.1 #percentage of balance of the maximum single bet
                [Stakes,Bets]=Experimental_Strategy(bets, true_market_odds, Probs, alpha, beta, max_exposure, max_bet_pcnt, strategy_balances[strats])



            elif Strategies_to_test[strats]==101: 
                #STRATEGY 101. Kelly with shorting capabilities  
                name[strats]='Kelly shorting'
                kf=0.2
                cutoff=0.05 #max percentage of balance to bet on single bet 
                [Stakes,Bets]=Kelly_shorting_Strategy(Probs,true_market_odds,kf,cutoff,strategy_balances[strats])
              
            elif Strategies_to_test[strats]==102:
                #STRATEGY 102. Supervised learning using a regressor  
                name[strats]='Supervised Learning'
                [Stakes,Bets]=Supervised_Learning_Strategy(Probs, true_market_odds,  strategy_balances[strats], np.mean(np.diff(running_balances[strats])>0),  len(N), strategy_balances[strats]-initial_balance)
                

            elif Strategies_to_test[strats]==103:
                #STRATEGY 103. Reinforcement Learning
                name[strats]='Reinforcement Learning'
                Reinforcement_Learning_Strategy(Probs, true_market_odds, initial_balance,  strategy_balances[strats])


            elif Strategies_to_test[strats]==999: 
                #STRATEGY 999. Manual 
                name[strats]='Manual'
                [Stakes,Bets]=Manual_Strategy(Probs, true_market_odds, strategy_balances[strats], np.mean(np.diff(running_balances[strats])>0),  len(N), strategy_balances[strats]-initial_balance)
    
            
            else:
                break



           #Fix stakes. i.e. round, min, max etc         
            #Stakes=np.round(Stakes) #Assumes Exchange betting only
            for i in range(len(Stakes)):
                if Stakes[i]<min_bet:
                    Stakes[i]=0 
                if Stakes[i]>max_bet:
                    Stakes[i]=max_bet #Assumes Exchange has some max bet amount 
                if Stakes[i]>strategy_balances[strats]:
                    Stakes[i]=strategy_balances[strats]


                      
            #store balance loading info
            loading_pcnt[strats].append(sum(Stakes)/(strategy_balances[strats]-sum(Stakes)) )

            #Evaluate bet WIN/LOSS
            for i in range(len(K)):
                #place bets
                bet=Bets[i]
                Bet_stake=Stakes[i]
                profit=0

                #Calculate expected value. We only calculate this for the plot
                p=Probs[i,bet-1]
                q=1-p
                d=market_odds[i,bet-1]
            


                if Bet_stake<=strategy_balances[strats]:
                    if Bet_stake != 0 and all(market_odds[i]) > 0: #place the bet or skip the bet (i.e.=0) altogether
                        if bet==results[K[i]]:
                            #win
                            profit =  Bet_stake*(market_odds[i,bet-1]-1)
                            comission= np.round(profit*comission_pcnt,2)
                            profit=profit-comission
 
 
                        else:
                            #lose
                            profit=-Bet_stake

 

                        strategy_balances[strats] =strategy_balances[strats] + profit

                        #We only update the balance when we have a bet. Since updating stationary balance when no bets are placed will add to the linear regression number of points
                        running_balances[strats].append(strategy_balances[strats])
                        running_mean_balances[strats].append(np.mean(running_balances[strats]))
                        running_stakes[strats].append(Bet_stake)
                        running_accuracies[strats].append(np.mean(np.diff(running_balances[strats])>0))
                        running_saving_balances[strats].append(running_saving_balances[strats][-1])



                        if bet==FULL_market_prediction[K[i]]:
                            running_bets_with_market[strats].append(1)
                        else:
                            running_bets_with_market[strats].append(0)

 
                else:
                    break


                if verbose:
                    print(i+1, "of ", len(K), "| Odds: ",market_odds[i,:],",  Bet placed: ",bet, ", Stake: ",Bet_stake,", Result: ", results[K[i]],", (P/L): ", profit,\
                    " (", strategy_balances[strats]-initial_balance, "), Resulting balance: ", strategy_balances[strats],", Accuracy: ",np.mean(np.diff(running_balances[strats])>0))  


 

                
                if strategy_balances[strats]<min_bet:
                    break


            #SAVING scheme. After the trading day
            if strategy_balances[strats] > initial_balance:
                amount =  (strategy_balances[strats]-initial_balance)*saving
                
                strategy_saving_balances[strats] = strategy_saving_balances[strats]+amount
                running_saving_balances[strats][-1] = strategy_saving_balances[strats]

                strategy_balances[strats] = strategy_balances[strats] - amount
                running_balances[strats][-1] = strategy_balances[strats]

         


            if verbose:
            
                ax.set_title("Strategy: "+name[strats])
                ax.plot(running_balances[strats], 'b') #print running balance

                ax2.set_ylabel('Savings balance')  
                ax2.plot(  running_saving_balances[strats], 'k--') # print running savings balance
                #ax2.plot(  running_accuracies[strats], 'k--') #print running Accuracy
                #ax2.plot(running_mean_balances[strats], 'k--') #print running mean balance
                
                ax.plot(list( map(add, running_balances[strats], running_saving_balances[strats]) ), 'r--') #print running balance + running savings balance


                ax.grid(True)
                plt.pause(0.05)

            
                print("\n")

    if verbose:
        plt.show()
 
    #ALL GAMES ARE FINISHED 

    #append any savings balance to (the end of) the final balance for all strategies  & calculate stats               
    for strats in range(num_of_strategies):
        strategy_balances[strats]  = strategy_balances[strats]  + strategy_saving_balances[strats]
        running_balances[strats][-1] = strategy_balances[strats]


        rB=running_balances[strats]
        rS= running_stakes[strats]
        rBWM= running_bets_with_market[strats]
        savings_balance = strategy_saving_balances[strats]
        strategy_saving_balances[strats] = 0 #not really necessary but reset the savings balances
        
        stats[strats].append(calculate_Strategy_Stats(rB,rS,min_bet,name[strats],rBWM,no_games,savings_balance, loading_pcnt))

        


    return stats


def LoadBackTestData(Market_history_file):
    #Load market data & additional ODDSHARK data



    BET_workbook = openpyxl.load_workbook(Market_history_file, data_only = True)
    BET_worksheet = BET_workbook.worksheets[0]
   
    no_games=BET_worksheet.max_row-2

    results=[] 
    FULL_our_prediction=np.zeros(no_games)
    FULL_our_probs=np.zeros([no_games,2])
    FULL_market_prediction=np.zeros(no_games)
    FULL_market_odds=np.zeros([no_games,2])


    CARMELO=[] #Carmelo probs for away
    COVERS=np.zeros([no_games,4]) #Covers consensus, sides, picks AWAY and picks HOME
    ODDSHARK=np.zeros([no_games,3]) #ODDSHARK moneyline away, home and spreads
    H2H=np.zeros([no_games,6]) #H2H record, score, fgp, rebounds, 3pp, steals

    
    ODDSHARK_LastN_Away=np.zeros([no_games,7]) #The extra ODDSHARK data from the Last N games for the AWAY team
    ODDSHARK_LastN_Home=np.zeros([no_games,7]) #The extra ODDSHARK data from the Last N games for the HOME team



    for i in range(no_games):
        
        
        results.append(int(BET_worksheet.cell(i+2+1, 3+1).value))  #cell indices start from 1
        FULL_our_prediction[i]=int(BET_worksheet.cell(i+2+1, 4+1).value) #cell indices start from 1


        if BET_worksheet.cell(i+2+1, 8+1).value == "": #cell indices start from 1
            FULL_market_prediction[i]=0
            FULL_market_odds[i,0]=0
            FULL_market_odds[i,1]=0
        else:
            FULL_market_prediction[i]=float(BET_worksheet.cell(i+2+1, 8+1).value)  #cell indices start from 1
            FULL_market_odds[i,0]=float(BET_worksheet.cell(i+2+1, 6+1).value)   #cell indices start from 1
            FULL_market_odds[i,1]=float(BET_worksheet.cell(i+2+1, 7+1).value) #cell indices start from 1

        
        if FULL_our_prediction[i]==1:
            FULL_our_probs[i,0]=float(BET_worksheet.cell(i+2+1, 5+1).value)    #cell indices start from 1
            FULL_our_probs[i,1]=100-float(BET_worksheet.cell(i+2+1, 5+1).value)  #cell indices start from 1
        else:
            FULL_our_probs[i,1]=float(BET_worksheet.cell(i+2+1, 5+1).value)       #cell indices start from 1
            FULL_our_probs[i,0]=100-float(BET_worksheet.cell(i+2+1, 5+1).value)   #cell indices start from 1
    

        #CARMELO prob. Convert to AWAY team
        carmelo_pred=int(BET_worksheet.cell(i+2+1, 11+1).value)          #cell indices start from 1                                    
        if carmelo_pred==1:
            CARMELO.append(np.float32(BET_worksheet.cell(i+2+1, 12+1).value))   #cell indices start from 1
        else:
            CARMELO.append(100.0-np.float32(BET_worksheet.cell(i+2+1, 12+1).value))   #cell indices start from 1


        #COVERS data
        if BET_worksheet.cell(i+2+1, 13+1).value == "":          #cell indices start from 1     
            COVERS[i,0]=0   #Consensus
        else:
            COVERS[i,0]=np.float32(BET_worksheet.cell(i+2+1, 13+1).value) #Consensus. AWAY TEAM only since the HOME TEAM is (100-Away).  Cell indices start from 1        
        
        if BET_worksheet.cell(i+2+1, 15+1).value == "":      #cell indices start from 1  
            COVERS[i,1]=0   #Sides
        else:
            COVERS[i,1]=np.float32(BET_worksheet.cell(i+2+1, 15+1).value) #Sides. AWAY TEAM. Cell indices start from 1
        
        if BET_worksheet.cell(i+2+1, 17+1).value == "":         #cell indices start from 1       
            COVERS[i,2]=0   #Picks away
            COVERS[i,3]=0   #Picks home
        else:
            COVERS[i,2]=int(BET_worksheet.cell(i+2+1, 17+1).value or 0)  #Picks AWAY TEAM. Cell indices start from 1
            COVERS[i,3]=int(BET_worksheet.cell(i+2+1, 18+1).value or 0) #Picks HOME TEAM.  Cell indices start from 1       

        #ODDSHARK data
        ODDSHARK[i,0]=np.float32(BET_worksheet.cell(i+2+1, 19+1).value) #Moneyline AWAY TEAM. Cell indices start from 1
        ODDSHARK[i,1]=np.float32(BET_worksheet.cell(i+2+1, 20+1).value) #Moneyline HOME TEAM.  Cell indices start from 1  
        ODDSHARK[i,2]=np.float32(BET_worksheet.cell(i+2+1, 21+1).value) #Spreads AWAY TEAM. Cell indices start from 1


        #H2H data
        h2hr_A=int(BET_worksheet.cell(i+2+1, 24+1).value)  #Cell indices start from 1
        h2hr_H=int(BET_worksheet.cell(i+2+1, 25+1).value)  #Cell indices start from 1
        TotalGames=h2hr_A+h2hr_H

        H2H[i,0]= (h2hr_A-h2hr_H)/TotalGames                                                                     #Record 
        H2H[i,1]=np.float32(BET_worksheet.cell(i+2+1, 26+1).value)-  np.float32(BET_worksheet.cell(i+2+1, 27+1).value)   #Score. Cell indices start from 1
        H2H[i,2]=np.float32(BET_worksheet.cell(i+2+1, 28+1).value) - np.float32(BET_worksheet.cell(i+2+1, 29+1).value)   #FGP. Cell indices start from 1
        H2H[i,3]=np.float32(BET_worksheet.cell(i+2+1, 30+1).value) -np.float32(BET_worksheet.cell(i+2+1, 31+1).value)    #Rebounds. Cell indices start from 1
        H2H[i,4]=np.float32(BET_worksheet.cell(i+2+1, 32+1).value) - np.float32(BET_worksheet.cell(i+2+1, 33+1).value)   #3PP. Cell indices start from 1
        H2H[i,5]=np.float32(BET_worksheet.cell(i+2+1, 34+1).value) - np.float32(BET_worksheet.cell(i+2+1, 35+1).value)   #Steals. Cell indices start from 1
        


        #ODDSHARK LastN data
        ODDSHARK_LastN_Away[i,0]=np.float32(BET_worksheet.cell(i+2+1, 37+1).value)  #AWAY LastN Win pct. Cell indices start from 1
        ODDSHARK_LastN_Away[i,1]=np.float32(BET_worksheet.cell(i+2+1, 38+1).value)  #AWAY LastN Score A. Cell indices start from 1
        ODDSHARK_LastN_Away[i,2]=np.float32(BET_worksheet.cell(i+2+1, 39+1).value)  #AWAY LastN Score H. Cell indices start from 1
        ODDSHARK_LastN_Away[i,3]=np.float32(BET_worksheet.cell(i+2+1, 40+1).value)  #AWAY LastN Line. Cell indices start from 1
        ODDSHARK_LastN_Away[i,4]=np.float32(BET_worksheet.cell(i+2+1, 41+1).value)  #AWAY LastN FG pct. Cell indices start from 1
        ODDSHARK_LastN_Away[i,5]=np.float32(BET_worksheet.cell(i+2+1, 42+1).value)  #AWAY LastN FT pct. Cell indices start from 1
        ODDSHARK_LastN_Away[i,6]=np.float32(BET_worksheet.cell(i+2+1, 43+1).value)  #AWAY LastN 3PTM pct. Cell indices start from 1

        ODDSHARK_LastN_Home[i,0]=np.float32(BET_worksheet.cell(i+2+1, 45+1).value)  #HOME LastN Win pct. Cell indices start from 1
        ODDSHARK_LastN_Home[i,1]=np.float32(BET_worksheet.cell(i+2+1, 46+1).value)  #HOME LastN Score A. Cell indices start from 1
        ODDSHARK_LastN_Home[i,2]=np.float32(BET_worksheet.cell(i+2+1, 47+1).value)  #HOME LastN Score H. Cell indices start from 1
        ODDSHARK_LastN_Home[i,3]=np.float32(BET_worksheet.cell(i+2+1, 48+1).value)  #HOME LastN Line. Cell indices start from 1
        ODDSHARK_LastN_Home[i,4]=np.float32(BET_worksheet.cell(i+2+1, 49+1).value)  #HOME LastN FG pct. Cell indices start from 1
        ODDSHARK_LastN_Home[i,5]=np.float32(BET_worksheet.cell(i+2+1, 50+1).value)  #HOME LastN FT pct. Cell indices start from 1
        ODDSHARK_LastN_Home[i,6]=np.float32(BET_worksheet.cell(i+2+1, 51+1).value)  #HOME LastN 3PTM pct. Cell indices start from 1


    FULL_our_probs=FULL_our_probs/100


    return no_games, results, FULL_our_probs, FULL_our_prediction, FULL_market_odds, FULL_market_prediction, CARMELO,  COVERS, ODDSHARK, H2H, ODDSHARK_LastN_Away, ODDSHARK_LastN_Home

	
def LoadAdditionalExpertsData(Experts_history_file,Market_history_file):

    from TeamsList import getTeam_by_partial_ANY, getTeam_by_Short

    #Load also the Market data so we align the games
    Market_workbook = openpyxl.load_workbook(Market_history_file)
    Market_worksheet = Market_workbook.worksheets[0]

    #Load the experts data
    Experts_workbook = openpyxl.load_workbook(Experts_history_file)
    Experts_worksheet = Experts_workbook.worksheets[0]

    #Note: Experts might be incomplete or include non-regular season so need to be crosschecked with Market data
    no_games_EX=Experts_worksheet.max_row-1  #Experts  
    no_games_MR= Market_worksheet.max_row-2 #Market
    

    line =np.zeros(no_games_MR)
    lineavg =np.zeros(no_games_MR)
    linesag =np.zeros(no_games_MR)
    linesage=np.zeros(no_games_MR)
    linesagp=np.zeros(no_games_MR)
    lineopen=np.zeros(no_games_MR)
    linemoore=np.zeros(no_games_MR)
    linepower=np.zeros(no_games_MR)   
    linesaggm=np.zeros(no_games_MR)
    linefox=np.zeros(no_games_MR)
    linedok=np.zeros(no_games_MR)
    linetalis=np.zeros(no_games_MR)

    linemassey=np.zeros(no_games_MR)	
    linepugh=np.zeros(no_games_MR)	
    linedonc=np.zeros(no_games_MR)
 

    for j in range(no_games_MR):  #Loop over Market games.   Assumes games order is the same for MarketData file so we can combine with standard features


        #Get teams, date and result
        teamA_MR= getTeam_by_Short(Market_worksheet.cell(j+2+1, 0+1).value)  #Away team, Market. cell indices start from 1
        teamH_MR= getTeam_by_Short(Market_worksheet.cell(j+2+1, 1+1).value)  #Home team, Market. cell indices start from 1


        date_MR = Market_worksheet.cell(j+2+1, 2+1).value  #Date, Market. cell indices start from 1
        dt_MR =  datetime.strptime(str(date_MR), '%m/%d/%y').date() #Convert to Python datetime object
 

        for i in range(no_games_EX):

            #Get teams, date and score
            teamA_EX=getTeam_by_partial_ANY(Experts_worksheet.cell(i+1+1, 3+1).value)  #Away team, Experts. cell indices start from 1
            teamH_EX=getTeam_by_partial_ANY(Experts_worksheet.cell(i+1+1, 1+1).value) #Home team, Experts. cell indices start from 1
       
            date_EX = int(Experts_worksheet.cell(i+1+1, 0+1).value)  #Date, Experts. cell indices start from 1
            dt_EX = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + date_EX - 2).date() #Convert to Python datetime object


            #Verify this games data with expert data. We can only use team names and date since scores are unreliable (i.e. missing from experts file)
            if teamA_EX == teamA_MR and teamH_EX == teamH_MR and  dt_EX == dt_MR:
            #Match found
                 
                #Append fatures or 0 if empty   
                line[j]= float(Experts_worksheet.cell(i+1+1, 5+1).value or 0)  # or float('NaN')). cell indices start from 1
                lineavg[j]= float(Experts_worksheet.cell(i+1+1, 6+1).value or 0)  # or float('NaN')). cell indices start from 1
                linesag[j]= float(Experts_worksheet.cell(i+1+1, 7+1).value or 0)  # or float('NaN')). cell indices start from 1
                linesage[j]= float(Experts_worksheet.cell(i+1+1, 8+1).value or 0)  # or float('NaN')). cell indices start from 1

                linesagp[j]=float(Experts_worksheet.cell(i+1+1, 9+1).value or 0)  # or float('NaN')). cell indices start from 1
                lineopen[j]=float(Experts_worksheet.cell(i+1+1, 10+1).value or 0)  # or float('NaN')). cell indices start from 1
                linemoore[j]=float(Experts_worksheet.cell(i+1+1, 11+1).value or 0)  # or float('NaN')). cell indices start from 1
                linepower[j]=float(Experts_worksheet.cell(i+1+1, 12+1).value or 0)  # or float('NaN')). cell indices start from 1

                linesaggm[j]=float(Experts_worksheet.cell(i+1+1, 13+1).value or 0)  # or float('NaN')). cell indices start from 1
                linefox[j]=float(Experts_worksheet.cell(i+1+1, 15+1).value or 0)  # or float('NaN')). cell indices start from 1
                linedok[j]=float(Experts_worksheet.cell(i+1+1, 16+1).value or 0)  # or float('NaN')). cell indices start from 1
                linetalis[j]=float(Experts_worksheet.cell(i+1+1, 17+1).value or 0)  # or float('NaN')). cell indices start from 1

                linemassey[j]=float(Experts_worksheet.cell(i+1+1, 18+1).value or 0)  # or float('NaN')). cell indices start from 1
                linepugh[j]=float(Experts_worksheet.cell(i+1+1, 19+1).value or 0)  # or float('NaN')). cell indices start from 1
                linedonc[j]=float(Experts_worksheet.cell(i+1+1, 20+1).value or 0)  # or float('NaN')). cell indices start from 1

                break

    

    return line, lineavg, linesag, linesage, linesagp, lineopen, linemoore, linepower, linesaggm, linefox, linedok, linetalis, linemassey, linepugh, linedonc

 
def getTrueOdds(FULL_market_odds,comission_pcnt):

    true_odds = np.zeros(FULL_market_odds.shape)
    for i in range(len(FULL_market_odds)):
    #Assume both sides are possible BACK odds, since we do not LAY binary bets
        true_odds[i,0]= 1+(1-comission_pcnt)*(FULL_market_odds[i,0] -1)
        true_odds[i,1]= 1+(1-comission_pcnt)*(FULL_market_odds[i,1] -1)
    return true_odds



def main():

    # SYSTEM VARIABLES
    Strategies_to_test=[100]
    strategy_runs=100
    randomize=True
    verbose=True
    Market_history_file='./Data/Backtest_data/BacktestData_2020-21.xlsx'



    comission_pcnt=0.02 # 2% Betfair comission
    initial_balance=1100
    min_bet=0 #Minimum wager (e.g. Betfair exchange)
    max_bet=500 #Maximum bet the market can take? (much higher for Sportsbook)
    f=0.1 #percentile of balance to bet
    Fixed_bet_amount=round(initial_balance*0.025)  #for all fixed bet strategies
    saving = 0.0  #save percentagex100 of balance above initial balance

 
    # LOAD BACKTEST EXCEL DATA
    no_games, results, FULL_our_probs, FULL_our_prediction, FULL_market_odds,\
     FULL_market_prediction, CARMELO,  COVERS, ODDSHARK, H2H,  ODDSHARK_LastN_Away, ODDSHARK_LastN_Home = LoadBackTestData(Market_history_file)

 

    # STRATEGIES EVALUATION
    #variables for each strategy
    num_of_strategies=len(Strategies_to_test)
    StratStats=[]
    running_stats=[]
    for strats in range(num_of_strategies):
        StratStats.append([])
        running_stats.append([])



    #run strategies multiple times
    if randomize is False:
        strategy_runs=1
    if strategy_runs>1 or num_of_strategies>1:
        verbose=False

    if verbose:

        plt.figure()
        ax = plt.axes()
        ax2 = ax.twinx()  # instantiate a second axes that shares the same x-axis
        ax.set_xlabel('Bets')
        ax.set_ylabel('Balance', color='blue')
        ax.tick_params(axis='y', labelcolor='blue')

    else:
        ax=None
        ax2=None
        

    


    #Single processing. Only for manual entries or for single runs 
    if  strategy_runs==1 : 
        running_stats= StrategiesRun(Strategies_to_test,min_bet,max_bet,results,f,Fixed_bet_amount, \
                               FULL_our_probs, FULL_our_prediction, FULL_market_odds, FULL_market_prediction,  \
                               initial_balance, no_games, randomize,verbose,comission_pcnt,ax,ax2, saving)


                            
    else:

        #Multiprocessing        
        cpus=12
        pool = multiprocessing.Pool(processes=cpus)
        results = [pool.apply_async(StrategiesRun, args=(Strategies_to_test,min_bet,max_bet,results,f,Fixed_bet_amount, \
                                FULL_our_probs, FULL_our_prediction, FULL_market_odds, FULL_market_prediction, \
                                initial_balance, no_games, randomize,verbose,comission_pcnt,ax,ax2, saving)) for i in range(strategy_runs)]
        pool.close()
        pool.join()
    


        #Gather the results
        for p in results:
            for strats in range(num_of_strategies):
                running_stats[strats].append(p.get()[strats][0])



    #average stats over runs
    for strats in range(num_of_strategies):
        StratStats[strats]=average_Strategy_Stats(running_stats[strats])





    #SAVE TO EXCEL
    workbook = openpyxl.Workbook()
    worksheet = workbook.worksheets[0]

    fields=dir(StratStats[0])
    for strats in range(num_of_strategies):  
        field_count=1

        worksheet.cell(0+1, strats+1+1).value= str(StratStats[strats].StrategyName) #Strategy name (Header). Cell indices start from 1

        for i in range(1, len(fields)):
            if "__" not in fields[i]: #skip over internal fields of the struct

                worksheet.cell(field_count+1, 0+1).value= str(fields[i]) #Field name (Header). Cell indices start from 1
                exec("worksheet.cell(field_count+1, strats+1+1).value=     StratStats[strats]."+fields[i]) #Field data. Cell indices start from 1
                field_count=field_count+1

 
  
    workbook.save("./Data/Backtest_data/Backtest_simulations.xlsx")  



 
  

if __name__ == '__main__':
    main()    