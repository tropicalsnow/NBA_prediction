def captureCarmeloData(game_file, date_string, starting_column=13):

    import requests
    from bs4 import BeautifulSoup, NavigableString, Tag



    import openpyxl
    
    from datetime import datetime

    from TeamsList import getTeam_by_partial_Name 


    #Parse game file
    GAME_workbook = openpyxl.load_workbook(game_file)
    GAME_worksheet = GAME_workbook.worksheets[0]

    #Check if header already exists and write header
    col_head = []
    hindx=1
    while True:
        header= GAME_worksheet.cell(1, hindx).value
        if header is None:
            break
        col_head.append(header)
        hindx=hindx+1
    #check if expert name is in the columns header
    try:
        starting_column = col_head.index("538_RAPTOR" )+1                        
    except:
        #expert does not exist already in header so add it
        starting_column=len(col_head)+1
 
    GAME_worksheet.cell(1, starting_column).value= "538_RAPTOR" 
    games_num=GAME_worksheet.max_row-2


    #Convert to python datetime object  = datetime.strptime(df['Date'][date_block_index],'%m/%d/%y') 
    datetime_object  = datetime.strptime(date_string,'%m/%d/%y') 


    #Navigate to main 538 nba  page
    url='https://projects.fivethirtyeight.com/2021-nba-predictions/games/'
    response = requests.get(url)
    HTMLtext = response.text.replace("<br />"," ")  #Remove line breaks and replace them with space
     
    soupify = BeautifulSoup(HTMLtext, 'lxml')  # Parse the HTML as a string.

    
    #Get all upcoming games of the day
    soup_g= soupify.find_all('div', {'class': 'day-group'})
    
    CARMELO_teams=[]
    CARMELO_Probs=[]

    for game in soup_g:
        #check date
        section=game.find_all('section', {'class': 'day'})
        for matchup in section:
            date=matchup.find('h3',{'class': 'h3'})
            date_text=date.text.split(", ")[1].split(" ")
            month=date_text[0][0:3]
            date=date_text[1]

            #Consider only specific day's games
            if (month == datetime_object.strftime("%b")) and (int(date) == int(datetime_object.day)):

                all_matches = matchup.find('div',{'class': 'games-section'})    
                for match in all_matches:
                    teams=match.find_all('tr',{'class': 'tr team'})

                    AWAY=teams[0].text.split(" ")[0]
                    # AWAY_probs=int(teams[0].text.split(" ")[-1][-3:-1])
                    AWAY_probs=int(teams[0].text.split(" ")[-1].split('%')[0][-2:])

                    HOME=teams[1].text.split(" ")[0]
                    # HOME_probs=int(teams[1].text.split(" ")[-1][-3:-1])
                    HOME_probs=int(teams[1].text.split(" ")[-1].split('%')[0][-2:])


                    CARMELO_teams.append([AWAY, HOME])
                    CARMELO_Probs.append([AWAY_probs, HOME_probs])
                    
                break



    #loop for games in excel file
    for i in range(games_num):

        #Get game info
        away_team_Short =   GAME_worksheet.cell(i+2+1, 0+1).value #cell indices start from 1
        home_team_Short =   GAME_worksheet.cell(i+2+1, 1+1).value #cell indices start from 1
        
        #loop for games from carmelo list    
        for j in range(len(CARMELO_teams)):
            
            Carmelo_Away = CARMELO_teams[j][0] 
            Carmelo_Home=  CARMELO_teams[j][1]
            
            #Convert to short names
            Carmelo_Away_Short = getTeam_by_partial_Name(Carmelo_Away).short
            Carmelo_Home_Short = getTeam_by_partial_Name(Carmelo_Home).short


            Carmelo_Away_Prob= CARMELO_Probs[j][0]
            Carmelo_Home_Prob= CARMELO_Probs[j][1]

            #Check that teams match in the current game
            if away_team_Short == Carmelo_Away_Short and home_team_Short == Carmelo_Home_Short:
                

                print(away_team_Short, " - ", home_team_Short)    

                #Write only the AWAY probabilities
                GAME_worksheet.cell(i+3, starting_column).value= Carmelo_Away_Prob/100  #Cell indices start from 1




    GAME_workbook.save(game_file)  
    
