import xlrd
from copy import deepcopy
import os.path

import sys


import openpyxl

import numpy as np

def vdir(obj):
    return [x for x in dir(obj) if not x.startswith('__') and x.isupper()]


class structtype():
    pass


def SaveTrainingData_To_Excel(features,labels,filename,weight=None):
    #TODO: NEEDS TO BE CONVERTED TO OPENPYXL

    if weight is None:
        weight=np.zeros_like(labels)

    #Create a workbook and add a worksheet.
    workbook = xlsxwriter.Workbook(filename)
    worksheet = workbook.add_worksheet()

    if type(features).__name__ == 'ndarray':
        observations=features.shape[0]
        dims=features.shape[1]
    else:
        observations=len(features)
        dims=len(features[0])


    
    
    
    for i in range(observations):
        for j in range(dims):
            worksheet.write(0, j, "f"+str(j)) #Data Header           
            worksheet.write(i+2, j, features[i][j]) #Data
        
        worksheet.write(0,   dims, "Y") #Label Header
        worksheet.write(i+2, dims, labels[i]) #Label

        worksheet.write(0,   dims+1, "Weight") # Weight Header
        worksheet.write(i+2, dims+1, weight[i]) #Data weight

    workbook.close()

 

def LoadGameData(filename,testflag):
    print("Loading Games from file: "+filename)
    
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.worksheets[0]
    games_sz=worksheet.max_row-2
    
       

    GamesData = [ structtype() for i in range(games_sz)]
    
    
    for i in range(games_sz):
        GamesData[i].Away = worksheet.cell(2+i+1, 0+1).value #cell indices start from 1
        GamesData[i].Home= worksheet.cell(2+i+1, 1+1).value  #cell indices start from 1
        GamesData[i].Date= worksheet.cell(2+i+1, 2+1).value #cell indices start from 1
        
        if not testflag: #old code??
            GamesData[i].Pts_Away= int(worksheet.cell(2+i+1, 2+1).value) #cell indices start from 1
            GamesData[i].Pts_Home= int(worksheet.cell(2+i+1, 3+1).value) #cell indices start from 1
            GamesData[i].Pts_Diff= int(worksheet.cell(2+i+1, 4+1).value) #cell indices start from 1
            GamesData[i].Win= int(worksheet.cell(2+i+1, 5+1).value) #cell indices start from 1
            GamesData[i].Date= worksheet.cell(2+i+1, 6+1).value #cell indices start from 1
        
        
        
    return (GamesData)


def LoadInjuryList(filename, offset=0): 
    #print("Loading injuries from file: "+filename)
    
    if  os.path.isfile(filename):

        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook.worksheets[0]
        games_sz=worksheet.max_row-2

    

    
        InjuryList = [ structtype() for i in range(games_sz)]
    
        for i in range(games_sz):
            InjuryList[i].Away = worksheet.cell(2+i+1, 0+1).value #cell indices start from 1
            InjuryList[i].Home= worksheet.cell(2+i+1, 1+1).value #cell indices start from 1
  
            InjuryList[i].Date= worksheet.cell(2+i+1, 2+offset+1).value #cell indices start from 1
        
            InjuryList[i].AwayList= worksheet.cell(2+i+1, 3+offset+1).value #cell indices start from 1
            InjuryList[i].HomeList= worksheet.cell(2+i+1, 4+offset+1).value #cell indices start from 1
    
        return (InjuryList)
    
    return (None)
 

def writePredictionsToFile(filename,p_test,p_probs):
    
    
    
    #Check that the test file exists
    if  not os.path.isfile(filename):
        print("test file "+filename+" does not exist.")
        return 0
    else:
        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook.worksheets[0]
        games_sz=worksheet.max_row-2

 
        

        #Test that the number of games equals the number of predictions
        if games_sz==len(p_test):
            for i in range(games_sz):

                    worksheet.cell(3+i, 4).value= int(p_test[i]) #Cell indices start from 1
                    worksheet.cell(3+i, 5).value=  float(p_probs[i,0]*100) #Cell indices start from 1
                    worksheet.cell(3+i, 6).value=  float(p_probs[i,1]*100) #Cell indices start from 1

                    #Also copy on the META predictions columns. In case we do not use META ensembles
                    worksheet.cell(3+i, 7).value= int(p_test[i]) #Cell indices start from 1
                    worksheet.cell(3+i, 8).value=  float(p_probs[i,0]*100) #Cell indices start from 1
                    worksheet.cell(3+i, 9).value=  float(p_probs[i,1]*100) #Cell indices start from 1


            workbook.save(filename) 
            return 1                  
        else:
            print("Number of predictions ("+str(len(p_test))+") and number of games ("+games_sz+") do not match")
            return 0

        
        