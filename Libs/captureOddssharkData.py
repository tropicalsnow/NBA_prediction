import time


import openpyxl

import numpy as np
import math

from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

from bs4 import BeautifulSoup as soup

from datetime import datetime

from TeamsList import getTeam_by_Short 

from ProcessHistoricOddssharkData import RotateCalendar



def captureOddssharkData(game_file, date_string, away_only=False):

    starting_column=21
    starting_row=2 # Starting row of first game entry



    #Automate data capture from Oddsshark with selenium
    options = FirefoxOptions()
    options.add_argument("--headless")
    driver = webdriver.Firefox(options=options)
    initial_url="https://www.oddsshark.com/nba/scores"       

    #Navigate to main ODDSHARK nba scores page
    driver.get(initial_url)
    time.sleep(5)



    #Force today's date
    historic_datetime_object  = datetime.strptime(date_string,'%m/%d/%y') #Convert to python datetime object  = datetime.strptime(df['Date'][date_block_index],'%m/%d/%y') #Convert to python datetime object
    RotateCalendar(driver, historic_datetime_object)



    # Get the current page source.
    time.sleep(5)
    source = driver.page_source


    # Parse into soup() the source of the page after the link is clicked and use "html.parser" as the Parser.
    soupify = soup(source, 'html.parser')

    #Get all upcoming games of the day
    soup_m= soupify.find_all('div', {'class': 'matchup pre'})

    gameURLlist=[]
    for match in soup_m:

        #Capture the URL links
        atags=match.find('a', {'class': 'scores-matchup__link'})
        game_url=atags['href']

        gameURLlist.append('https://www.oddsshark.com'+game_url)


    gameURLs_size=len(gameURLlist)




    #Parse game file
    GAME_workbook = openpyxl.load_workbook(game_file)
    GAME_worksheet = GAME_workbook.worksheets[0]


    if away_only==True:
        #write header
        #Check if header already exists and write header
        col_head = []
        hindx=1
        while True:
            header= GAME_worksheet.cell(1, hindx).value
            if header is None:
                break
            col_head.append(header)
            hindx=hindx+1
        #check if expert name is in the columns header
        try:
            starting_column = col_head.index("ODDSHARK")+1                        
        except:
            #expert does not exist already in header so add it
            starting_column=len(col_head)+1

        GAME_worksheet.cell(1, starting_column).value= "ODDSHARK"

    history_games=GAME_worksheet.max_row-2


    for i in range(history_games):
        
        #Get game info
        away_team_Short =           GAME_worksheet.cell(i+2+1, 0+1).value  #cell indices start from 1
        home_team_Short =           GAME_worksheet.cell(i+2+1, 1+1).value #cell indices start from 1

        historic_datetime_object  = datetime.strptime(GAME_worksheet.cell(i+2+1, 2+1).value,'%m/%d/%y')  #cell indices start from 1
        TeamAWAY=getTeam_by_Short(away_team_Short)
        TeamHOME=getTeam_by_Short(home_team_Short)

        #find url in list
        for j in range(gameURLs_size):
            #Partial matching of team names in URL
            if  TeamAWAY.name[0:5].lower().replace(" ","-")  in gameURLlist[j] and TeamHOME.name[0:5].lower().replace(" ","-")  in gameURLlist[j]:
                print(i,": Processing: ",TeamAWAY.name," - ",TeamHOME.name,": ",gameURLlist[j])                    
    

                #Goto that url and capture data
                driver.get(gameURLlist[j])
                time.sleep(5)


                ############## GET MONEYLINES #########################

                # Find the matchup tab element using a css selector and click it.
                try:
                    matchup_tab=WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, 'react-tabs-0')))
                    matchup_tab.click()
                except:   
                    time.sleep(5)
                    print("\t\tCannot click Matchup tab. Obscured by popup? Trying to click")
                    accept_cookies_button=WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, 'agree-button')))
                    accept_cookies_button.click()
                    time.sleep(5)
                    matchup_tab.click()


                # Get the current page source.
                source = driver.page_source

                # Parse into soup() the source of the page after the link is clicked and use "html.parser" as the Parser.
                soupify = soup(source, 'html.parser')



                ############## GET MEDIAN MONEYLINES FOR BOTH TEAMS ##############
                MoneyLineList = []

                for ultag in soupify.find_all('table', {'class': 'table table--compact-odds table--right-only table--reduced-padding table--react table--double-striped'}):
                            for row in ultag.find_all('tr'):
                                columns = row.find_all('td')
                                if len(columns)>0 and len(columns[len(columns)-2].text)>0:
                                    MoneyLineList.append(int(columns[len(columns)-2].text))


                #Remove the first two lines (opening lines)
                MoneyLineList.pop(0)
                MoneyLineList.pop(0)

                #Get median Moneyline for both teams
                awayMoneyline=np.median(MoneyLineList[::2])
                if math.isnan(awayMoneyline):
                    awayMoneyline=0
                homeMoneyline=np.median(MoneyLineList[1::2])
                if math.isnan(homeMoneyline):
                    homeMoneyline=0


                ############## GET H2H (LAST 10 GAMES) DATA #########################
                H2HList = np.zeros([6,2])


                ultag = soupify.find_all('table', {'class': 'table--teamstats'})


                rows =ultag[2].find_all('tr')  #The H2Htable is the 3rd table (the other two are the Team records table and the Edge Finder table)


                Record=rows[1].find_all('td')[0].text.split("-") 
                H2HList[0,0]=int(Record[0]) #Away record
                H2HList[0,1]=int(Record[1]) #Home record

                H2HList[1,0]=float(rows[4].find_all('td')[0].text)   #Away score
                H2HList[1,1]=float(rows[4].find_all('td')[1].text)  #Home score

                H2HList[2,0]=float(rows[6].find_all('td')[0].text.replace("%","")) #Away FGP
                H2HList[2,1]=float(rows[6].find_all('td')[1].text.replace("%","")) #Home FGP

                H2HList[3,0]=float(rows[4].find_all('td')[0].text) #Away Rebounds
                H2HList[3,1]=float(rows[4].find_all('td')[1].text)  #Home Rebounds

                H2HList[4,0]=eval(rows[8].find_all('td')[0].text)*100 #Away 3PP
                H2HList[4,1]=eval(rows[8].find_all('td')[1].text)*100 #Away 3PP

                H2HList[5,0]=float(rows[9].find_all('td')[0].text) #Away steals
                H2HList[5,1]=float(rows[9].find_all('td')[1].text) #Home steals


                ############## GET CONSENSUS SPREADS #########################
                try:

                    main_header = soupify.find_all('div', {'class': 'graph--duel'})
                    spread_header= main_header[0].find('div', {'class': 'graph--duel__header--middle'})
                    spreads=spread_header.find_all('div', {'class': 'graph--duel__header-value'})

                    away_spread = float(spreads[0].text)
                    home_spread = float(spreads[1].text)
                except:
                    print("Spread data does not exist. Setting to zeros")
                    away_spread=0
                    home_spread=0
                

                ################ GET LAST N GAMES DATA ################################
                LastNdata = np.zeros([7,2])

                #AWAY Team
                WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, 'react-tabs-22'))).click()
                tbl= soupify.find('table', {'class': 'table table--react table--striped table--reduced-padding'})

                for row in tbl.find_all('tr', {'class': 'last-10-games-row gc-row'}):
                    columns = row.find_all('td')
                    
                    scr= columns[2].text.split("-") #score
                    LastNdata[0,0]=LastNdata[0,0]+int(scr[0])
                    LastNdata[1,0]=LastNdata[1,0]+int(scr[1])

                    if int(scr[0])>int(scr[1]):
                        LastNdata[2,0]=LastNdata[2,0]+1 #win
                                        
                    try:
                        LastNdata[3,0]=LastNdata[3,0]+float(columns[4].text or 0) #line
                    except:
                        LastNdata[3,0]=0


                    LastNdata[4,0]=LastNdata[4,0]+float(columns[7].text) #FG%
                    LastNdata[5,0]=LastNdata[5,0]+float(columns[8].text) #FT%
                    LastNdata[6,0]=LastNdata[6,0]+float(columns[9].text) #3PTM




                #HOME Team
                WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, 'react-tabs-24'))).click()
                #we need to reload the source here    
                source = driver.page_source
                soupify = soup(source, 'html.parser')
            
                tbl= soupify.find('table', {'class': 'table table--react table--striped table--reduced-padding'})



                for row in tbl.find_all('tr', {'class': 'last-10-games-row gc-row'}):
                    columns = row.find_all('td')
                
                    scr= columns[2].text.split("-") #score
                    LastNdata[0,1]=LastNdata[0,1]+int(scr[0])
                    LastNdata[1,1]=LastNdata[1,1]+int(scr[1])

                    if int(scr[0])>int(scr[1]):
                        LastNdata[2,1]=LastNdata[2,1]+1 #win
                    
                    try:
                        LastNdata[3,1]=LastNdata[3,1]+float(columns[4].text or 0) #line
                    except:
                        LastNdata[3,1]=0

                    LastNdata[4,1]=LastNdata[4,1]+float(columns[7].text) #FG%
                    LastNdata[5,1]=LastNdata[5,1]+float(columns[8].text) #FT%
                    LastNdata[6,1]=LastNdata[6,1]+float(columns[9].text) #3PTM

                #Get the averages
                LastNdata=LastNdata/10


                #Populate game ODDSSHARK stats in Excel file from captured URL
                if away_only==True:

                    #Get AWAY team moneyline and convert it to probability
                    if awayMoneyline>0:
                        away_prob = 100/(awayMoneyline+100)
                    else:
                        away_prob =(-awayMoneyline)/(-awayMoneyline+100)

                    GAME_worksheet.cell(i+starting_row+1, starting_column).value= away_prob    #cell indices start from 1
                
                else:
                    GAME_worksheet.cell(i+starting_row+1, starting_column+1).value  = awayMoneyline #Moneyline AWAY. cell indices start from 1
                    GAME_worksheet.cell(i+starting_row+1, starting_column+1+1).value  = homeMoneyline #Moneyline HOME. cell indices start from 1

                    GAME_worksheet.cell(i+starting_row+1, starting_column+2+1).value  = away_spread #spread AWAY. cell indices start from 1
                    GAME_worksheet.cell(i+starting_row+1, starting_column+3+1).value  = home_spread #spread HOME. cell indices start from 1

                    GAME_worksheet.cell(i+starting_row+1, starting_column+5+1).value  = int(H2HList[0][0]) #H2H record AWAY. cell indices start from 1
                    GAME_worksheet.cell(i+starting_row+1, starting_column+6+1).value  = int(H2HList[0][1]) #H2H record HOME. cell indices start from 1
                    
                    GAME_worksheet.cell(i+starting_row+1, starting_column+7+1).value  = float(H2HList[1][0]) #H2H Score AWAY. cell indices start from 1
                    GAME_worksheet.cell(i+starting_row+1, starting_column+8+1).value  = float(H2HList[1][1]) #H2H Score HOME. cell indices start from 1

                    GAME_worksheet.cell(i+starting_row+1, starting_column+9+1).value  =  float(H2HList[2][0]) #H2H FGP AWAY. cell indices start from 1
                    GAME_worksheet.cell(i+starting_row+1, starting_column+10+1).value  = float(H2HList[2][1]) #H2H FGP HOME. cell indices start from 1
                    
                    GAME_worksheet.cell(i+starting_row+1, starting_column+11+1).value  = float(H2HList[3][0]) #H2H Rebounds AWAY. cell indices start from 1
                    GAME_worksheet.cell(i+starting_row+1, starting_column+12+1).value  = float(H2HList[3][1]) #H2H Rebounds HOME. cell indices start from 1
                    
                    GAME_worksheet.cell(i+starting_row+1, starting_column+13+1).value  = float(H2HList[4][0]) #H2H 3PP AWAY. cell indices start from 1
                    GAME_worksheet.cell(i+starting_row+1, starting_column+14+1).value  = float(H2HList[4][1]) #H2H 3PP HOME. cell indices start from 1

                    
                    GAME_worksheet.cell(i+starting_row+1, starting_column+15+1).value  = float(H2HList[5][0]) #H2H Steals AWAY. cell indices start from 1
                    GAME_worksheet.cell(i+starting_row+1, starting_column+16+1).value  = float(H2HList[5][1]) #H2H Steals HOME. cell indices start from 1
                    
                    
                    GAME_worksheet.cell(i+starting_row+1, starting_column+18+1).value  = LastNdata[2,0] #Aw Last 10GA Win pct. cell indices start from 1
                    GAME_worksheet.cell(i+starting_row+1, starting_column+19+1).value  = LastNdata[0,0] #Aw Last 10GA Score. cell indices start from 1
                    GAME_worksheet.cell(i+starting_row+1, starting_column+20+1).value  = LastNdata[1,0] #           cell indices start from 1
                    GAME_worksheet.cell(i+starting_row+1, starting_column+21+1).value  = LastNdata[3,0] #Aw Last 10GA Line. cell indices start from 1
                    GAME_worksheet.cell(i+starting_row+1, starting_column+22+1).value  = LastNdata[4,0] #Aw Last 10GA FG pct. cell indices start from 1
                    GAME_worksheet.cell(i+starting_row+1, starting_column+23+1).value  = LastNdata[5,0] #Aw Last 10GA FT pct. cell indices start from 1
                    GAME_worksheet.cell(i+starting_row+1, starting_column+24+1).value  = LastNdata[6,0] #Aw Last 10GA 3PTM pct. cell indices start from 1
                    
                
                    GAME_worksheet.cell(i+starting_row+1, starting_column+26+1).value  = LastNdata[2,1] #Hm Last 10GA Win pct. cell indices start from 1
                    GAME_worksheet.cell(i+starting_row+1, starting_column+27+1).value  = LastNdata[0,1] #Hm Last 10GA Score. cell indices start from 1
                    GAME_worksheet.cell(i+starting_row+1, starting_column+28+1).value  = LastNdata[1,1] #           cell indices start from 1
                    GAME_worksheet.cell(i+starting_row+1, starting_column+29+1).value  = LastNdata[3,1] #Hm Last 10GA Line. cell indices start from 1
                    GAME_worksheet.cell(i+starting_row+1, starting_column+30+1).value  = LastNdata[4,1] #Hm Last 10GA FG pct. cell indices start from 1
                    GAME_worksheet.cell(i+starting_row+1, starting_column+31+1).value  = LastNdata[5,1] #Hm Last 10GA FT pct. cell indices start from 1
                    GAME_worksheet.cell(i+starting_row+1, starting_column+32+1).value  = LastNdata[6,1] #Hm Last 10GA 3PTM pct. cell indices start from 1                


    # We are done with the driver so quit.
    driver.quit()

    #write to the excel history file
    GAME_workbook.save(game_file)  
    
