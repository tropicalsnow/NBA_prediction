import urllib
import urllib.request
import urllib.error
import json
import sys

import numpy as np

import openpyxl

from datetime import datetime
from dateutil import tz


from TeamsList import getTeam_by_partial_Name

URL = "https://api.betfair.com/exchange/betting/json-rpc/v1"
APPKEY = "BAgRJxMl7W7C3qK0" #delayed for demo. FIXED
SESSION_TOKEN=""  

def callAping(jsonrpc_req, url, sessionToken, appKey):

    headers = {'X-Application': appKey, 'X-Authentication': sessionToken, 'content-type': 'application/json'}


    try:
        req = urllib.request.Request(url, jsonrpc_req.encode('utf-8'), headers)
        response = urllib.request.urlopen(req)
        jsonResponse = response.read()
        return jsonResponse.decode('utf-8')
    except urllib.error.URLError as e:
        print (e.reason) 
        print ('Oops no service available at ' + str(url))
        exit()
    except urllib.error.HTTPError:
        print ('Oops not a valid operation from the service ' + str(url))
        exit()


def getEventTypes():
    event_type_req = '{"jsonrpc": "2.0", "method": "SportsAPING/v1.0/listEventTypes", "params": {"filter":{ }}, "id": 1}'
    eventTypesResponse = callAping(event_type_req, URL, SESSION_TOKEN, APPKEY)
    eventTypeLoads = json.loads(eventTypesResponse)


    try:
        eventTypeResults = eventTypeLoads['result']
        return eventTypeResults
    except:
        print ('Exception from API-NG' + str(eventTypeLoads['error']))
        exit()


def getEventTypeIDForEventTypeName(eventTypesResult, requestedEventTypeName):
    if(eventTypesResult is not None):
        for event in eventTypesResult:
            eventTypeName = event['eventType']['name']
            if( eventTypeName == requestedEventTypeName):
                return event['eventType']['id']
    else:
        print ('Oops there is an issue with the input')
        exit()

def getCompetitions(eventTypeID):
    #Find competitions with given eventTypeID
    competitions_req='{"jsonrpc": "2.0", "method": "SportsAPING/v1.0/listCompetitions", "params": {"filter":{ "eventTypeIds":["' + eventTypeID + '"] }}, "id": 1}'

    competitionsResponse = callAping(competitions_req, URL, SESSION_TOKEN, APPKEY)
    competitionLoads = json.loads(competitionsResponse)

    try:
        competitionsResults = competitionLoads['result']
        return competitionsResults
    except:
        print ('Exception from API-NG' + str(competitionLoads['error']))
        exit()

def getCompetitionIDForCompetitionName(competitionsResults, requestedCompetitionName):
    if(competitionsResults is not None):
        for competition in competitionsResults:
            competitionName = competition['competition']['name']
            if( competitionName == requestedCompetitionName):
                return competition['competition']['id']
    else:
        print ('Oops there is an issue with the input')
        exit()   

def getEvents(competitionID):
    #Find events with given competitionID 
    events_req='{"jsonrpc": "2.0", "method": "SportsAPING/v1.0/listEvents", "params": {"filter":{ "competitionIds":["' + competitionID + '"] }}, "id": 1}'

    eventsResponse = callAping(events_req, URL, SESSION_TOKEN, APPKEY)
    eventsLoads = json.loads(eventsResponse)

    try:
        eventsResults = eventsLoads['result']
        return eventsResults
    except:
        print ('Exception from API-NG' + str(eventsLoads['error']))
        exit()


def getMarketCatalogueForEvent( eventID):
    #Get market catalogue with given event ID
    market_catalogue_req = '{"jsonrpc": "2.0", "method": "SportsAPING/v1.0/listMarketCatalogue", "params": {"filter":{ "eventIds":["' + eventID + '"],  "marketCountries":["GB"],"marketTypeCodes":["MATCH_ODDS"]},"maxResults":"1"}, "id": 1}'


    market_catalogue_response = callAping(market_catalogue_req, URL, SESSION_TOKEN, APPKEY)    
    market_catalogue_loads = json.loads(market_catalogue_response)


    try:
        market_catalogue_results = market_catalogue_loads['result']
        return market_catalogue_results
    except: 
        print ('Exception from API-NG' + str(market_catalogue_loads['error']))
        exit()


def getMarketId(market_catalogue_results):
    if( market_catalogue_results is not None):
        for market in market_catalogue_results:
            return market['marketId']



def getMarketBookBestOffers(marketId):
    market_book_req = '{"jsonrpc": "2.0", "method": "SportsAPING/v1.0/listMarketBook", "params": {"marketIds":["' + marketId + '"],"priceProjection":{"priceData":["EX_BEST_OFFERS"]}}, "id": 1}'

    market_book_response = callAping(market_book_req, URL, SESSION_TOKEN, APPKEY)

    market_book_loads = json.loads(market_book_response)
    try:
        market_book_result = market_book_loads['result']
        return market_book_result
    except:
        print ('Exception from API-NG' + str(market_book_result['error']))
        exit()

  
def getPriceInfo(market_book_result):
    odds=np.zeros(2)
    if(market_book_result is not None):
        for marketBook in market_book_result:
            runners = marketBook['runners']
            counter=0
            for runner in runners:
                if (runner['status'] == 'ACTIVE'):
                    odds[counter]=runner['ex']['availableToBack'][0]['price']
                    counter=counter+1

    return odds


def captureBetfairOdds(game_file, token, starting_column=10):

    global SESSION_TOKEN
    SESSION_TOKEN=token

    from_zone = tz.gettz('GMT')
    to_zone = tz.gettz('EST')



    #PARSE GAME FILE
    GAME_workbook = openpyxl.load_workbook(game_file)
    GAME_worksheet = GAME_workbook.worksheets[0]
 
    games_num=GAME_worksheet.max_row-2

    


    #GET PRELIMINARY BETFAIR DATA

    #Query all events and get event ID for basketball games
    eventTypesResult = getEventTypes()
    basketballEventTypeID = getEventTypeIDForEventTypeName(eventTypesResult, 'Basketball')
    
    #Using basketball event ID query all markets and get NBA competition ID
    competitionsResults=getCompetitions(basketballEventTypeID)
    NBACompetitionID= getCompetitionIDForCompetitionName(competitionsResults, 'NBA')
    

    #Using NBA competion ID to get all the available games
    eventsResults=getEvents(NBACompetitionID)




    #loop for games in excel file
    for i in range(games_num):
        
        #Get game info
        away_team_Short =   GAME_worksheet.cell(i+2+1, 0+1).value   #cell indices start from 1
        home_team_Short =   GAME_worksheet.cell(i+2+1, 1+1).value   #cell indices start from 1
        date_string =       GAME_worksheet.cell(i+2+1, 2+1).value   #cell indices start from 1


        #Convert to python datetime object  = datetime.strptime(df['Date'][date_block_index],'%m/%d/%y') 
        dateEST  = datetime.strptime(date_string,'%m/%d/%y') 



        #loop for events in Betfair exchange NBA list 
        for event in eventsResults:
            eventName = event['event']['name'].split('@')
            
            dateGMT=  event['event']['openDate'][0:16]
            dateGMT =  datetime.strptime(dateGMT, '%Y-%m-%dT%H:%M')
            dateGMT = dateGMT.replace(tzinfo=from_zone)
            dateGMT = dateGMT.astimezone(to_zone)
            
            eventID =  event['event']['id']
            
            if len(eventName)==2:
                #Convert to short names
                Betfair_Away_Short = getTeam_by_partial_Name(eventName[0].rstrip().lstrip()).short
                Betfair_Home_Short = getTeam_by_partial_Name(eventName[1].rstrip().lstrip()).short

                 #Check that the dates match and that the teams match   
                if (dateGMT.year == dateEST.year and dateGMT.month == dateEST.month and dateGMT.day == dateEST.day)  and (Betfair_Away_Short == away_team_Short and Betfair_Home_Short == home_team_Short  ):      

                    catalogue=getMarketCatalogueForEvent(eventID)
                    marketId=getMarketId(catalogue)
 
                    market_book_result=getMarketBookBestOffers(marketId)
                    odds=getPriceInfo(market_book_result)
                    
                    print(Betfair_Away_Short, Betfair_Home_Short, odds)    

                    #Write the odds to the file
                    GAME_worksheet.cell(2+i+1, starting_column+1).value = float(odds[0])
                    GAME_worksheet.cell(2+i+1, starting_column+1+1).value = float(odds[1])

                    
                    #Add predictions and probs
                    if odds[0]<odds[1]:
                        GAME_worksheet.cell(2+i+1, starting_column+2+1).value = float(1)
                    else:
                        GAME_worksheet.cell(2+i+1, starting_column+2+1).value = float(2)

                    GAME_worksheet.cell(2+i+1, starting_column+3+1).value = float(1./odds[0]*100)
                    GAME_worksheet.cell(2+i+1, starting_column+4+1).value = float(1./odds[1]*100)



    GAME_workbook.save(game_file)  


