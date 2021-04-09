#new version: Captures from DONBEST.com instead of COVERS.com

import numpy as np

import openpyxl
import requests


from datetime import datetime
#from bs4 import BeautifulSoup, NavigableString, Tag

import json  

from TeamsList import getTeam_by_partial_ANY, getTeam_by_partial_Name



def captureInjuryList(data_file):
    #Grab PRE-GAME predicted injury data. This is only used for PRE-game prediction


    donbest_header={
    'Host': 'sports-prd-ghosts-api-widgets.sports.gracenote.com',
    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:83.0) Gecko/20100101 Firefox/83.0',
    'Accept': 'application/json, text/plain, */*',
    'Accept-Language': 'en-US,en;q=0.5',
    'Accept-Encoding': 'gzip, deflate, br',
    'Referer': 'https://widgets.sports.gracenote.com/latest/index.html',
    'Pragma': 'no-cache',
    'Cache-Control': 'no-cache',
    'Origin': 'https://widgets.sports.gracenote.com',
    'Connection': 'keep-alive',
    'TE': 'Trailers'
    }


    url='http://www.donbest.com/nba/injuries/'

    #query the GRACENOTE api, faking a request from donbest sports
    customerId='1412'  #fixed for Donbest?
    seasonID='664624' #Does this change every season?
    url='https://sports-prd-ghosts-api-widgets.sports.gracenote.com/api/Injuries?customerId='+customerId+\
    '&editionId=%2Fsport%2Fbasketball%2Fseason:'+seasonID+\
    '&filter=%7B%22include%22:%5B%22team%22,%22players%22%5D,%22fields%22:%7B%22teamInjuries%22:'\
    '%7B%22injuries%22:%7B%22playerId%22:true,%22location%22:true,%22teamId%22:true,%22startDate%22:'\
    'true,%22injury%22:true,%22status%22:true,%22displayStatus%22:true,%22note%22:true,%22lastUpdated%22:'\
    'true%7D%7D,%22players%22:%7B%22playerFirstName%22:true,%22playerLastName%22:true,%22thumbnailUrl%22:'\
    'true,%22seasonDetails%22:%7B%22position%22:%7B%22positionShortName%22:true%7D%7D%7D,%22team%22:'\
    '%7B%22teamName%22:true,%22teamShortName%22:true,%22type%22:true%7D%7D%7D&languageCode=2&module=na_teamsports'\
    '&sportId=%2Fsport%2Fbasketball&type=injuries'

    
    response = requests.get(url, headers=donbest_header)
    
    
    if response.status_code==200: 
    
        injuries_text = json.loads(response.text)['injuries']['teamInjuries']

        
        #Parse injury file
        workbook = openpyxl.load_workbook(data_file)
        worksheet = workbook.worksheets[0]

    

        no_games=worksheet.max_row-2
        
        #Loop over games data
        for i in  range(no_games):

            away_players_string=''
            home_players_string=''

            #Get game info
            away_team_Short =   worksheet.cell(i+2+1, 0+1).value #cell indices start from 1
            home_team_Short =   worksheet.cell(i+2+1, 1+1).value #cell indices start from 1

            print('\n',away_team_Short, "-", home_team_Short)

          
        
            #loop for all available teams
            for j in range(len(injuries_text)):
                team_name = injuries_text[j]['team'][0]['teamName']   

                try:
                    Inj_team=getTeam_by_partial_Name(team_name).short
                except:
                    try:
                        Inj_team=getTeam_by_partial_ANY(team_name).short
                    except:
                        continue


                if Inj_team==away_team_Short:
                    TA_match=True
                    TH_match=False
                elif Inj_team==home_team_Short:
                    TH_match=True
                    TA_match=False
                else:
                    TH_match=False
                    TA_match=False

                if (TA_match or TH_match) and len(injuries_text[j]['injuries'])>0:
                    #loop for all potential players in the injury list of that team
                    for p in range(len(injuries_text[j]['injuries'])):
                        player=injuries_text[j]['injuries'][p]['playerFirstName'] +' '+ injuries_text[j]['injuries'][p]['playerLastName']

                        if TA_match:
                            away_players_string=away_players_string+player+', '
            
                        if TH_match:
                            home_players_string=home_players_string+player+', '

            #Add player strings to excel. Remove the last comma and space
            worksheet.cell(2+i+1, 3+1).value= away_players_string[:-2]  #Away. Cell indices start from 1
            worksheet.cell(2+i+1, 4+1).value= home_players_string[:-2]  #Home. Cell indices start from 1
    
    workbook.save(data_file)  



