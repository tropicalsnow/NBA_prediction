import openpyxl
import sqlite3
import numpy as np
import pandas as pd

from capturePredictionTrackerData import capturePredictionTrackerData
from captureCarmeloData import captureCarmeloData
from ProcessHistoricCoversData import ProcessCoversData
from captureOddssharkData import  captureOddssharkData
from ProcessHistoricMarketData import Capture_OddsPortal_UpcomingData

import os.path, sys
sys.path.append('Libs\Expert_ensemble')

from ensemble_methods import GlobalBrier_optimiser_method, Stacker_method
from generateExpertPredictionDatabase import create_Database


def extract_data(data, expert_ids, predictions):
    #extract predictions and labels to numpy arrays
    
    no_games = len(data)
    no_experts = len(expert_ids)

    #store the extracted features & labels here
    features = np.zeros([no_games, no_experts])
    weights  = np.zeros([no_games])
    labels =  np.zeros([no_games])
    soft_labels =  np.zeros([no_games])

    for i in range(no_games):

        print("\tExtracting features from game: %i/%i  " %(i+1, no_games) , end="\r", flush=True)

        #grab game info
        game_id, result, scoreA, scoreH, season= data.iloc[i][['game_id','Result','Away_Score', 'Home_Score','Season']]

        #Get all the predictions for this game.
        game_preds = predictions[predictions['game_id']==game_id]

        #loop for all the experts
        for j in range(no_experts):
            #get that expert's Away only probabilities. We dont need the probs for home since they sum up to 1

            try:
                [[pA]]=  game_preds[game_preds['expert_id'] == expert_ids[j]][['Prob_Away']].values
            except:
                #no predictions found for that expert so set them to 0.5 (unknown outcome)
                pA=0.5
        
            #add probs as features
            features[i, j] = pA

            


        #add label
        labels[i]=result
        
        #add weight. In this case just the season year
        weights[i] = int(season)
        
        #soft label wrt the AWAY team
        # soft_labels[i]=Spread2Prob_Convert(scoreH-scoreA, [-0.04927,  0.1403, 0.5472, -0.04684]) #probability
        soft_labels[i]=scoreH-scoreA  #point spreads
        

    print("\n")
    return features, labels, soft_labels, weights



def rearrangeTest_data(testDF, expertsDF):

    #Rearrange test data according to the order of the experts in the Experts sql database

    num_experts = expertsDF.shape[0]
    num_games = testDF.shape[0]

    X_test=np.zeros([num_games,num_experts])
 
    
    for i in range(num_experts):
    
        expert_name = expertsDF['expert_name'].iloc[i]

        #Only match the experts that exist in the Database. If the extracted file has more experts then we ignore them
        try:
            index = list(testDF.columns).index(expert_name)
            out = np.array(testDF.iloc[:,index])
        except:
            #expert does not exist. Set all values to 0.5
            out = 0.5*np.ones(num_games)


        X_test[:,i] = out
 
    return X_test



def capture_all_experts(experts_file,  date_string):

    print("Capturing experts data")

    #Add CARMELO data
    print('\t--Appending CARMELO data')
    captureCarmeloData(experts_file, date_string)
    
    
    #Add COVERS data
    print('\t--Appending COVERS data')
    ProcessCoversData(experts_file, away_only=True)

    #Add ODDSHARK data
    print('\t--Appending ODDSHARK data')
    captureOddssharkData(experts_file,date_string, away_only=True) 

    
    #Add prediction tracker data
    print('\t--Appending Prediction Tracker data')
    capturePredictionTrackerData(experts_file,  starting_column=8)

    
    #Add oddsportal data
    print('\t--Appending Oddsportal data')
    Capture_OddsPortal_UpcomingData(experts_file)

    #Load the excel file and fill NaNs and zeros with 0.5, which is a non-prediction
    df = pd.read_excel(experts_file)
    df= df.drop(0,axis=0)
    df=df.fillna(0.5)
    df=df.replace(0, 0.5)

    return df


def ensemble_metaPredict(experts_file, scheduled_games_file, our_predictions_AWAY):

    #We only consider AWAY team probabilities

    
    #Write our probs to file
    game_workbook = openpyxl.load_workbook(experts_file)
    game_worksheet = game_workbook.worksheets[0]
    game_worksheet.cell(1,4).value ="OUR_Predictor"   

    n_games=game_worksheet.max_row-2
    date_string=game_worksheet.cell(3, 3).value

    for i in range(n_games):
        game_worksheet.cell(i+3, 4).value  =   our_predictions_AWAY[i]   #cell indices start from 1
    game_workbook.save(experts_file)  







    print("Starting ensemble prediction...")

    #First capture all the experts for the CURRENT games
    captured_df=capture_all_experts(experts_file, date_string)

    
    
    #TRAIN FROM EXPERTS DATABASE
    
    #load data from database
    db_filename='./Libs/Expert_ensemble/ExpertPredictions_DB.db'
    if  not os.path.isfile(db_filename):
        #create database if it doesnt exist
        create_Database(db_filename, [2015, 2016, 2017, 2018, 2019, 2020])


    conn=sqlite3.connect(db_filename)
    cursor = conn.cursor()

    #Get all the Predictions, Experts & Games data and save them to Pandas df
    all_game_rows=cursor.execute("SELECT game_id, Away_Score, Home_Score, Result, Season from GAMES").fetchall()
    col_names= np.array(cursor.description)
    games_DF = pd.DataFrame(all_game_rows, columns=col_names[:,0])

    all_pred_rows=cursor.execute("SELECT * from Predictions").fetchall()
    col_names= np.array(cursor.description)
    predictions_DF = pd.DataFrame(all_pred_rows, columns=col_names[:,0])

    all_expert_rows=cursor.execute("SELECT * from Experts").fetchall()
    col_names= np.array(cursor.description)
    experts_DF = pd.DataFrame(all_expert_rows, columns=col_names[:,0])

    #Close database
    conn.close()


    #extract the data
    expert_ids = experts_DF['expert_id'].values #these are already unique by definition of the database

    try:
        X_train= np.loadtxt("./temp/temp_Xtrain.csv", delimiter=',')
        y_train, weights_train= np.loadtxt("./temp/temp_y_train.csv", delimiter=',' , usecols=(0, 1), unpack=True)
    except:
        X_train, y_train, _, weights_train = extract_data(games_DF, expert_ids, predictions_DF)
        np.savetxt("./temp/temp_Xtrain.csv", X_train, delimiter=',', fmt='%s')
        np.savetxt("./temp/temp_y_train.csv", np.column_stack((y_train,weights_train)) , delimiter=',', fmt='%s')

    
    
    #TRAIN & PREDICT 
    #convert test data to approriate format, then train ensemble predictor and predict
    print("\nPredicting ensemble using %i samples from %i experts"%(X_train.shape))
    X_test = rearrangeTest_data(captured_df, experts_DF)
    #use one of the ensemble predictor methods
    # X_test_preds= Stacker_method(X_train, y_train, X_test)  #STACKER
    X_test_preds, _, _ = GlobalBrier_optimiser_method(X_train, y_train, X_test, weights_train) #GLOBAL OPTIMISER

    #WRITE TO FILE
    scheduled_games_workbook = openpyxl.load_workbook(scheduled_games_file)
    scheduled_games_worksheet = scheduled_games_workbook.worksheets[0]

    #just write the data in the order that was read
    for i in range(len(X_test_preds)):

        if  len(X_test_preds.shape)==1:
            #wrt to AWAY team
            if X_test_preds[i]>0.5:
                result = 1
            else:
                result = 2
          

            scheduled_games_worksheet.cell(i+3, 7).value  = result
            scheduled_games_worksheet.cell(i+3, 8).value  = X_test_preds[i]*100   #AWAY
            scheduled_games_worksheet.cell(i+3, 9).value  =  (1-X_test_preds[i])*100  #HOME

        else:
            if X_test_preds[i,0]>X_test_preds[i,1]:
                result = 1
            else:
                result = 2

            scheduled_games_worksheet.cell(i+3, 7).value  = result
            scheduled_games_worksheet.cell(i+3, 8).value  = X_test_preds[i,0]*100   #AWAY
            scheduled_games_worksheet.cell(i+3, 9).value  =  X_test_preds[i,1]*100  #HOME



    scheduled_games_workbook.save(scheduled_games_file)  
