import configparser
import json  
from collections import namedtuple
from scrape_NBA_Data import scrape_season_Database, calculateAvg_season_Player_Stats, append_team_roster
import os.path
from featureUtilities import selectFeatures, generateAttributeIndices, generate_Features_fromDB, engineerFeatures, Predict_Features_forLineup, Spread2Prob_Convert
from PredictPlayerStats import getPlayerPrediction
import joblib
import numpy as np
import sys
from seasonDB_utils import open_database, create_season_Database
import pandas as pd
import openpyxl
from datetime import datetime


sys.path.append('Betting_strategies')


import functools
import operator
from shutil import copyfile

import requests


import math

sys.path.append('../')
sys.path.append('../Betting_strategies')
from runBacktest_modular import LoadBackTestData

from LoadSaveXLSData import LoadInjuryList, SaveTrainingData_To_Excel
  
from sklearn.metrics import brier_score_loss
from sklearn.utils import class_weight
import sklearn.model_selection



import matplotlib.pyplot as plt


sys.path.append('../')

 
import TeamsList
from TeamsList import getTeam_by_partial_ANY, getTeam_by_Short, getTeam_by_ID


from betacal import BetaCalibration

from ensemble_metaPredict import extract_data
from ensemble_methods import GlobalBrier_optimiser_method







def str2bool(v):
  return str(v).lower() in ("yes", "true", "t", "1")


def WriteScoresToExcel(AccuArray, BrierArray, ProbsArray, filename):
    #Create a workbook and add a worksheet.
    workbook = xlsxwriter.Workbook(filename)  
    worksheet = workbook.add_worksheet()

    #headers
    worksheet.write(0,   0, "cumm. Accuracy") 
    worksheet.write(0,   1, "cumm. Brier") 
    worksheet.write(0,   2, "Probs") 

    for i in range(len(AccuArray)):
        worksheet.write(i+1, 0, AccuArray[i]) 
        worksheet.write(i+1, 1, BrierArray[i]) 
        worksheet.write(i+1, 2, ProbsArray[i]) 

    workbook.close()


def LoadSettings(settings_file):

    config = configparser.ConfigParser()
    config.read(settings_file)

    #Variables
    trainingSeasons = json.loads(config.get("Variables","trainingSeasons"))
    testSeason = json.loads(config.get("Variables","testSeason"))
    chosen_attributes   = json.loads(config.get("Variables","chosen_attributes"))
    databaseDir=  json.loads(config.get("Variables","databaseDir"))
    plot_graphs= str2bool(json.loads(config.get("Variables","plot_graphs")))
    useMetaEnsemble=str2bool(json.loads(config.get("Variables","useMetaEnsemble")))
    
    #Classifier settings
    num_of_features=  json.loads(config.get("Classifier settings","num_of_features"))
    use_feature_selector= str2bool(json.loads(config.get("Classifier settings","use_feature_selector")))
    feature_engineer =  str2bool(json.loads(config.get("Classifier settings","use_feature_engineering")))
    label_type = json.loads(config.get("Classifier settings","labels"))
    add_gameno_as_feature = str2bool(json.loads(config.get("Classifier settings","add_gameno_as_feature")))
    reverse_home_advantage = str2bool(json.loads(config.get("Classifier settings","reverse_home_advantage")))
    backtestMetaEnsemble=   str2bool(json.loads(config.get("Classifier settings","backtestMetaEnsemble")))
 
    #Train options
    saveToExcel = str2bool(json.loads(config.get("Train options","saveToExcel")))
    burn_in= json.loads(config.get("Train options","burn_in"))
    live_retrain = str2bool(json.loads(config.get("Train options","live_retrain")))

    #Prediction
    topN = json.loads(config.get("Prediction","topN"))
    MCruns = json.loads(config.get("Prediction","MCruns"))

 
    
    ParamStruct = namedtuple("Params", "trainingSeasons testSeason databaseDir plot_graphs useMetaEnsemble \
                                        chosen_attributes   num_of_features use_feature_selector feature_engineer label_type add_gameno_as_feature reverse_home_advantage backtestMetaEnsemble\
                                        saveToExcel burn_in live_retrain topN MCruns")
    return ParamStruct(trainingSeasons, testSeason, databaseDir, plot_graphs, useMetaEnsemble,
                       chosen_attributes, num_of_features, use_feature_selector, feature_engineer, label_type, add_gameno_as_feature, reverse_home_advantage, backtestMetaEnsemble,
                       saveToExcel, burn_in, live_retrain, topN, MCruns)

def scrape_all_databases(Params):

    for i in range(len(Params.trainingSeasons)+1):
  
        if i <len(Params.trainingSeasons):
            season_str = str(Params.trainingSeasons[i]-1)+"-"+str(Params.trainingSeasons[i])[2:4]
        else:
            season_str =   str(Params.testSeason-1)+"-"+str(Params.testSeason)[2:4]

        
        print("Season: "+season_str)

        scrape_season_Database(season_str, Params.databaseDir)

def generateTrainingData(Params, Features_FILENAME):

    if  not os.path.isfile("./Features/"+Features_FILENAME+"sav"):
        #if feature file does not exist then generate it
        print("Feature file not found. Generating...")

 

        Features=[]  #holds the team & player features (individual game boxscores) for all the games in all the training seasons
        Labels =[]   #holds the labels for all the games in all the training seasons
        Weights=[]   #Holds the weights for each datapoint. Every datapoint from each season has the same weight
        

        #generate the weights
        #w_seasons = np.ones(len(Params.trainingSeasons)) #Constant i.e. no weighting
        #w_seasons = np.linspace(0.1,1,len(Params.trainingSeasons)) #Linear
        #w_seasons =  1/np.logspace(1, 0, len(Params.trainingSeasons)) #logarithmic
        #w_seasons = [0.9787, 0.9477, 0.9604, 0.9679, 0.9793, 0.9862, 1.0000, 0.9866, 0.9491, 0.9881, 0.9676, 0.9832, 0.9741, 0.9850, 0.9892, 0.9962, 0.9815, 0.9868, 0.9628, 0.9682] #manual based on cross validation
        w_seasons = np.power(10, np.linspace(0.1,1,len(Params.trainingSeasons))) #powerlaw

        #normalise weights
        w_seasons= w_seasons/max(w_seasons)

        #loop over seasons
        for i in range(len(Params.trainingSeasons)):

            s=Params.trainingSeasons[i]

            season_str = str(s-1)+"-"+str(s)[2:4]
            #get database file associated with season
            databaseFile = Params.databaseDir+"/NBA_season"+season_str+".db"
            season_Features, seasonLabels = generate_Features_fromDB(Params, databaseFile)      

            #Append current season features and labels to the total
            if len(Labels)==0:
                Labels=np.array(seasonLabels,dtype=np.int)
                Weights=w_seasons[i]*np.ones_like(seasonLabels)

                Features=np.array(season_Features)
            else:
                Labels = np.append(Labels, seasonLabels)
                Weights = np.append(Weights, w_seasons[i]*np.ones_like(seasonLabels) )

                Features = np.vstack([Features, season_Features])





        
        #Write Features and labels to file
        joblib.dump((Features,Labels, Weights), "./Features/"+Features_FILENAME+"sav")
       
 
    else:
        #otherwise just load it
        print("Feature file found. Loading...")
        Features,Labels,Weights=joblib.load("./Features/"+Features_FILENAME+"sav")



    if Params.saveToExcel==True and not os.path.isfile("./Features/"+Features_FILENAME+"xlsx"):
        print("Saving features+labels to excel file: ",Features_FILENAME+"xlsx")
        SaveTrainingData_To_Excel(Features,Labels,"./Features/"+Features_FILENAME+"xlsx",Weights)


    return Features, Labels, Weights
 

def PredictLineup(Params, team_ID, injury_list_FILENAME, curr_season_database, player_dbms):
    
    lineup=[]
    expected_lineup=[]
    cursor=curr_season_database.conn.cursor()
    player_cursor=player_dbms.conn.cursor()
    
    #Take the whole roster as the lineup
    season_roster=np.array(cursor.execute('SELECT * FROM Teams WHERE team_id='+team_ID).fetchone())[3:]
    season_roster=list(np.delete(season_roster, np.where(season_roster==None)).astype(int))


    
    #Use roster to determine an expected 10 player lineup based on expected minutes of playtime
    exp_minutes=[]
    for i in season_roster:
        
        player_exp_min = player_cursor.execute('SELECT MIN FROM PredictedStats WHERE player_id="'+str(i)+'" and season="'+str(Params.testSeason)+'"').fetchone()
     

        if player_exp_min==None:
            #player does not exist in prediction database so add him
            player_exp_stats, _=getPlayerPrediction(Params, str(i))
            if len(player_exp_stats)==0:
                #Rookie or some fault. Essentially ignore player
                player_exp_min=0
            else:
                player_exp_min=player_exp_stats[0]
            
            exp_minutes.append(float(player_exp_min))
        else: 
            exp_minutes.append(float(player_exp_min[0]))

    #order by descending number of minutes
    sorted_index=np.argsort(-np.array(exp_minutes))
    #take N most probable players by expected number of minutes
    N=10 
    for i in range(min(N,len(season_roster))):
        expected_lineup.append(season_roster[sorted_index[i]])







    #First check for existing games in database. If the team has already played some games then the most probable lineup is the one from the previous
    #game, minus the injury list. If no games exist in the database then try to predict the most likely lineup from the roster and the number of PREDICTED minutes played 
    games_curr = cursor.execute('SELECT * FROM games WHERE Away_Team_id='+team_ID+' or  Home_Team_id='+team_ID).fetchall()
    if len(games_curr)>0:
        #Take last game lineup
        game_ID = games_curr[-1][0] #TODO: Check if this is a reliable way to get last game player. This assumes correct ordering in the SQL database
        lineup =  cursor.execute('SELECT player_id FROM Lineups WHERE game_id="'+game_ID+'" and  team_id='+team_ID).fetchall()

        lineup = functools.reduce(operator.concat, lineup)
        lineup = [int(i) for i in lineup]
    else:
        lineup=expected_lineup
    

    #Filter players in the injury list
    if len(lineup)>0:
        #Remove injury list if we have it available
        ALL_InjuryLists = LoadInjuryList(injury_list_FILENAME)
        if ALL_InjuryLists is not None:
            for i in ALL_InjuryLists:
                if i.AwayList is not None:
                    for ia in i.AwayList.split(", "):
                        player_id=cursor.execute('SELECT player_id FROM Players WHERE player_name="'+ia+'"').fetchone()
                        if player_id is not None:
                            player_id = functools.reduce(operator.concat, player_id)

                            if int(player_id) in lineup:
                                lineup.remove(int(player_id))

                if i.HomeList is not None:
                    for ih in i.HomeList.split(", "):
                        player_id=cursor.execute('SELECT player_id FROM Players WHERE player_name="'+ih+'"').fetchone()
                        if player_id is not None:
                            player_id = functools.reduce(operator.concat, player_id)
                            
                            if int(player_id) in lineup:
                                lineup.remove(int(player_id))



    #filter by roster. TODO: Fix this so we dont have less than 5 players
    temP=   list(set(lineup ) & set(season_roster))                         
    if len(temP)<5:
        temP=expected_lineup
    return temP 

 
def addToTempDatabase(gm_indx_list, temp_dbms, dbms):
    players_that_played_index=[] #the index of all players that have played in the games inside gm_indx_list

    cursor= dbms.conn.cursor()
    temp_cursor= temp_dbms.conn.cursor()


    for game_id in gm_indx_list:

        #append game info to temp_dbms from dmbs database. Also update team/player averages

        #Game
        g_row = cursor.execute('SELECT * FROM games where game_id="'+game_id+'"').fetchone()   
        temp_cursor.execute('INSERT  OR REPLACE INTO games (game_id,Away_Team_id,Home_Team_id,Away_Score,Home_Score,Result,Date) VALUES(?,?,?,?,?,?,?)',
                                              (g_row[0], g_row[1], g_row[2], g_row[3], g_row[4], g_row[5], g_row[6] ) )

        #Lineups.  
        l_rows = cursor.execute('SELECT * FROM Lineups where game_id="'+game_id+'"').fetchall()   
        for i in range(len(l_rows)):

            players_that_played_index.append(l_rows[i][0])    

            temp_cursor.execute('INSERT OR REPLACE  INTO lineups (player_id, game_id, team_id) VALUES(?,?,?)',
                                            (l_rows[i][0], l_rows[i][1], l_rows[i][2]) )
            #Players. This will do a lot of unecessary inserts
            p_row =  cursor.execute('SELECT player_id, player_name, age, exp, weight, height FROM Players where player_id="'+l_rows[i][0]+'"').fetchone()   
            if p_row is not None:  #why a player in the lineup will not be in the database?
                temp_cursor.execute('INSERT OR REPLACE  INTO Players (player_id, player_name, age, exp, weight, height) VALUES(?,?,?,?,?,?)',
                                             (p_row[0],  p_row[1], p_row[2], p_row[3], p_row[4], p_row[5] ) )


        
        #Boxscores player
        bp_rows= cursor.execute('SELECT * FROM BoxscoresPlayer where game_id="'+game_id+'"').fetchall()   
        for i in range(len(bp_rows)):
            temp_cursor.execute('INSERT OR REPLACE INTO boxscoresPlayer (game_id, player_id,  MIN, FGM, FGA, FG_PCT, FG3M, FG3A, FG3_PCT, FTM, \
                                                                                FTA, FT_PCT, OREB, DREB, REB, AST, STL, BLK, TOV, PF, \
                                                                                PTS, PLUS_MINUS,  E_OFF_RATING, OFF_RATING, E_DEF_RATING, DEF_RATING, E_NET_RATING, NET_RATING,  AST_PCT, AST_TOV, \
                                                                                AST_RATIO, OREB_PCT, DREB_PCT, REB_PCT, TOV_PCT, EFG_PCT, TS_PCT, USG_PCT, E_USG_PCT, E_PACE, \
                                                                                PACE, PIE, PTS_OFF_TOV,  PTS_2ND_CHANCE, PTS_FB, PTS_PAINT, OPP_PTS_OFF_TOV, OPP_PTS_2ND_CHANCE, OPP_PTS_FB, OPP_PTS_PAINT, \
                                                                                BLKA, PFD, PCT_FGA_2PT, PCT_FGA_3PT, PCT_PTS_2PT, PCT_PTS_2PT_MR, PCT_PTS_3PT, PCT_PTS_FB, PCT_PTS_FT, PCT_PTS_OFF_TOV, \
                                                                                PCT_PTS_PAINT, PCT_AST_2PM, PCT_UAST_2PM, PCT_AST_3PM, PCT_UAST_3PM, PCT_AST_FGM, PCT_UAST_FGM, USG_PCT_USG, PCT_FGM, PCT_FGA, \
                                                                                PCT_FG3M, PCT_FG3A, PCT_FTM, PCT_FTA, PCT_OREB, PCT_DREB, PCT_REB, PCT_AST, PCT_TOV, PCT_STL, \
                                                                                PCT_BLK, PCT_BLKA, PCT_PF, PCT_PFD, PCT_PTS,  EFG_PCT_FF, FTA_RATE, TM_TOV_PCT, OREB_PCT_FF, OPP_EFG_PCT, \
                                                                                OPP_FTA_RATE, OPP_TOV_PCT, OPP_OREB_PCT, SPD, DIST, ORBC, DRBC, RBC, TCHS, SAST, \
                                                                                FTAST, PASS, CFGM, CFGA, CFG_PCT, UFGM, UFGA, UFG_PCT, DFGM, DFGA, \
                                                                                DFG_PCT,  CONTESTED_SHOTS, CONTESTED_SHOTS_2PT, CONTESTED_SHOTS_3PT, DEFLECTIONS, CHARGES_DRAWN, SCREEN_ASSISTS, SCREEN_AST_PTS, OFF_LOOSE_BALLS_RECOVERED, DEF_LOOSE_BALLS_RECOVERED, \
                                                                                LOOSE_BALLS_RECOVERED, OFF_BOXOUTS, DEF_BOXOUTS, BOX_OUT_PLAYER_TEAM_REBS, BOX_OUT_PLAYER_REBS,  BOX_OUTS)' 
                                                'VALUES(?,?,?,?,?,?,?,?,?,?, ?,?,?,?,?,?,?,?,?,?, ?,?,?,?,?,?,?,?,?,?,\
                                                        ?,?,?,?,?,?,?,?,?,?, ?,?,?,?,?,?,?,?,?,?, ?,?,?,?,?,?,?,?,?,?,\
                                                        ?,?,?,?,?,?,?,?,?,?, ?,?,?,?,?,?,?,?,?,?, ?,?,?,?,?,?,?,?,?,?,\
                                                        ?,?,?,?,?,?,?,?,?,?, ?,?,?,?,?,?,?,?,?,?, ?,?,?,?,?,?,?,?,?,?,\
                                                        ?,?,?,?,?,?)',
                                            (bp_rows[i][0], bp_rows[i][1], bp_rows[i][2], bp_rows[i][3], bp_rows[i][4], bp_rows[i][5], bp_rows[i][6], bp_rows[i][7], bp_rows[i][8], bp_rows[i][9],
                                            bp_rows[i][10], bp_rows[i][11], bp_rows[i][12], bp_rows[i][13], bp_rows[i][14], bp_rows[i][15], bp_rows[i][16], bp_rows[i][17], bp_rows[i][18], bp_rows[i][19],
                                            bp_rows[i][20], bp_rows[i][21], bp_rows[i][22], bp_rows[i][23], bp_rows[i][24], bp_rows[i][25], bp_rows[i][26], bp_rows[i][27], bp_rows[i][28], bp_rows[i][29],
                                            bp_rows[i][30], bp_rows[i][31], bp_rows[i][32], bp_rows[i][33], bp_rows[i][34], bp_rows[i][35], bp_rows[i][36], bp_rows[i][37], bp_rows[i][38], bp_rows[i][39],
                                            bp_rows[i][40], bp_rows[i][41], bp_rows[i][42], bp_rows[i][43], bp_rows[i][44], bp_rows[i][45], bp_rows[i][46], bp_rows[i][47], bp_rows[i][48], bp_rows[i][49],
                                            bp_rows[i][50], bp_rows[i][51], bp_rows[i][52], bp_rows[i][53], bp_rows[i][54], bp_rows[i][55], bp_rows[i][56], bp_rows[i][57], bp_rows[i][58], bp_rows[i][59],
                                            bp_rows[i][60], bp_rows[i][61], bp_rows[i][62], bp_rows[i][63], bp_rows[i][64], bp_rows[i][65], bp_rows[i][66], bp_rows[i][67], bp_rows[i][68], bp_rows[i][69],
                                            bp_rows[i][70], bp_rows[i][71], bp_rows[i][72], bp_rows[i][73], bp_rows[i][74], bp_rows[i][75], bp_rows[i][76], bp_rows[i][77], bp_rows[i][78], bp_rows[i][79],
                                            bp_rows[i][80], bp_rows[i][81], bp_rows[i][82], bp_rows[i][83], bp_rows[i][84], bp_rows[i][85], bp_rows[i][86], bp_rows[i][87], bp_rows[i][88], bp_rows[i][89],
                                            bp_rows[i][90], bp_rows[i][91], bp_rows[i][92], bp_rows[i][93], bp_rows[i][94], bp_rows[i][95], bp_rows[i][96], bp_rows[i][97], bp_rows[i][98], bp_rows[i][99],
                                            bp_rows[i][100], bp_rows[i][101], bp_rows[i][102], bp_rows[i][103], bp_rows[i][104], bp_rows[i][105], bp_rows[i][106], bp_rows[i][107], bp_rows[i][108], bp_rows[i][109],
                                            bp_rows[i][110], bp_rows[i][111], bp_rows[i][112], bp_rows[i][113], bp_rows[i][114], bp_rows[i][115], bp_rows[i][116], bp_rows[i][117], bp_rows[i][118], bp_rows[i][119],
                                            bp_rows[i][120], bp_rows[i][121], bp_rows[i][122], bp_rows[i][123], bp_rows[i][124], bp_rows[i][125]))      
    


    temp_dbms.conn.commit()



    #Calculate season game-average stats for the players
    calculateAvg_season_Player_Stats(temp_dbms, players_that_played_index, False)


def shuffle_weights(model, weights=None):
    """Randomly permute the weights in `model`, or the given `weights`.
    This is a fast approximation of re-initializing the weights of a model.
    Assumes weights are distributed independently of the dimensions of the weight tensors
      (i.e., the weights have the same distribution along each dimension).
    :param Model model: Modify the weights of the given model.
    :param list(ndarray) weights: The model's weights will be replaced by a random permutation of these weights.
      If `None`, permute the model's current weights.
    """
    if weights is None:
        weights = model.get_weights()
    weights = [np.random.permutation(w.flat).reshape(w.shape) for w in weights]
    # Faster, but less random: only permutes along the first dimension
    # weights = [np.random.permutation(w) for w in weights]
    model.set_weights(weights)


def trainClassifier(Params, model, TrainFeatures, TrainLabels, TrainWeights):
    
    from sklearn.model_selection import KFold, cross_val_score, GridSearchCV, train_test_split

    model_filename="./temp/trained_model.dat"
    if  os.path.isfile(model_filename):
        # load model from file
        model = joblib.load(model_filename)
    else:

        answer = input("No trained model found. Do you want to search for meta-parameters Y/N? ")
        if answer.capitalize() == "Y":

            print("Searching for meta-parameters.")
            
            param_grid= { 'n_estimators':[100, 250, 500, 1000], 'max_depth':[5, 10, 25,50 ], 
            'min_samples_split':[2,4,8,16,32,64,128,256], 'min_samples_leaf':[2,4,8,16,32,64,128,256]
                } 


            #setup grid search on the train+val data.
            kfold = KFold(n_splits=5, shuffle=True)
            grid_search = GridSearchCV(model, param_grid, cv=kfold,  scoring='neg_brier_score', n_jobs=-1, verbose=3) 
            
            
            if Params.label_type=='hard':
                grid_search.fit(TrainFeatures, TrainLabels-1,  sample_weight=TrainWeights)
            else:           
                grid_search.fit(TrainFeatures, TrainLabels,  sample_weight=TrainWeights)


            print("The best parameters recovered: ", grid_search.best_params_)
            print("Cross-validation score: ", grid_search.best_score_)

            #Keep the best model recovered from the Grid search
            model=grid_search.best_estimator_


        print("Fitting classifier model with [%i x %i] features"%TrainFeatures.shape)
        if Params.label_type=='hard':
            model.fit(TrainFeatures, TrainLabels-1, sample_weight=TrainWeights)
        else:
            model.fit(TrainFeatures, TrainLabels, sample_weight=TrainWeights) 

        # save model to file
        joblib.dump(model, model_filename)
 

 

    # if Params.label_type=='hard':
    #     y_prob=model.predict_proba(TrainFeatures)
    #     BS=brier_score_loss(TrainLabels, y_prob[:,0], pos_label=1)
    #     MS=model.score(TrainFeatures, TrainLabels-1)
    # else:

    #     y=model.predict(TrainFeatures)
    #     TL=(TrainLabels<0).astype(int) 
    #     Ty=(y<0).astype(int) 
    #     MS= np.sum(Ty==TL)/len(Ty)
        
    #     Spread2Prob_Convert_v = np.vectorize(Spread2Prob_Convert)
    #     y_conv= Spread2Prob_Convert_v (y, [9.96240982e-01,  1.34591168e-12, -1.34107120e+00, -6.40128871e-01])
    #     y_prob = np.array([y_conv, 1-y_conv]).T

    #     BS=brier_score_loss(TL, y_prob[:,0], pos_label=1)
    
    #print('Train acc. = %.4f, Train Brier = %.4f'%(MS,BS))
   
    return model


    
def SortGamesByDate(gamesList):
    """fix the sorting of the games list by date since: 1)stats.nba.com might have inconsistent  dates+game_ids  and 2)sqlite does not support DATE type"""
    
    #first convert date strings to datetime objects
    datetimes_list=[datetime.strptime(game[6], "%m/%d/%y").date() for game in gamesList]

    #now sort the array and return the index
    sorted_index = np.argsort(datetimes_list)

    return sorted_index
 


def backTestSeason(Params, TrainFeatures, TrainLabels, TrainWeights, CM, Market_history_file, FS, verbose=False):   
 
    #get database file associated with current (BACKTESTED) season
    test_season = str(Params.testSeason-1)+"-"+str(Params.testSeason)[2:4]
    databaseFile = Params.databaseDir+"/NBA_season"+test_season+".db"
    if  os.path.isfile(databaseFile):
        dbms = open_database(databaseFile)   #open the database file
        cursor= dbms.conn.cursor()
    else:
        print("Testing database was not found. Aborting")
        sys.exit()

    
    #Load Player Prediction database as well
    player_dbms= open_database(Params.databaseDir+"/PlayerStatDatabase.db") #TODO: Pass this from Params file


    #Load Expert Predictions database
    if Params.backtestMetaEnsemble:
        Experts_dbms= open_database("./Libs/Expert_ensemble/ExpertPredictions_DB.db") #TODO: Pass this from Params file
        experts_cursor= Experts_dbms.conn.cursor()

        #Get all the Predictions, Experts & Games data and save them to Pandas df
        all_game_rows=experts_cursor.execute("SELECT game_id, Away_Score, Home_Score, Result, Season from GAMES").fetchall()
        col_names= np.array(experts_cursor.description)
        games_DF = pd.DataFrame(all_game_rows, columns=col_names[:,0])

        all_pred_rows=experts_cursor.execute("SELECT * from Predictions").fetchall()
        col_names= np.array(experts_cursor.description)
        predictions_DF = pd.DataFrame(all_pred_rows, columns=col_names[:,0])

        all_expert_rows=experts_cursor.execute("SELECT * from Experts").fetchall()
        col_names= np.array(experts_cursor.description)
        experts_DF = pd.DataFrame(all_expert_rows, columns=col_names[:,0])
        expert_ids = experts_DF['expert_id'].values #these are already unique by definition of the database

        #Train ensemble prediction method from past data (but ignoring current season so far)
        past_games_DF= games_DF[games_DF['Season'] != "2020"]
        past_predictions_DF= pd.merge(past_games_DF, predictions_DF, on=['game_id'], how='inner')[['game_id','expert_id','Prob_Away','Prob_Home','Success']]

        print("Extracting ensemble data...")
        X_train, y_train, _, weights_train = extract_data(past_games_DF, expert_ids, past_predictions_DF) 
        
        _, _, x_opt = GlobalBrier_optimiser_method(X_train, y_train) #GLOBAL OPTIMISER

        
        Experts_dbms.closeConnection()


    #Create temporary backtest database
    temp_Database_file='temp/tempBackTesting_Database.db'
    copyfile(databaseFile, temp_Database_file)  
    temp_dbms =  open_database(temp_Database_file)     
    temp_cursor= temp_dbms.conn.cursor()   
    tbl_list = temp_cursor.execute('SELECT name from sqlite_master where type= "table"').fetchall()  
    for i in range(len(tbl_list)):  
        tbl_name=tbl_list[i][0]     
        if tbl_name != 'Teams':   
            temp_cursor.execute('DELETE FROM ' +tbl_name)   
    temp_cursor.connection.commit()   




    #Create empty pandas dataframe to store temporary prediction data, so that it can be saved into the excel file in the end after the backtesting
    dfObj = pd.DataFrame(columns=['Away', 'Home', 'Date', 'Result', 'Our prediction', 'Our probs'])
    
    #Create empty list to store game indices so that the games may be inserted to the temp database in one go after every day (much faster than inserting to database after every game)
    gm_indx_list=[]


    #Fit the classifier
    CM=trainClassifier(Params, CM, TrainFeatures, TrainLabels, TrainWeights )

    
    g_rws = cursor.execute('SELECT * FROM games').fetchall() #Get all the games in the test_season ground truth database. 
    sorted_games_index= SortGamesByDate(g_rws)


    print("Backtesting test season %s. %i games found." %(test_season,len(g_rws)))


    if verbose:
        #TODO: Print Brier as well as market accuracy and brier
        plt.figure()
        ax = plt.axes()
        ax2 = ax.twinx()  # instantiate a second axes that shares the same x-axis

        ax.set_ylabel('Accuracy')
        ax2.set_ylabel('Brier')
        ax.set_xlabel('Game')
        ax.set_title("Backesting season: "+test_season)
        ax.set_ylim([0.45,0.85])
        ax.grid(True)
        ax.tick_params(axis='y', labelcolor='blue')



    #loop over games
    CompArray=[]
    ProbsArray=[]
    AccuArray=[]
    BrierArray=[]

 

    ###################################################################         MAIN BACKTESTING LOOP       ##############################################################    
    old_timestamp=0
    np.random.seed(1000)
    new_day_flag=False
    day_counter=1
    for g_row in range(len(g_rws)):
        
        #Get game info
        game_id= g_rws[sorted_games_index[g_row]][0]
        awayTeam_ID=g_rws[sorted_games_index[g_row]][1]
        homeTeam_ID=g_rws[sorted_games_index[g_row]][2]
        TestLabel = int(g_rws[sorted_games_index[g_row]][5])
        game_date=g_rws[sorted_games_index[g_row]][6]



        dt=datetime.strptime(game_date, "%m/%d/%y")
        cur_timestamp= datetime.timestamp(dt)

        #Track changes to date
        if old_timestamp==0:
            old_timestamp = cur_timestamp
        else:
            if old_timestamp != cur_timestamp:
                new_day_flag=True
                old_timestamp=cur_timestamp
            else:
                new_day_flag=False

            
        #Retrain classifier here by incorporating the ACTUAL features after a new day
        if new_day_flag:
 
            #Add games collected in temporary list to temporary database
            addToTempDatabase(gm_indx_list, temp_dbms, dbms) 
            #reset game index list
            gm_indx_list =[]


            if Params.live_retrain:
                #generate the features from the collected games so far
                observed_Features, observed_Labels = generate_Features_fromDB(Params, temp_Database_file)   
                observed_Weights = 2*np.ones_like(observed_Labels) #maximum weights of 2

                #append to training features & labels
                updatedFeatures = np.vstack([TrainFeatures, observed_Features])
                updatedLabels = np.append(TrainLabels, observed_Labels)
                updatedWeights = np.append(TrainWeights, observed_Weights)

                #retrain the classifier
                CM=trainClassifier(Params, CM, updatedFeatures, updatedLabels, updatedWeights )


            print("\tDAY %i (Burn-in: %i days)." %(day_counter, Params.burn_in))    
            day_counter=day_counter+1

 




        awObj = getTeam_by_ID(awayTeam_ID)
        hmObj = getTeam_by_ID(homeTeam_ID)
        awayTeam_name = awObj.name
        awayTeam_short = awObj.short
        homeTeam_name = hmObj.name
        homeTeam_short = hmObj.short



        #Predict expected lineups them or scrape team roster
        awayTeam_lineup=PredictLineup(Params, awayTeam_ID, '', temp_dbms, player_dbms) 
        homeTeam_lineup=PredictLineup(Params, homeTeam_ID, '', temp_dbms, player_dbms) 
        
        
        print("Game %i (%s):  %s vs %s \t\t\t" % (g_row+1,game_id, awayTeam_name,homeTeam_name))


        #Predict features for the  game
        Feats_Away=Predict_Features_forLineup(Params, awayTeam_lineup,  temp_cursor)
        Feats_Home=Predict_Features_forLineup(Params, homeTeam_lineup,  temp_cursor)
        TestFeatures=np.hstack([Feats_Away, Feats_Home]) 
        
        
        
        if Params.reverse_home_advantage:
            ReverseTestFeatures=np.hstack([Feats_Home, Feats_Away]) 


        if Params.add_gameno_as_feature:   
            TestFeatures=np.hstack([TestFeatures,  (g_row+1)*np.ones((Params.MCruns,1)) ])  #Add the game number as a feature
            if Params.reverse_home_advantage:
                ReverseTestFeatures=np.hstack([ReverseTestFeatures,  (g_row+1)*np.ones((Params.MCruns,1)) ])  #Add the game number as a feature
        
        if TestFeatures.shape[0]==1:
            TestFeatures=TestFeatures.reshape(1,-1)
            if Params.reverse_home_advantage:
                ReverseTestFeatures=ReverseTestFeatures.reshape(1,-1)


        if Params.use_feature_selector:
            TestFeatures = FS.transform(TestFeatures)
            if Params.reverse_home_advantage:
                ReverseTestFeatures = FS.transform(ReverseTestFeatures)
        
        if Params.feature_engineer:
            TestFeatures = engineerFeatures(TestFeatures)
            if Params.reverse_home_advantage:
                ReverseTestFeatures = engineerFeatures(ReverseTestFeatures)

        
        #Predict outcome of game
        if Params.label_type=='hard':
            out_probs = CM.predict_proba(TestFeatures) 
            out_probs= np.mean(out_probs, axis=0) #take the mean of the MC probs 

            if Params.reverse_home_advantage:
                reverse_out_probs = CM.predict_proba(ReverseTestFeatures) 
                reverse_out_probs= np.mean(reverse_out_probs, axis=0) #take the mean of the MC probs 
                out_probs = np.array([(out_probs[0]+reverse_out_probs[1])/2, (out_probs[1]+reverse_out_probs[0])/2])
            
            
            out_probs[0] = 1- out_probs[1] #make sure they sum to 1
                     


            if Params.backtestMetaEnsemble:

                #Inject our probabilities instead of what is in the database
                try:
                    df_indx= predictions_DF.index[(predictions_DF['game_id']==game_id) & (predictions_DF['expert_id']=='0004') ][0]                
                    predictions_DF.loc[df_indx,'Prob_Away']=out_probs[0]
                    predictions_DF.loc[df_indx,'Prob_Home']=out_probs[1]
                except:
                    #inject manually
                    new_row = {'game_id':game_id, 'expert_id':'0004', 'Prob_Away':out_probs[0], 'Prob_Home':out_probs[1], 'Success':1}
                    predictions_DF= predictions_DF.append(new_row, ignore_index=True)

                #obtain the expert predictions for this game
                current_game_DF = games_DF[games_DF['game_id'] == game_id]
                current_preds_DF = predictions_DF[predictions_DF['game_id']==game_id]

            
                X_test, y_test, _, _, = extract_data(current_game_DF, expert_ids, current_preds_DF) 

                out_probs[0]= (X_test @ x_opt)[0]  #the away probs
                out_probs[1] = 1- out_probs[0] #make sure they sum to 1


                #since we have the y_test label we can re-train the ensemble predictor here already
                #So append current test data to training data (+ labels) and re-train the ensemble predictor
                X_train = np.vstack([X_train, X_test])
                y_train = np.hstack([y_train, y_test])

                _, _, x_opt = GlobalBrier_optimiser_method(X_train, y_train) #GLOBAL OPTIMISER

          

            p_test = np.argmax(out_probs)+1 #Remember here that labels are 1-Away, 2-Home
            p_probs =  np.max(out_probs)

            
        else:  #SOFT LABELS  TODO: MC and Experts Ensemble

            out_regr= CM.predict(TestFeatures)
            out_probs = Spread2Prob_Convert(out_regr, [9.96240982e-01,  1.34591168e-12, -1.34107120e+00, -6.40128871e-01])

   

            if out_probs > 0.5: #HOME
                p_test= 2 #home
                p_probs = out_probs
            else:               #AWAY
                p_test=1 #away
                p_probs = 1-out_probs
            


        if TestLabel == p_test:
            CompArray.append(1)            
        else:
            CompArray.append(0)
        ProbsArray.append(p_probs)


        AccuArray.append(np.mean(CompArray))
        BrierArray.append(brier_score_loss(CompArray, ProbsArray))
        print("\tg.t. label %i, predicted label %i @ %.2f prob., Accuracy: %.4f, Brier: %.4f "%(TestLabel, p_test, p_probs, AccuArray[-1], BrierArray[-1]))

        if verbose:
            ax.plot(g_row, AccuArray[-1],'b.')  #Print running Accuracy
            ax2.plot(g_row, BrierArray[-1], 'kx')    # print running Brier score
            plt.pause(0.001)
        



        #Append to temporary dataframe
        dfObj = dfObj.append({'Away':  awayTeam_short, 'Home': homeTeam_short, 'Date' : game_date, 'Result' :  TestLabel, "Our prediction" : int(p_test), "Our probs" :  float(100*p_probs)   }, ignore_index=True)
        #Append game id to list
        gm_indx_list.append(game_id) 

          

    if verbose:
        plt.savefig("Backtest.png")
        plt.show()
        



    #WRITE TO EXCEL FILE

    if Market_history_file is not None:
        market_workbook = openpyxl.load_workbook(Market_history_file)
        worksheet = market_workbook.worksheets[0]
 
        #find game entry in worksheet
        for ws_g in range(worksheet.max_row-2):
            
            ws_away_short = str(worksheet.cell(2+ws_g+1, 0+1).value)  #cell indices start from 1
            ws_home_short = str(worksheet.cell(2+ws_g+1, 1+1).value)  #cell indices start from 1
            ws_date = str(worksheet.cell(2+ws_g+1, 2+1).value)        #cell indices start from 1
            ws_result = int(worksheet.cell(2+ws_g+1, 3+1).value)      #cell indices start from 1


            #Check game against stored dataframe
            query_row = dfObj.loc[(dfObj['Away'] == ws_away_short) & (dfObj['Home'] == ws_home_short) &  (dfObj['Date'] == ws_date) &  (dfObj['Result'] == ws_result)]
            
            if not query_row.empty:
                worksheet.cell(2+ws_g+1, 4+1 ).value= query_row.iloc[0]['Our prediction'] #Prediction 1/2 (Away/Home). cell indices start from 1
                worksheet.cell(2+ws_g+1, 5+1 ).value= query_row.iloc[0]['Our probs']  #Probability  [50-100]. cell indices start from 1

                  
        market_workbook.save(Market_history_file)  



    #Delete temporary backtesting database here
    temp_dbms.closeConnection()
    dbms.closeConnection()
    player_dbms.closeConnection()
    os.remove('temp/tempBackTesting_Database.db')  