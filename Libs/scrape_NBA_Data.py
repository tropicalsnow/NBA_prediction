import numpy as np
import requests
from requests.exceptions import Timeout
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.retry import Retry

import json  
import openpyxl
import time
import pandas as pd
import math
from datetime import datetime
import sys
from bs4 import BeautifulSoup, NavigableString, Tag
import os

sys.path.append('../')

import TeamsList



from seasonDB_utils import create_season_Database, open_database



from captureInjuryList import captureInjuryList

    
from captureBetfairOdds import captureBetfairOdds


headers={
    'Host': 'stats.nba.com',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:68.0) Gecko/20100101 Firefox/68.0',
    'Accept': 'application/json, text/plain, */*',
    'Accept-Language': 'en-US,en;q=0.5',
    'Accept-Encoding': 'gzip, deflate, br',
    'x-nba-stats-origin': 'stats',
    'x-nba-stats-token': 'true',
    'DNT': '1',
    'Connection':'keep-alive',
    'Referer': 'https://stats.nba.com/',
    'Pragma': 'no-cache',
    'Cache-Control': 'no-cache'
}


headers_ALT = {
    'Accept-Encoding': 'gzip, deflate, sdch',
    'Accept-Language': 'en-US,en;q=0.8',
    'Upgrade-Insecure-Requests': '1',
    'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Cache-Control': 'max-age=0',
    'Connection': 'keep-alive',
}

 


mytimeout = (20, 30)  # the program will wait X seconds to establish a connection, and if this successful, it will wait for another Y for the server response.

retry_strategy = Retry(
    total=15,
    status_forcelist=[413, 429, 503],
    method_whitelist=["HEAD", "GET", "PUT", "DELETE", "OPTIONS", "TRACE"],
    backoff_factor=3
)
    

NBA_web_session = requests.Session() 
adapter = HTTPAdapter(max_retries=retry_strategy)
# Mount it for both http and https usage
NBA_web_session.mount("https://", adapter)
NBA_web_session.mount("http://", adapter)
NBA_web_session.header = headers



overide_games_list=["0021201214", "0021900707"] #MANUAL OVERRIDE FOR MISSING GAME BETWEEN IND-BOS in 2012-13 and LAC-LAL 2019-20
override_player_ids= {"78412": "919"}  #override any player ids that might be wrongly recorder in stats.nba.com where an alternative one exists




 


#Scrape player & team BOXSCORE STATS data for given game/season and save them to database
def process_BOXSCORE_stats(season,GameID,date_string,cursor, override_player_ids):

    #TRADITIONAL STATS
    traditional_url='https://stats.nba.com/stats/boxscoretraditionalv2?EndPeriod=10&EndRange=28800&GameID='+GameID+\
    '&RangeType=0&Season='+season+'&SeasonType=Regular+Season&StartPeriod=1&StartRange=0' 
    response_TRADITIONAL =   requests.get(traditional_url, headers=headers)

    #ADVANCED STATS
    advanced_url='https://stats.nba.com/stats/boxscoreadvancedv2?EndPeriod=10&EndRange=28800&GameID='+GameID+\
    '&RangeType=0&Season='+season+'&SeasonType=Regular+Season&StartPeriod=1&StartRange=0' 
    response_ADVANCED = requests.get(advanced_url, headers=headers)  
    

    #MISC STATS
    misc_url='https://stats.nba.com/stats/boxscoremiscv2?EndPeriod=10&EndRange=28800&GameID='+GameID+\
    '&RangeType=0&Season='+season+'&SeasonType=Regular+Season&StartPeriod=1&StartRange=0' 
    response_MISC = requests.get(misc_url, headers=headers)     
    

    #SCORING STATS
    scoring_url='https://stats.nba.com/stats/boxscorescoringv2?EndPeriod=10&EndRange=28800&GameID='+GameID+\
    '&RangeType=0&Season='+season+'&SeasonType=Regular+Season&StartPeriod=1&StartRange=0' 
    response_SCORING = requests.get(scoring_url, headers=headers)     
    

    #USAGE STATS
    usage_url='https://stats.nba.com/stats/boxscoreusagev2?EndPeriod=10&EndRange=28800&GameID='+GameID+\
    '&RangeType=0&Season='+season+'&SeasonType=Regular+Season&StartPeriod=1&StartRange=0' 
    response_USAGE = requests.get(usage_url, headers=headers) 
    
    
    #FOUR FACTORS
    fourfactors_url='https://stats.nba.com/stats/boxscorefourfactorsv2?EndPeriod=10&EndRange=28800&GameID='+GameID+\
    '&RangeType=0&Season='+season+'&SeasonType=Regular+Season&StartPeriod=1&StartRange=0' 
    response_FOURFACTORS = requests.get(fourfactors_url, headers=headers)  
    

    #TRACKING
    tracking_url='https://stats.nba.com/stats/boxscoreplayertrackv2?EndPeriod=10&EndRange=28800&GameID='+GameID+\
    '&RangeType=0&Season='+season+'&SeasonType=Regular+Season&StartPeriod=1&StartRange=0'
    response_TRACKING = requests.get(tracking_url, headers=headers)    
    

    #HUSTLE    
    hustle_url='https://stats.nba.com/stats/hustlestatsboxscore?GameID='+GameID
    response_HUSTLE = requests.get(hustle_url, headers=headers)      
    



    if response_TRADITIONAL.status_code==200: 
        boxscore_traditional = json.loads(response_TRADITIONAL.text)
        Playerstats_TRADITIONAL = boxscore_traditional['resultSets'][0]['rowSet']


    if response_ADVANCED.status_code==200: 
        boxscore_advanced = json.loads(response_ADVANCED.text)
        Playerstats_ADVANCED = boxscore_advanced['resultSets'][0]['rowSet']


    if response_MISC.status_code==200: 
        boxscore_misc = json.loads(response_MISC.text)
        Playerstats_MISC = boxscore_misc['resultSets'][0]['rowSet']


    if response_SCORING.status_code==200: 
        boxscore_scoring = json.loads(response_SCORING.text)
        Playerstats_SCORING = boxscore_scoring['resultSets'][0]['rowSet']


    if response_USAGE.status_code==200: 
        boxscore_USAGE = json.loads(response_USAGE.text)
        Playerstats_USAGE = boxscore_USAGE['resultSets'][0]['rowSet']
        

    if response_FOURFACTORS.status_code==200: 
        boxscore_fourfactors = json.loads(response_FOURFACTORS.text)
        Playerstats_FOURFACTORS = boxscore_fourfactors['resultSets'][0]['rowSet']


    if response_TRACKING.status_code==200: 
        boxscore_tracking = json.loads(response_TRACKING.text)
        Playerstats_TRACKING = boxscore_tracking['resultSets'][0]['rowSet']

       
    if response_HUSTLE.status_code==200: 
        boxscore_hustle = json.loads(response_HUSTLE.text)
        Playerstats_HUSTLE = boxscore_hustle['resultSets'][1]['rowSet']
        
 



    #Insert to BoxscoresPlayer AND Lineups table
    for p in range(len(Playerstats_TRADITIONAL)):  #loop for all players in the game


        player_id = Playerstats_TRADITIONAL[p][4]
        #check if player_id needs to be overriden
        if player_id in override_player_ids:
            player_id = override_player_ids[player_id]


        if  Playerstats_TRADITIONAL[p][7] != '':
            pass #skip players who did not play
        else:


            

            #Check for hustle stats. Older games do not have hustle stats and new games follow active players indexing (so not all players have HUSTLE stats)

            CONTESTED_SHOTS=0; CONTESTED_SHOTS_2PT=0; CONTESTED_SHOTS_3PT=0; DEFLECTIONS=0
            CHARGES_DRAWN=0; SCREEN_ASSISTS=0; SCREEN_AST_PTS=0; OFF_LOOSE_BALLS_RECOVERED=0
            DEF_LOOSE_BALLS_RECOVERED=0; LOOSE_BALLS_RECOVERED=0; OFF_BOXOUTS= 0; DEF_BOXOUTS= 0
            BOX_OUT_PLAYER_TEAM_REBS= 0; BOX_OUT_PLAYER_REBS= 0; BOX_OUTS=  0

            for hp in range(len(Playerstats_HUSTLE)):



                #check if traditional player id matches hustle player id, because hustle stats may not be available for everyone or in order
                if  player_id  ==  Playerstats_HUSTLE[hp][4]: 

                    CONTESTED_SHOTS     =            float(Playerstats_HUSTLE[hp][10] or 0)                     #CONTESTED_SHOTS
                    CONTESTED_SHOTS_2PT =            float(Playerstats_HUSTLE[hp][11] or 0)                     #CONTESTED_SHOTS_2PT
                    CONTESTED_SHOTS_3PT =            float(Playerstats_HUSTLE[hp][12] or 0)                     #CONTESTED_SHOTS_3PT
                    DEFLECTIONS =                    float(Playerstats_HUSTLE[hp][13] or 0)                     #DEFLECTIONS
                    CHARGES_DRAWN=                   float(Playerstats_HUSTLE[hp][14] or 0)                     #CHARGES_DRAWN
                    SCREEN_ASSISTS=                  float(Playerstats_HUSTLE[hp][15] or 0)                     #SCREEN_ASSISTS
                    SCREEN_AST_PTS=                  float(Playerstats_HUSTLE[hp][16] or 0)                     #SCREEN_AST_PTS
                    OFF_LOOSE_BALLS_RECOVERED=       float(Playerstats_HUSTLE[hp][17] or 0)                     #OFF_LOOSE_BALLS_RECOVERED
                    DEF_LOOSE_BALLS_RECOVERED=       float(Playerstats_HUSTLE[hp][18] or 0)                     #DEF_LOOSE_BALLS_RECOVERED
                    LOOSE_BALLS_RECOVERED=           float(Playerstats_HUSTLE[hp][19] or 0)                     #LOOSE_BALLS_RECOVERED
                    OFF_BOXOUTS=                     float(Playerstats_HUSTLE[hp][20] or 0)                     #OFF_BOXOUTS
                    DEF_BOXOUTS=                     float(Playerstats_HUSTLE[hp][21] or 0)                     #DEF_BOXOUTS
                    BOX_OUT_PLAYER_TEAM_REBS=        float(Playerstats_HUSTLE[hp][22] or 0)                     #BOX_OUT_PLAYER_TEAM_REBS
                    BOX_OUT_PLAYER_REBS=             float(Playerstats_HUSTLE[hp][23] or 0)                     #BOX_OUT_PLAYER_REBS
                    BOX_OUTS=                        float(Playerstats_HUSTLE[hp][24] or 0)                     #BOX_OUTS

                    break #once matching player_ids found break the for loop



            #check for availability of ADVANCED stats
            if len(Playerstats_ADVANCED)==0:
                E_OFF_RATING=0; OFF_RATING=0; E_DEF_RATING=0; DEF_RATING=0; E_NET_RATING=0; NET_RATING=0
                AST_PCT=0; AST_TOV=0; AST_RATIO=0; OREB_PCT=0; DREB_PCT=0; REB_PCT=0; TOV_PCT=0; EFG_PCT=0
                TS_PCT=0; USG_PCT=0; E_USG_PCT=0; E_PACE=0; PACE=0; PIE=0
            else:
                E_OFF_RATING =          float(Playerstats_ADVANCED[p][9] or 0)
                OFF_RATING=             float(Playerstats_ADVANCED[p][10] or 0)
                E_DEF_RATING=           float(Playerstats_ADVANCED[p][11] or 0)
                DEF_RATING=             float(Playerstats_ADVANCED[p][12] or 0)
                E_NET_RATING=           float(Playerstats_ADVANCED[p][13] or 0)
                NET_RATING=             float(Playerstats_ADVANCED[p][14] or 0)
                AST_PCT=                float(Playerstats_ADVANCED[p][15] or 0)       
                AST_TOV=                float(Playerstats_ADVANCED[p][16] or 0)    
                AST_RATIO=              float(Playerstats_ADVANCED[p][17] or 0)
                OREB_PCT=               float(Playerstats_ADVANCED[p][18] or 0)
                DREB_PCT=               float(Playerstats_ADVANCED[p][19] or 0)
                REB_PCT=                float(Playerstats_ADVANCED[p][20] or 0)
                TOV_PCT=                float(Playerstats_ADVANCED[p][21] or 0)
                EFG_PCT=                float(Playerstats_ADVANCED[p][22] or 0) 
                TS_PCT=                 float(Playerstats_ADVANCED[p][23] or 0) 
                USG_PCT=                float(Playerstats_ADVANCED[p][24] or 0)       
                E_USG_PCT=              float(Playerstats_ADVANCED[p][25] or 0)       
                E_PACE=                 float(Playerstats_ADVANCED[p][26] or 0)       
                PACE=                   float(Playerstats_ADVANCED[p][27] or 0)              
                PIE=                    float(Playerstats_ADVANCED[p][28] or 0)    

            #check for availability of MISC stats
            if len(Playerstats_MISC)==0:
                PTS_OFF_TOV=0; PTS_2ND_CHANCE=0; PTS_FB=0; PTS_PAINT=0; OPP_PTS_OFF_TOV=0
                OPP_PTS_2ND_CHANCE=0; OPP_PTS_FB=0; OPP_PTS_PAINT=0; BLKA=0; PFD=0

            else:
                PTS_OFF_TOV=            float(Playerstats_MISC[p][9] or 0)
                PTS_2ND_CHANCE=         float(Playerstats_MISC[p][10] or 0)
                PTS_FB=                 float(Playerstats_MISC[p][11] or 0)
                PTS_PAINT=              float(Playerstats_MISC[p][12] or 0)  
                OPP_PTS_OFF_TOV=        float(Playerstats_MISC[p][13] or 0) 
                OPP_PTS_2ND_CHANCE=     float(Playerstats_MISC[p][14] or 0)
                OPP_PTS_FB=             float(Playerstats_MISC[p][15] or 0) 
                OPP_PTS_PAINT=          float(Playerstats_MISC[p][16] or 0)
                BLKA=                   float(Playerstats_MISC[p][18] or 0)               
                PFD=                    float(Playerstats_MISC[p][20] or 0)  

            #check for availability of SCORING stats
            if len(Playerstats_SCORING)==0:
                PCT_FGA_2PT=0; PCT_FGA_3PT=0; PCT_PTS_2PT=0; PCT_PTS_2PT_MR=0; PCT_PTS_3PT=0; PCT_PTS_FB=0
                PCT_PTS_FT=0; PCT_PTS_OFF_TOV=0; PCT_PTS_PAINT=0; PCT_AST_2PM=0; PCT_UAST_2PM=0; PCT_AST_3PM=0
                PCT_UAST_3PM=0; PCT_AST_FGM=0; PCT_UAST_FGM=0
            else:
                PCT_FGA_2PT=            float(Playerstats_SCORING[p][9] or 0)  
                PCT_FGA_3PT=            float(Playerstats_SCORING[p][10] or 0)   
                PCT_PTS_2PT=            float(Playerstats_SCORING[p][11] or 0)   
                PCT_PTS_2PT_MR=         float(Playerstats_SCORING[p][12] or 0) 
                PCT_PTS_3PT=            float(Playerstats_SCORING[p][13] or 0) 
                PCT_PTS_FB=             float(Playerstats_SCORING[p][14] or 0)  
                PCT_PTS_FT=             float(Playerstats_SCORING[p][15] or 0) 
                PCT_PTS_OFF_TOV=        float(Playerstats_SCORING[p][16] or 0)  
                PCT_PTS_PAINT=          float(Playerstats_SCORING[p][17] or 0)     
                PCT_AST_2PM=            float(Playerstats_SCORING[p][18] or 0)
                PCT_UAST_2PM=           float(Playerstats_SCORING[p][19] or 0)       
                PCT_AST_3PM=            float(Playerstats_SCORING[p][20] or 0) 
                PCT_UAST_3PM=           float(Playerstats_SCORING[p][21] or 0)
                PCT_AST_FGM=            float(Playerstats_SCORING[p][22] or 0) 
                PCT_UAST_FGM=           float(Playerstats_SCORING[p][23] or 0)

            #check for availability of USAGE stats
            if len(Playerstats_USAGE)==0:
                USG_PCT_USG=0; PCT_FGM=0; PCT_FGA=0; PCT_FG3M=0; PCT_FG3A=0; PCT_FTM=0; PCT_FTA=0; PCT_OREB=0; PCT_DREB=0
                PCT_REB=0; PCT_AST=0; PCT_TOV=0; PCT_STL=0; PCT_BLK=0; PCT_BLKA=0; PCT_PF=0; PCT_PFD=0; PCT_PTS=0
            else:
                USG_PCT_USG=            float(Playerstats_USAGE[p][9] or 0)
                PCT_FGM=                float(Playerstats_USAGE[p][10] or 0)  
                PCT_FGA=                float(Playerstats_USAGE[p][11] or 0)
                PCT_FG3M=               float(Playerstats_USAGE[p][12] or 0)
                PCT_FG3A=               float(Playerstats_USAGE[p][13] or 0)
                PCT_FTM=                float(Playerstats_USAGE[p][14] or 0)
                PCT_FTA=                float(Playerstats_USAGE[p][15] or 0)
                PCT_OREB=               float(Playerstats_USAGE[p][16] or 0)
                PCT_DREB=               float(Playerstats_USAGE[p][17] or 0) 
                PCT_REB=                float(Playerstats_USAGE[p][18] or 0) 
                PCT_AST=                float(Playerstats_USAGE[p][19] or 0)
                PCT_TOV=                float(Playerstats_USAGE[p][20] or 0)
                PCT_STL=                float(Playerstats_USAGE[p][21] or 0)
                PCT_BLK=                float(Playerstats_USAGE[p][22] or 0)
                PCT_BLKA=               float(Playerstats_USAGE[p][23] or 0)
                PCT_PF=                 float(Playerstats_USAGE[p][24] or 0)  
                PCT_PFD=                float(Playerstats_USAGE[p][25] or 0)
                PCT_PTS=                float(Playerstats_USAGE[p][26] or 0)     

            #check for availability of FOURFACTORS stats
            if len(Playerstats_FOURFACTORS)==0:
                EFG_PCT_FF=0; FTA_RATE=0; TM_TOV_PCT=0; OREB_PCT_FF=0; OPP_EFG_PCT=0; OPP_FTA_RATE=0; OPP_TOV_PCT=0; OPP_OREB_PCT=0
            else:
                EFG_PCT_FF=             float(Playerstats_FOURFACTORS[p][9] or 0)
                FTA_RATE=               float(Playerstats_FOURFACTORS[p][10] or 0)  
                TM_TOV_PCT=             float(Playerstats_FOURFACTORS[p][11] or 0)
                OREB_PCT_FF=            float(Playerstats_FOURFACTORS[p][12] or 0)
                OPP_EFG_PCT=            float(Playerstats_FOURFACTORS[p][13] or 0)
                OPP_FTA_RATE=           float(Playerstats_FOURFACTORS[p][14] or 0)
                OPP_TOV_PCT=            float(Playerstats_FOURFACTORS[p][15] or 0)
                OPP_OREB_PCT=           float(Playerstats_FOURFACTORS[p][16] or 0)

            #check for availability of TRACKING stats
            if len(Playerstats_TRACKING)==0:
                SPD=0; DIST=0;  ORBC=0; DRBC=0; RBC=0; TCHS=0; SAST=0; FTAST=0; PASS=0; CFGM=0
                CFGA=0; CFG_PCT=0; UFGM=0; UFGA=0; UFG_PCT=0; DFGM=0; DFGA=0; DFG_PCT=0
            else:
                SPD=                    float(Playerstats_TRACKING[p][9] or 0)    
                DIST=                   float(Playerstats_TRACKING[p][10] or 0)
                ORBC=                   float(Playerstats_TRACKING[p][11] or 0)
                DRBC=                   float(Playerstats_TRACKING[p][12] or 0)      
                RBC=                    float(Playerstats_TRACKING[p][13] or 0) 
                TCHS=                   float(Playerstats_TRACKING[p][14] or 0)
                SAST=                   float(Playerstats_TRACKING[p][15] or 0)
                FTAST=                  float(Playerstats_TRACKING[p][16] or 0)
                PASS=                   float(Playerstats_TRACKING[p][17] or 0)   
                CFGM=                   float(Playerstats_TRACKING[p][19] or 0)
                CFGA=                   float(Playerstats_TRACKING[p][20] or 0)
                CFG_PCT=                float(Playerstats_TRACKING[p][20] or 0)
                UFGM=                   float(Playerstats_TRACKING[p][22] or 0)  
                UFGA=                   float(Playerstats_TRACKING[p][23] or 0)
                UFG_PCT=                float(Playerstats_TRACKING[p][24] or 0)
                DFGM=                   float(Playerstats_TRACKING[p][26] or 0) 
                DFGA=                   float(Playerstats_TRACKING[p][27] or 0)
                DFG_PCT=                float(Playerstats_TRACKING[p][28] or 0)





            if   Playerstats_TRADITIONAL[p][8] is None:
                MIN=0          
            else:
                if isinstance(Playerstats_TRADITIONAL[p][8], str):
                    MIN=   float((Playerstats_TRADITIONAL[p][8]).replace(":",".")) 
                else:
                    MIN=   float(Playerstats_TRADITIONAL[p][8])




            cursor.execute('INSERT OR REPLACE INTO boxscoresPlayer (game_id, player_id, MIN, \
                                            FGM, FGA, FG_PCT, FG3M, FG3A, FG3_PCT, FTM, FTA, FT_PCT, OREB, DREB, REB, AST, STL, BLK, TOV, PF, PTS, PLUS_MINUS, \
                                            E_OFF_RATING, OFF_RATING, E_DEF_RATING, DEF_RATING, E_NET_RATING, NET_RATING,  AST_PCT, AST_TOV, AST_RATIO, OREB_PCT, DREB_PCT, REB_PCT, TOV_PCT, EFG_PCT, TS_PCT, USG_PCT, E_USG_PCT, E_PACE, PACE, PIE, \
                                            PTS_OFF_TOV,  PTS_2ND_CHANCE, PTS_FB, PTS_PAINT, OPP_PTS_OFF_TOV, OPP_PTS_2ND_CHANCE, OPP_PTS_FB, OPP_PTS_PAINT, BLKA, PFD, \
                                            PCT_FGA_2PT, PCT_FGA_3PT, PCT_PTS_2PT, PCT_PTS_2PT_MR, PCT_PTS_3PT, PCT_PTS_FB, PCT_PTS_FT, PCT_PTS_OFF_TOV, PCT_PTS_PAINT, PCT_AST_2PM, PCT_UAST_2PM, PCT_AST_3PM, PCT_UAST_3PM, PCT_AST_FGM, PCT_UAST_FGM,\
                                            USG_PCT_USG, PCT_FGM, PCT_FGA, PCT_FG3M, PCT_FG3A, PCT_FTM, PCT_FTA, PCT_OREB, PCT_DREB, PCT_REB, PCT_AST, PCT_TOV, PCT_STL, PCT_BLK, PCT_BLKA, PCT_PF, PCT_PFD, PCT_PTS, \
                                            EFG_PCT_FF, FTA_RATE, TM_TOV_PCT, OREB_PCT_FF, OPP_EFG_PCT, OPP_FTA_RATE, OPP_TOV_PCT, OPP_OREB_PCT, \
                                            SPD, DIST, ORBC, DRBC, RBC, TCHS, SAST, FTAST, PASS, CFGM, CFGA, CFG_PCT, UFGM, UFGA, UFG_PCT,  DFGM, DFGA, DFG_PCT, \
                                            CONTESTED_SHOTS, CONTESTED_SHOTS_2PT, CONTESTED_SHOTS_3PT, DEFLECTIONS, CHARGES_DRAWN, SCREEN_ASSISTS, SCREEN_AST_PTS, OFF_LOOSE_BALLS_RECOVERED, DEF_LOOSE_BALLS_RECOVERED, LOOSE_BALLS_RECOVERED, OFF_BOXOUTS, DEF_BOXOUTS, BOX_OUT_PLAYER_TEAM_REBS, BOX_OUT_PLAYER_REBS,  BOX_OUTS)' 
                                            'VALUES(?,?,?, \
                                            ?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,\
                                            ?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?, \
                                            ?,?,?,?,?,?,?,?,?,?,\
                                            ?,?,?,?,?,?,?,?,?,?,?,?,?,?,?, \
                                            ?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?, \
                                            ?,?,?,?,?,?,?,?, \
                                            ?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?, \
                                            ?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',
                                        (GameID,                                                   #game_id
                                        player_id,                                                 #player_id
                                        MIN,                                                       #MIN
                                        #TRADITIONAL  19
                                        float(Playerstats_TRADITIONAL[p][9] or 0),                 #FGM
                                        float(Playerstats_TRADITIONAL[p][10] or 0),                #FGA
                                        float(Playerstats_TRADITIONAL[p][11] or 0),                #FG_PCT
                                        float(Playerstats_TRADITIONAL[p][12] or 0),                #FG3M
                                        float(Playerstats_TRADITIONAL[p][13] or 0),                #FG3A
                                        float(Playerstats_TRADITIONAL[p][14] or 0),                #FG3_PCT
                                        float(Playerstats_TRADITIONAL[p][15] or 0),                #FTM
                                        float(Playerstats_TRADITIONAL[p][16] or 0),                #FTA
                                        float(Playerstats_TRADITIONAL[p][17] or 0),                #FT_PCT
                                        float(Playerstats_TRADITIONAL[p][18] or 0),                #OREB
                                        float(Playerstats_TRADITIONAL[p][19] or 0),                #DREB
                                        float(Playerstats_TRADITIONAL[p][20] or 0),                #REB
                                        float(Playerstats_TRADITIONAL[p][21] or 0),                #AST
                                        float(Playerstats_TRADITIONAL[p][22] or 0),                #STL
                                        float(Playerstats_TRADITIONAL[p][23] or 0),                #BLK
                                        float(Playerstats_TRADITIONAL[p][24] or 0),                #TOV
                                        float(Playerstats_TRADITIONAL[p][25] or 0),                #PF
                                        float(Playerstats_TRADITIONAL[p][26] or 0),                #PTS
                                        float(Playerstats_TRADITIONAL[p][27] or 0),                #PLUS_MINUS
                                        #ADVANCED      20                                                          
                                        E_OFF_RATING, OFF_RATING, E_DEF_RATING, DEF_RATING, E_NET_RATING,                 
                                        NET_RATING, AST_PCT, AST_TOV, AST_RATIO, OREB_PCT, DREB_PCT, REB_PCT,
                                        TOV_PCT, EFG_PCT, TS_PCT, USG_PCT, E_USG_PCT, E_PACE, PACE, PIE,
                                        #MISC  10
                                        PTS_OFF_TOV, PTS_2ND_CHANCE, PTS_FB, PTS_PAINT, OPP_PTS_OFF_TOV, 
                                        OPP_PTS_2ND_CHANCE, OPP_PTS_FB, OPP_PTS_PAINT, BLKA, PFD,
                                        #SCORING  15
                                        PCT_FGA_2PT, PCT_FGA_3PT, PCT_PTS_2PT, PCT_PTS_2PT_MR, PCT_PTS_3PT, PCT_PTS_FB,
                                        PCT_PTS_FT, PCT_PTS_OFF_TOV, PCT_PTS_PAINT, PCT_AST_2PM, PCT_UAST_2PM, PCT_AST_3PM,
                                        PCT_UAST_3PM, PCT_AST_FGM, PCT_UAST_FGM,
                                        #USAGE 18
                                        USG_PCT_USG, PCT_FGM, PCT_FGA, PCT_FG3M, PCT_FG3A, PCT_FTM, PCT_FTA, PCT_OREB, PCT_DREB,
                                        PCT_REB, PCT_AST, PCT_TOV, PCT_STL, PCT_BLK, PCT_BLKA, PCT_PF, PCT_PFD, PCT_PTS,
                                        #FOUR FACTORS  8
                                        EFG_PCT_FF, FTA_RATE, TM_TOV_PCT, OREB_PCT_FF, OPP_EFG_PCT, OPP_FTA_RATE, OPP_TOV_PCT, OPP_OREB_PCT,
                                        #TRACKING   18
                                        SPD, DIST, ORBC, DRBC, RBC, TCHS, SAST, FTAST, PASS, CFGM, CFGA, CFG_PCT, UFGM, UFGA, UFG_PCT,
                                        DFGM, DFGA, DFG_PCT,
                                        #HUSTLE  15
                                        CONTESTED_SHOTS, CONTESTED_SHOTS_2PT, CONTESTED_SHOTS_3PT, DEFLECTIONS, CHARGES_DRAWN, SCREEN_ASSISTS,
                                        SCREEN_AST_PTS, OFF_LOOSE_BALLS_RECOVERED, DEF_LOOSE_BALLS_RECOVERED, LOOSE_BALLS_RECOVERED, OFF_BOXOUTS,
                                        DEF_BOXOUTS, BOX_OUT_PLAYER_TEAM_REBS, BOX_OUT_PLAYER_REBS, BOX_OUTS                                          
                                    )
            )

            
            cursor.execute('INSERT OR REPLACE  INTO lineups (player_id, game_id, team_id)'
                                        'VALUES(?,?,?)',
                                        (player_id,                                        #player_id
                                        GameID,                                            #game_id
                                        Playerstats_TRADITIONAL[p][1]                      #team_id
                                        )
            )


 




#Scrape the average (per game) stats of a player at a given season and save them to database or return them.
#This is only used for older seasons where players may have problematic Boxscore stats
def process_STATS(stats_cursor, player_id, season_id, season_type='Regular'):
    
    TRADITIONAL_DATA_ORDERED=np.zeros(20)
    ADVANCED_DATA_ORDERED=np.zeros(20)
    MISC_DATA_ORDERED=np.zeros(10)
    SCORING_DATA_ORDERED=  np.zeros(15)
    USAGE_DATA_ORDERED= np.zeros(18)

    NBA_web_session.headers = headers
    
    type='Base'

    splits_url = "https://stats.nba.com/stats/playerdashboardbygeneralsplits?DateFrom=&DateTo=&GameSegment=&LastNGames=0&LeagueID=00&Location=&MeasureType="+type+\
                    "&Month=0&OpponentTeamID=0&Outcome=&PORound=0&PaceAdjust=N&PerMode=PerGame&Period=0&PlayerID="+str(player_id)+\
                    "&PlusMinus=N&Rank=N&Season="+season_id+\
                    "&SeasonSegment=&SeasonType="+season_type+\
                    "+Season&ShotClockRange=&Split=general&VsConference=&VsDivision="



    career_url="https://stats.nba.com/stats/playercareerstats?LeagueID=00&PerMode=PerGame&PlayerID="+str(player_id)




    #first check if the player has a splits URL at all. Really old players do not
    try:
        time.sleep(3)
        #response_splits_TRADITIONAL = NBA_web_session.get(splits_url, timeout=mytimeout)
        response_splits_TRADITIONAL = requests.get(splits_url, headers=headers)
        JSON_splits_traditional = json.loads(response_splits_TRADITIONAL.text)
        TRADITIONAL_split_stats = np.array(JSON_splits_traditional['resultSets'][0]['rowSet'])
    except Timeout as ex:
        sys.exit('Traditional splits request failed: ', ex)
       

    if TRADITIONAL_split_stats.size==0:
        #no SPLITS data then just capture TRADITIONAL STATS only from CAREER tab

        try:
            time.sleep(3)
            #response_career_TRADITIONAL = NBA_web_session.get(career_url, timeout=mytimeout)
            response_career_TRADITIONAL = requests.get(career_url, headers=headers)
            JSON_career_traditional = json.loads(response_career_TRADITIONAL.text)
            TRADITIONAL_career_stats = np.array(JSON_career_traditional['resultSets'][0]['rowSet'])
        except Timeout as ex:
            sys.exit('Traditional career request failed: ', ex) 
    
        
        #check if season_id is unique (i.e. multiple teams within a year)
        try:
            indx=np.where(TRADITIONAL_career_stats[:,1]==season_id)[0]   
        except:
            indx=[]
        
        if len(indx)==1:
            data=TRADITIONAL_career_stats[indx][0]
        elif len(indx)>1:
            for i in indx:
                #we have multiple years. So only consider the TOTAL (i.e. where the TEAM short name is 'TOT' and team_id = 0)
                if int(TRADITIONAL_career_stats[i,3])== 0 and TRADITIONAL_career_stats[i,4] =='TOT':
                    data=TRADITIONAL_career_stats[i]
                    break
        else: #for some reason the Career stats and Boxscore stats are inconsisent for the player. So skip
            return TRADITIONAL_DATA_ORDERED, ADVANCED_DATA_ORDERED, MISC_DATA_ORDERED, SCORING_DATA_ORDERED, USAGE_DATA_ORDERED
        

        if stats_cursor != None:
            stats_cursor.execute('UPDATE Players SET  \
                MIN= ?, FGM= ?, FGA= ?,  FG_PCT= ?, FG3M= ?, FG3A= ?, FG3_PCT= ?, FTM= ?, FTA= ?, FT_PCT= ?, OREB= ?, DREB= ?, REB= ?, AST= ?, STL= ?,  BLK= ?, TOV= ?, PF= ?, PTS= ? \
                WHERE player_id = ?',\
                (
                float(data[8] or 0),        #MIN
                float(data[9] or 0),        #FGM
                float(data[10] or 0),       #FGA
                float(data[11] or 0),       #FG_PCT
                float(data[12] or 0),       #FG3M
                float(data[13] or 0),       #FG3A
                float(data[14] or 0),       #FG3_PCT
                float(data[15] or 0),       #FTM
                float(data[16] or 0),       #FTA
                float(data[17] or 0),       #FT_PCT
                float(data[18] or 0),       #OREB
                float(data[19] or 0),       #DREB
                float(data[20] or 0),       #REB
                float(data[21] or 0),       #AST
                float(data[22] or 0),       #STL           
                float(data[23] or 0),       #BLK
                float(data[24] or 0),       #TOV
                float(data[25] or 0),        #PF
                float(data[26] or 0),        #PTS   
                
                player_id
                ))      

            
    else:
        #grab all stats from here. Note that TRADITIONAL stats from the splits page have additional data
        if TRADITIONAL_split_stats[0][0]=='Overall':
            data=TRADITIONAL_split_stats[0]

            if stats_cursor == None: #save the data and return it in the end
                TRADITIONAL_DATA_ORDERED=np.array([
                    float(data[6] or 0),        #MIN
                    float(data[7] or 0),        #FGM
                    float(data[8] or 0),       #FGA
                    float(data[9] or 0),       #FG_PCT
                    float(data[10] or 0),       #FG3M
                    float(data[11] or 0),       #FG3A
                    float(data[12] or 0),       #FG3_PCT
                    float(data[13] or 0),       #FTM
                    float(data[14] or 0),       #FTA
                    float(data[15] or 0),       #FT_PCT
                    float(data[16] or 0),       #OREB
                    float(data[17] or 0),       #DREB
                    float(data[18] or 0),       #REB
                    float(data[19] or 0),       #AST
                    float(data[21] or 0),       #STL           
                    float(data[22] or 0),       #BLK
                    float(data[20] or 0),       #TOV
                    float(data[24] or 0),       #PF
                    float(data[26] or 0),       #PTS   
                    float(data[27] or 0)        #PLUS_MINUS   
                    ])


            else: #add it to database   
                stats_cursor.execute('UPDATE Players SET  \
                MIN= ?, FGM= ?, FGA= ?,  FG_PCT= ?, FG3M= ?, FG3A= ?, FG3_PCT= ?, FTM= ?, FTA= ?, FT_PCT= ?, OREB= ?, DREB= ?, REB= ?, AST= ?, STL= ?,  BLK= ?, TOV= ?, PF= ?, PTS= ?, PLUS_MINUS=?\
                WHERE player_id = ?',\
                (
                float(data[6] or 0),        #MIN
                float(data[7] or 0),        #FGM
                float(data[8] or 0),       #FGA
                float(data[9] or 0),       #FG_PCT
                float(data[10] or 0),       #FG3M
                float(data[11] or 0),       #FG3A
                float(data[12] or 0),       #FG3_PCT
                float(data[13] or 0),       #FTM
                float(data[14] or 0),       #FTA
                float(data[15] or 0),       #FT_PCT
                float(data[16] or 0),       #OREB
                float(data[17] or 0),       #DREB
                float(data[18] or 0),       #REB
                float(data[19] or 0),       #AST
                float(data[21] or 0),       #STL           
                float(data[22] or 0),       #BLK
                float(data[20] or 0),       #TOV
                float(data[24] or 0),       #PF
                float(data[26] or 0),       #PTS   
                float(data[27] or 0),       #PLUS_MINUS   

                player_id
                )) 
  



        #Grab the rest of the splits  (ADVANCED, MISC, SCORING, USAGE). FOURFACTORS and TRACKING do not exist here and are only available from the Games Boxscores
        type='Advanced'
        splits_url = "https://stats.nba.com/stats/playerdashboardbygeneralsplits?DateFrom=&DateTo=&GameSegment=&LastNGames=0&LeagueID=00&Location=&MeasureType="+type+\
                    "&Month=0&OpponentTeamID=0&Outcome=&PORound=0&PaceAdjust=N&PerMode=PerGame&Period=0&PlayerID="+str(player_id)+\
                    "&PlusMinus=N&Rank=N&Season="+season_id+\
                    "&SeasonSegment=&SeasonType="+season_type+\
                    "+Season&ShotClockRange=&Split=general&VsConference=&VsDivision="


        try:  
            time.sleep(3)
            response_splits_ADVANCED = requests.get(splits_url, headers=headers)
            #response_splits_ADVANCED = NBA_web_session.get(splits_url, timeout=mytimeout)
            JSON_splits_advanced = json.loads(response_splits_ADVANCED.text)
            ADVANCED_split_stats = np.array(JSON_splits_advanced['resultSets'][0]['rowSet'])
        except Timeout as ex:
            sys.exit('Advanced splits request failed: ', ex)


        
        if ADVANCED_split_stats[0][0]=='Overall':
            data=ADVANCED_split_stats[0]


            if stats_cursor == None: #save the data and return it in the end
                ADVANCED_DATA_ORDERED=np.array([
                float(data[7] or 0),        #E_OFF_RATING
                float(data[8] or 0),        #OFF_RATING
                float(data[10] or 0),        #E_DEF_RATING
                float(data[11] or 0),        #DEF_RATING
                float(data[13] or 0),       #E_NET_RATING
                float(data[14] or 0),       #NET_RATING
                float(data[16] or 0),       #AST_PCT
                float(data[17] or 0),       #AST_TOV
                float(data[18] or 0),       #AST_RATIO
                float(data[19] or 0),       #OREB_PCT
                float(data[20] or 0),       #DREB_PCT
                float(data[21] or 0),       #REB_PCT
                float(data[22] or 0),       #TOV_PCT
                float(data[24] or 0),       #EFG_PCT
                float(data[25] or 0),       #TS_PCT           
                float(data[26] or 0),       #USG_PCT
                float(data[27] or 0),       #E_USG_PCT
                float(data[28] or 0),       #E_PACE
                float(data[29] or 0),       #PACE   
                float(data[32] or 0)       #PIE   
                ])
            
            else:  #add it to the database

                stats_cursor.execute('UPDATE Players SET  \
                E_OFF_RATING= ?, OFF_RATING= ?, E_DEF_RATING= ?,  DEF_RATING= ?, E_NET_RATING= ?, NET_RATING= ?, AST_PCT= ?, AST_TOV= ?, AST_RATIO= ?, OREB_PCT= ?, DREB_PCT= ?, REB_PCT= ?, TOV_PCT= ?, EFG_PCT= ?, TS_PCT= ?,  USG_PCT= ?, E_USG_PCT= ?, E_PACE= ?, PACE= ?, PIE=?\
                WHERE player_id = ?',\
                (
                float(data[7] or 0),        #E_OFF_RATING
                float(data[8] or 0),        #OFF_RATING
                float(data[10] or 0),        #E_DEF_RATING
                float(data[11] or 0),        #DEF_RATING
                float(data[13] or 0),       #E_NET_RATING
                float(data[14] or 0),       #NET_RATING
                float(data[16] or 0),       #AST_PCT
                float(data[17] or 0),       #AST_TOV
                float(data[18] or 0),       #AST_RATIO
                float(data[19] or 0),       #OREB_PCT
                float(data[20] or 0),       #DREB_PCT
                float(data[21] or 0),       #REB_PCT
                float(data[22] or 0),       #TOV_PCT
                float(data[24] or 0),       #EFG_PCT
                float(data[25] or 0),       #TS_PCT           
                float(data[26] or 0),       #USG_PCT
                float(data[27] or 0),       #E_USG_PCT
                float(data[28] or 0),       #E_PACE
                float(data[29] or 0),       #PACE   
                float(data[32] or 0),       #PIE   

                player_id
                )) 
      



        type='Misc'
        splits_url = "https://stats.nba.com/stats/playerdashboardbygeneralsplits?DateFrom=&DateTo=&GameSegment=&LastNGames=0&LeagueID=00&Location=&MeasureType="+type+\
                    "&Month=0&OpponentTeamID=0&Outcome=&PORound=0&PaceAdjust=N&PerMode=PerGame&Period=0&PlayerID="+str(player_id)+\
                    "&PlusMinus=N&Rank=N&Season="+season_id+\
                    "&SeasonSegment=&SeasonType="+season_type+\
                    "+Season&ShotClockRange=&Split=general&VsConference=&VsDivision="
    
        try:
            time.sleep(3)
            response_splits_MISC = requests.get(splits_url, headers=headers)
            #response_splits_MISC = NBA_web_session.get(splits_url, timeout=mytimeout)
            JSON_splits_misc = json.loads(response_splits_MISC.text)
            MISC_split_stats = np.array(JSON_splits_misc['resultSets'][0]['rowSet'])
        except Timeout as ex:
            sys.exit('MISC splits request failed: ', ex)
        
        if MISC_split_stats[0][0]=='Overall':
            data=MISC_split_stats[0]

            if stats_cursor == None: #save the data and return it in the end
                MISC_DATA_ORDERED=np.array([
                    float(data[7] or 0),        #PTS_OFF_TOV
                    float(data[8] or 0),        #PTS_2ND_CHANCE
                    float(data[9] or 0),       #PTS_FB
                    float(data[10] or 0),       #PTS_PAINT
                    float(data[11] or 0),       #OPP_PTS_OFF_TOV
                    float(data[12] or 0),       #OPP_PTS_2ND_CHANCE
                    float(data[13] or 0),       #OPP_PTS_FB
                    float(data[14] or 0),       #OPP_PTS_PAINT
                    float(data[16] or 0),       #BLKA
                    float(data[18] or 0)       #PFD
                ])

            else:
                stats_cursor.execute('UPDATE Players SET  \
                PTS_OFF_TOV= ?, PTS_2ND_CHANCE= ?, PTS_FB= ?,  PTS_PAINT= ?, OPP_PTS_OFF_TOV= ?, OPP_PTS_2ND_CHANCE= ?, OPP_PTS_FB= ?, OPP_PTS_PAINT= ?, BLKA= ?, PFD= ?\
                WHERE player_id = ?',\
                (
                float(data[7] or 0),        #PTS_OFF_TOV
                float(data[8] or 0),        #PTS_2ND_CHANCE
                float(data[9] or 0),       #PTS_FB
                float(data[10] or 0),       #PTS_PAINT
                float(data[11] or 0),       #OPP_PTS_OFF_TOV
                float(data[12] or 0),       #OPP_PTS_2ND_CHANCE
                float(data[13] or 0),       #OPP_PTS_FB
                float(data[14] or 0),       #OPP_PTS_PAINT
                float(data[16] or 0),       #BLKA
                float(data[18] or 0),       #PFD

                player_id
                ))   
 


        type='Scoring'
        splits_url = "https://stats.nba.com/stats/playerdashboardbygeneralsplits?DateFrom=&DateTo=&GameSegment=&LastNGames=0&LeagueID=00&Location=&MeasureType="+type+\
                    "&Month=0&OpponentTeamID=0&Outcome=&PORound=0&PaceAdjust=N&PerMode=PerGame&Period=0&PlayerID="+str(player_id)+\
                    "&PlusMinus=N&Rank=N&Season="+season_id+\
                    "&SeasonSegment=&SeasonType="+season_type+\
                    "+Season&ShotClockRange=&Split=general&VsConference=&VsDivision="

        try:
            time.sleep(3)
            #response_splits_SCORING = NBA_web_session.get(splits_url, timeout=mytimeout)
            response_splits_SCORING = requests.get(splits_url, headers=headers)
            JSON_splits_scoring = json.loads(response_splits_SCORING.text)
            SCORING_split_stats = np.array(JSON_splits_scoring['resultSets'][0]['rowSet'])
        except Timeout as ex:
            sys.exit('SCORING splits request failed: ', ex)



        if SCORING_split_stats[0][0]=='Overall':
            data=SCORING_split_stats[0]

            
            if stats_cursor == None: #save the data and return it in the end
                
                SCORING_DATA_ORDERED=np.array([
                float(data[7] or 0),        #PCT_FGA_2PT
                float(data[8] or 0),        #PCT_FGA_3PT
                float(data[9] or 0),        #PCT_PTS_2PT
                float(data[10] or 0),       #PCT_PTS_2PT_MR
                float(data[11] or 0),       #PCT_PTS_3PT
                float(data[12] or 0),       #PCT_PTS_FB
                float(data[13] or 0),       #PCT_PTS_FT
                float(data[14] or 0),       #PCT_PTS_OFF_TOV
                float(data[15] or 0),       #PCT_PTS_PAINT
                float(data[16] or 0),       #PCT_AST_2PM
                float(data[17] or 0),       #PCT_UAST_2PM
                float(data[18] or 0),       #PCT_AST_3PM
                float(data[19] or 0),       #PCT_UAST_3PM
                float(data[20] or 0),       #PCT_AST_FGM
                float(data[21] or 0)       #PCT_UAST_FGM

                ])

            else:
                stats_cursor.execute('UPDATE Players SET  \
                PCT_FGA_2PT= ?, PCT_FGA_3PT= ?, PCT_PTS_2PT= ?, PCT_PTS_2PT_MR= ?, PCT_PTS_3PT= ?, PCT_PTS_FB= ?,\
                PCT_PTS_FT= ?, PCT_PTS_OFF_TOV= ?, PCT_PTS_PAINT= ?, PCT_AST_2PM= ?, PCT_UAST_2PM= ?, PCT_AST_3PM= ?,\
                PCT_UAST_3PM= ?, PCT_AST_FGM= ?, PCT_UAST_FGM= ?\
                WHERE player_id = ?',\
                (
                float(data[7] or 0),        #PCT_FGA_2PT
                float(data[8] or 0),        #PCT_FGA_3PT
                float(data[9] or 0),        #PCT_PTS_2PT
                float(data[10] or 0),       #PCT_PTS_2PT_MR
                float(data[11] or 0),       #PCT_PTS_3PT
                float(data[12] or 0),       #PCT_PTS_FB
                float(data[13] or 0),       #PCT_PTS_FT
                float(data[14] or 0),       #PCT_PTS_OFF_TOV
                float(data[15] or 0),       #PCT_PTS_PAINT
                float(data[16] or 0),       #PCT_AST_2PM
                float(data[17] or 0),       #PCT_UAST_2PM
                float(data[18] or 0),       #PCT_AST_3PM
                float(data[19] or 0),       #PCT_UAST_3PM
                float(data[20] or 0),       #PCT_AST_FGM
                float(data[21] or 0),       #PCT_UAST_FGM

                player_id
                ))   



        type='Usage'
        splits_url = "https://stats.nba.com/stats/playerdashboardbygeneralsplits?DateFrom=&DateTo=&GameSegment=&LastNGames=0&LeagueID=00&Location=&MeasureType="+type+\
            "&Month=0&OpponentTeamID=0&Outcome=&PORound=0&PaceAdjust=N&PerMode=PerGame&Period=0&PlayerID="+str(player_id)+\
            "&PlusMinus=N&Rank=N&Season="+season_id+\
            "&SeasonSegment=&SeasonType="+season_type+\
            "+Season&ShotClockRange=&Split=general&VsConference=&VsDivision="



        try:
            time.sleep(3)
            #response_splits_USAGE = NBA_web_session.get(splits_url, timeout=mytimeout)
            response_splits_USAGE = requests.get(splits_url, headers=headers)
            JSON_splits_usage = json.loads(response_splits_USAGE.text)
            USAGE_split_stats = np.array(JSON_splits_usage['resultSets'][0]['rowSet'])
        except Timeout as ex:
            sys.exit('USAGE splits request failed: ', ex)



        if USAGE_split_stats[0][0]=='Overall':
            data=USAGE_split_stats[0]


            if stats_cursor == None: #save the data and return it in the end
                
                USAGE_DATA_ORDERED=np.array([
                    float(data[7] or 0),        #USG_PCT_USG
                    float(data[8] or 0),        #PCT_FGM
                    float(data[9] or 0),        #PCT_FGA
                    float(data[10] or 0),       #PCT_FG3M
                    float(data[11] or 0),       #PCT_FG3A
                    float(data[12] or 0),       #PCT_FTM
                    float(data[13] or 0),       #PCT_FTA
                    float(data[14] or 0),       #PCT_OREB
                    float(data[15] or 0),       #PCT_DREB
                    float(data[16] or 0),       #PCT_REB
                    float(data[17] or 0),       #PCT_AST
                    float(data[18] or 0),       #PCT_TOV
                    float(data[19] or 0),       #PCT_STL
                    float(data[20] or 0),       #PCT_BLK
                    float(data[21] or 0),       #PCT_BLKA
                    float(data[22] or 0),       #PCT_PF
                    float(data[23] or 0),       #PCT_PFD
                    float(data[24] or 0)       #PCT_PTS
                ])


            else:

                stats_cursor.execute('UPDATE Players SET  \
                USG_PCT_USG= ?, PCT_FGM= ?, PCT_FGA= ?, PCT_FG3M= ?, PCT_FG3A= ?, PCT_FTM= ?, PCT_FTA= ?, PCT_OREB= ?, PCT_DREB= ?, \
                PCT_REB= ?, PCT_AST= ?, PCT_TOV= ?, PCT_STL= ?, PCT_BLK= ?, PCT_BLKA= ?, PCT_PF= ?, PCT_PFD= ?, PCT_PTS= ?\
                WHERE player_id = ?',\
                (
                float(data[7] or 0),        #USG_PCT_USG
                float(data[8] or 0),        #PCT_FGM
                float(data[9] or 0),        #PCT_FGA
                float(data[10] or 0),       #PCT_FG3M
                float(data[11] or 0),       #PCT_FG3A
                float(data[12] or 0),       #PCT_FTM
                float(data[13] or 0),       #PCT_FTA
                float(data[14] or 0),       #PCT_OREB
                float(data[15] or 0),       #PCT_DREB
                float(data[16] or 0),       #PCT_REB
                float(data[17] or 0),       #PCT_AST
                float(data[18] or 0),       #PCT_TOV
                float(data[19] or 0),       #PCT_STL
                float(data[20] or 0),       #PCT_BLK
                float(data[21] or 0),       #PCT_BLKA
                float(data[22] or 0),       #PCT_PF
                float(data[23] or 0),       #PCT_PFD
                float(data[24] or 0),       #PCT_PTS

                player_id
                ))   



    return TRADITIONAL_DATA_ORDERED, ADVANCED_DATA_ORDERED, MISC_DATA_ORDERED, SCORING_DATA_ORDERED, USAGE_DATA_ORDERED


#Scrape PAST games data from stats.nba.com and save them to file. Games (and related) table(s) will be re-written from scratch
def scrape_NBA_past_games(database,counter,season,override_games):

    curs = database.conn.cursor()

    counter=counter+1
    no_game_counter=0
    print("Scraping Games from season "+season)

    
    
    while True:

        #Compose the url address for each game of the season       
        gameID="002"+season[2:4]+str(counter).zfill(5)

        if gameID in override_games:
            print("Manual override for game "+gameID+ "\t\t")
        else:

            #### MAKE WEB REQUESTS ######
                    
            #SUMMARY
            time.sleep(3)
            summary_url='https://stats.nba.com/stats/boxscoresummaryv2?GameID='+gameID
            response = requests.get(summary_url, headers=headers)


            if response.status_code==200: #check if we have a valid page
                boxscore_summary = json.loads(response.text)   
               
                #check if we have a valid, but empty page
                if boxscore_summary['resultSets'][0]['rowSet'] != []:

                    #Get game info
                    game_date=boxscore_summary['resultSets'][0]['rowSet'][0][0][0:10]
                    datetime_object=datetime.strptime(game_date, '%Y-%m-%d').date()
                    date_string= str(datetime_object.strftime("%m")+"/"+datetime_object.strftime("%d")+"/"+datetime_object.strftime("%y"))


                    teamID_Away=  boxscore_summary['resultSets'][0]['rowSet'][0][7]
                    teamID_Home=  boxscore_summary['resultSets'][0]['rowSet'][0][6]
                    
                    #determine which is the home and away teams (i.e. ordering)
                    if boxscore_summary['resultSets'][5]['rowSet'][0][3] == teamID_Away:
                        home_score = int(boxscore_summary['resultSets'][5]['rowSet'][1][22])
                        away_score = int(boxscore_summary['resultSets'][5]['rowSet'][0][22])

                        teamName_Home=boxscore_summary['resultSets'][5]['rowSet'][1][4]
                        teamName_Away=boxscore_summary['resultSets'][5]['rowSet'][0][4]
                    else:
                        away_score = int(boxscore_summary['resultSets'][5]['rowSet'][1][22])
                        home_score = int(boxscore_summary['resultSets'][5]['rowSet'][0][22])

                        teamName_Home=boxscore_summary['resultSets'][5]['rowSet'][0][4]
                        teamName_Away=boxscore_summary['resultSets'][5]['rowSet'][1][4]
                
                    #Insert to Games table
                    if home_score>away_score:
                        result=2
                    else:
                        result=1
                    curs.execute('INSERT  OR REPLACE INTO games (game_id,Away_Team_id,Home_Team_id,Away_Score,Home_Score,Result,Date) VALUES(?,?,?,?,?,?,?)',
                                        (gameID, teamID_Away, teamID_Home, away_score, home_score, result, date_string ) )



                    
                    #Scrape Lineups and Boxscore stats for the current game, here
                    process_BOXSCORE_stats(season,gameID,date_string,curs, override_player_ids)


                    database.conn.commit() #Best to commit after every game in case we lose connectivitiy


                    print("\tGame %i:   %s @ %s  %i-%i\t\t"  % (counter,teamName_Away,teamName_Home,away_score,home_score))                
                    no_game_counter=0

                else:
                   print("\t\t\t%ith game page is invalid. Skipping. \t\t "%(counter))       



            else: 
                no_game_counter=no_game_counter+1
                print("\t\t\tNo %ith game found. Trying %s of 10.\t\t "%(counter, no_game_counter))
                if no_game_counter>=10:      
                    print("\t\tNo more games\t\t\t\t\t")     
                    #End of games for the season


                    #Check games beyond 1230 (i.e. special season 2020)
                    if counter<1231:
                        counter=1230
                        no_game_counter=0
                        print('\t\tChecking games beyond 1230\t\t\t\t\t')

                    else:
                        break
            
                        
        counter=counter+1                      



def append_team_roster(cursor, team_id, season):

    team_roster_array=[]

    roster_url= 'https://stats.nba.com/stats/commonteamroster?LeagueID=00&Season='+season+'&TeamID='+str(team_id)
    response = requests.get(roster_url, headers=headers)

    if response.status_code==200: 
        roster_JSON = json.loads(response.text)
        roster = roster_JSON['resultSets'][0]['rowSet']
    
        for i in range(len(roster)):
            player_id = str(roster[i][-1])
            team_roster_array.append(player_id)
            col_name= 'Roster '+str(i+1)

            try:
                cursor.execute("alter table Teams add column '%s' 'TEXT'" % col_name)
            except:
                pass

            cursor.execute('UPDATE Teams SET "%s" = "%s"  WHERE team_id = "%s" ' % (
                    col_name, player_id, str(team_id)))
     
        
        #commit for each team 
        cursor.connection.commit()

        return team_roster_array

    else:
        print('Incorrect response code for Roster data. ', response.status_code)


 
#Scrape player & teams data for given season. Necessary since teams can change their roster throughout the season
def scrape_NBA_teamsNplayers(database, season):

    cursor = database.conn.cursor()

    #TEAMS 
    ALL_teams_rosters=[]
    teams_sz=len(TeamsList.Teams)
    for i in range(teams_sz): #loop over teams
        #Insert to Teams table
        cursor.execute('INSERT OR REPLACE INTO teams (team_id, team_name, short_name) VALUES(?,?,?)',
                                  (TeamsList.Teams[i].ID, TeamsList.Teams[i].name, TeamsList.Teams[i].short) )

        #ADD team roster for the team for that season
        time.sleep(2) 
        team_roster=append_team_roster(cursor, TeamsList.Teams[i].ID, season)
        ALL_teams_rosters= ALL_teams_rosters+ team_roster

    #PLAYERS
    #scrape players directly from games (Linueps) data  +  From current ALL team rosters
    
    lineups= cursor.execute('SELECT player_id FROM Lineups').fetchall()
    lineups  = [''.join(i) for i in lineups]  #from tuple to list

    ALL_teams_rosters=ALL_teams_rosters+lineups #add lineups to all roster data
    ALL_teams_rosters = list(set(ALL_teams_rosters)) #remove duplicates

    NBA_web_session.headers= headers_ALT
    for j in range(len(ALL_teams_rosters)): #loop over lineups
        player_id=ALL_teams_rosters[j]


        
        #Check if player already exists in Players table
        exist_check = cursor.execute('SELECT * FROM Players WHERE player_id='+player_id).fetchall()
        if len(exist_check)==0:
            #Add player to the table

            player_url='https://stats.nba.com/player/'+player_id+'/career/'
            time.sleep(2)  
            
            response = NBA_web_session.get(player_url)
            
            HTMLtext = response.text.replace("<br />"," ")  #Remove line breaks and replace them with space
            soupify = BeautifulSoup(HTMLtext, 'html.parser')  # Parse the HTML as a string.


            #TODO: FIND A BETTER WAY TO SCAPE PLAYERS PERSONAL INFO

            #get summary tab for player
            summary_tab= soupify.find('div', {'class': 'player-name'})
            first_name = summary_tab.find_all('div', {'class': 'stats-teamplayer-summary-text large'})[0].text.strip()
            last_name = summary_tab.find_all('div', {'class': 'stats-teamplayer-summary-text large'})[1].text.strip()
            player_name = first_name +" "+last_name





            personal = soupify.find('div', {'class': 'personal'}) 
            personal_section= personal.find_all('div', {'class': 'personal-section'})[1]
            personal_info_border_right =  personal_section.find_all('div', {'class': 'personal-info border-right'})[1]
            born =  personal_info_border_right.find_all('div', {'class': 'stats-teamplayer-summary-text'})[1].text[-4:]

            draft_year = personal_section.find_all('div', {'class': 'personal-info border-right'})[2].find_all('div', {'class': 'stats-teamplayer-summary-text'})[1].text.split()[0]
            age= int(season[0:4])-int(born)

            if player_name !=' ' and (age>17 and age <60) : #Verify that we have a valid player

                exp = personal_section.find_all('div', {'class': 'personal-info'})[3].find_all('div', {'class': 'stats-teamplayer-summary-text'})[1].find_all('span')[0].text.split()[0]
                if exp =="R":
                    exp=0
                else:
                    exp=int(exp)
         

  
                personal_section= personal.find_all('div', {'class': 'personal-section'})[0]
                try:
                    weight= int(personal_section.find_all('div', {'class': 'personal-info border-right'})[1].find_all('div', {'class': 'stats-teamplayer-summary-text'})[1].find('span', {'ng-if': 'playerInfo.WEIGHT'}).text[0:-2] or 0)
                except:
                    pass
                
                personal_info_border_right =  personal_section.find_all('div', {'class': 'personal-info border-right'})[0]
                
                #height broken on the new version of the webpage
                height = 0.0
                

                #Insert to Players table
                try:
                    print("\tAdding : %s \t\t\t\t\t\t"% player_name , end="\r", flush=True)
                    cursor.execute('INSERT  INTO Players (player_id, player_name, age, exp, weight, height) VALUES(?,?,?,?,?,?)',
                                                (player_id,  player_name, age, exp, weight, height ) )
                except:
                        print("\tPlayer ",player_name, "already exists. Skipping")

            else:
                print('\n\tinvalid player found: %s (%s)\n'%(player_name, player_id))


    database.conn.commit()





def calculateAvg_season_Player_Stats(database, players_that_played_index, verbose=True):

    
    cursor = database.conn.cursor()

    if verbose==True:
        print("Calculating SELECTED average player stats  \t\t")

    #get the boxscore attributes. Includes game_id and player_id cols
    column_names=[]
    BoxscoresPlayer_array = np.array(cursor.execute('SELECT * FROM BoxscoresPlayer').fetchall())
    for i in range(len(cursor.description)):
        column_names.append(cursor.description[i][0])
    ncols= len(column_names)
    Players_array= np.array(cursor.execute('SELECT * FROM Players').fetchall())
    ncols_target=len(cursor.description)



    #first add the new columns to the database Do this only the first time around
    if ncols_target<7:
        for cols in range(2,ncols): #skip the game_id and player_id cols and only look at the statistics
            cursor.execute("ALTER TABLE Players ADD COLUMN "+column_names[cols]+" FLOAT")  #use sqlite directly


    if len(Players_array)>0 and len(BoxscoresPlayer_array)>0:
            
        #loop over players
        if players_that_played_index is None:
            players_that_played_index =  Players_array[:,0] #all the players in the database

        for pl_row in range(len(players_that_played_index)):
            
            if verbose==True:
                print("\tPlayer: %i of %i" % (pl_row+1, len(players_that_played_index)) , end="\r", flush=True)
            
            
            player_id= players_that_played_index[pl_row]





            #Find all games played by that player 
            games_played_index = BoxscoresPlayer_array[:,1]==player_id
            nrows=np.sum(games_played_index)
        
            AvgStats = np.mean(BoxscoresPlayer_array[games_played_index,2:].astype(float),axis=0)


            #add average values to table Single hardcoded update. Much faster than for loop
            cursor.execute('UPDATE Players SET \
                    MIN=?, FGM=?, FGA=?, FG_PCT=?, FG3M=?, FG3A=?, FG3_PCT=?, FTM=?, FTA=?, FT_PCT=?, OREB=?, DREB=?, REB=?, AST=?, STL=?, BLK=?, TOV=?, PF=?, PTS=?, PLUS_MINUS=?, \
                    E_OFF_RATING=?, OFF_RATING=?, E_DEF_RATING=?, DEF_RATING=?, E_NET_RATING=?, NET_RATING=?, AST_PCT=?, AST_TOV=?, AST_RATIO=?, OREB_PCT=?, DREB_PCT=?, REB_PCT=?, TOV_PCT=?, EFG_PCT=?, TS_PCT=?, USG_PCT=?, E_USG_PCT=?, E_PACE=?, PACE=?, PIE=?, \
                    PTS_OFF_TOV=?,  PTS_2ND_CHANCE=?, PTS_FB=?, PTS_PAINT=?, OPP_PTS_OFF_TOV=?, OPP_PTS_2ND_CHANCE=?, OPP_PTS_FB=?, OPP_PTS_PAINT=?, BLKA=?, PFD=?, \
                    PCT_FGA_2PT=?, PCT_FGA_3PT=?, PCT_PTS_2PT=?, PCT_PTS_2PT_MR=?, PCT_PTS_3PT=?, PCT_PTS_FB=?, PCT_PTS_FT=?, PCT_PTS_OFF_TOV=?, PCT_PTS_PAINT=?, PCT_AST_2PM=?, PCT_UAST_2PM=?, PCT_AST_3PM=?, PCT_UAST_3PM=?, PCT_AST_FGM=?, PCT_UAST_FGM=?, \
                    USG_PCT_USG=?, PCT_FGM=?, PCT_FGA=?, PCT_FG3M=?, PCT_FG3A=?, PCT_FTM=?, PCT_FTA=?, PCT_OREB=?, PCT_DREB=?, PCT_REB=?, PCT_AST=?, PCT_TOV=?, PCT_STL=?, PCT_BLK=?, PCT_BLKA=?, PCT_PF=?, PCT_PFD=?, PCT_PTS=?, \
                    EFG_PCT_FF=?, FTA_RATE=?, TM_TOV_PCT=?, OREB_PCT_FF=?, OPP_EFG_PCT=?, OPP_FTA_RATE=?, OPP_TOV_PCT=?, OPP_OREB_PCT=?, \
                    SPD=?, DIST=?, ORBC=?, DRBC=?, RBC=?, TCHS=?, SAST=?, FTAST=?, PASS=?, CFGM=?, CFGA=?, CFG_PCT=?, UFGM=?, UFGA=?, UFG_PCT=?, DFGM=?, DFGA=?, DFG_PCT=?, \
                    CONTESTED_SHOTS=?, CONTESTED_SHOTS_2PT=?, CONTESTED_SHOTS_3PT=?, DEFLECTIONS=?, CHARGES_DRAWN=?, SCREEN_ASSISTS=?, SCREEN_AST_PTS=?, OFF_LOOSE_BALLS_RECOVERED=?, DEF_LOOSE_BALLS_RECOVERED=?, LOOSE_BALLS_RECOVERED=?, OFF_BOXOUTS=?, DEF_BOXOUTS=?, BOX_OUT_PLAYER_TEAM_REBS=?, BOX_OUT_PLAYER_REBS=?,  BOX_OUTS=?  '
            
            'WHERE player_id = '+player_id, \
            (AvgStats[0],  AvgStats[1],  AvgStats[2],  AvgStats[3],  AvgStats[4], AvgStats[5], AvgStats[6], AvgStats[7], AvgStats[8], AvgStats[9], AvgStats[10], AvgStats[11], AvgStats[12], AvgStats[13], AvgStats[14], AvgStats[15], AvgStats[16], AvgStats[17], AvgStats[18], AvgStats[19],
            AvgStats[20], AvgStats[21], AvgStats[22], AvgStats[23], AvgStats[24], AvgStats[25], AvgStats[26], AvgStats[27], AvgStats[28], AvgStats[29], AvgStats[30], AvgStats[31], AvgStats[32], AvgStats[33], AvgStats[34], AvgStats[35], AvgStats[36], AvgStats[37], AvgStats[38], AvgStats[39],
            AvgStats[40], AvgStats[41], AvgStats[42], AvgStats[43], AvgStats[44], AvgStats[45], AvgStats[46], AvgStats[47], AvgStats[48], AvgStats[49],
            AvgStats[50], AvgStats[51], AvgStats[52], AvgStats[53], AvgStats[54], AvgStats[55], AvgStats[56], AvgStats[57], AvgStats[58], AvgStats[59], AvgStats[60], AvgStats[61], AvgStats[62], AvgStats[63], AvgStats[64],
            AvgStats[65], AvgStats[66], AvgStats[67], AvgStats[68], AvgStats[69], AvgStats[70], AvgStats[71], AvgStats[72], AvgStats[73], AvgStats[74], AvgStats[75], AvgStats[76], AvgStats[77], AvgStats[78], AvgStats[79], AvgStats[80], AvgStats[81], AvgStats[82],
            AvgStats[83], AvgStats[84], AvgStats[85], AvgStats[86], AvgStats[87], AvgStats[88], AvgStats[89], AvgStats[90],
            AvgStats[91], AvgStats[92], AvgStats[93], AvgStats[94], AvgStats[95], AvgStats[96], AvgStats[97], AvgStats[98], AvgStats[99], AvgStats[100], AvgStats[101], AvgStats[102], AvgStats[103], AvgStats[104], AvgStats[105], AvgStats[106], AvgStats[107], AvgStats[108],
            AvgStats[109], AvgStats[110], AvgStats[111], AvgStats[112], AvgStats[113], AvgStats[114], AvgStats[115], AvgStats[116], AvgStats[117], AvgStats[118], AvgStats[119], AvgStats[120], AvgStats[121], AvgStats[122], AvgStats[123]
            ))


        database.conn.commit()



def refine_Player_Stats(database,season, verbose=True):
    #Scrape avg stats directly for players since some old games have problematic Boxscore data

    stats_cursor = database.conn.cursor()
    all_players = stats_cursor.execute('SELECT * FROM Players ').fetchall()

    counter=1
    for player in all_players:
        player_id=player[0]

        if verbose:
            print('\t\tCHECKING %i: player_id = %s\t\t'%(counter, player_id), end="\r", flush=True)
            counter = counter +1

        #Test if the plater has a SPLITS tab in their profile. If not then most likely and older player, whose Boxscore data
        #is not reliable and needs to be refined
        splits_url = "https://stats.nba.com/stats/playerdashboardbygeneralsplits?DateFrom=&DateTo=&GameSegment=&LastNGames=0&LeagueID=00&Location=&MeasureType=Base"\
                    "&Month=0&OpponentTeamID=0&Outcome=&PORound=0&PaceAdjust=N&PerMode=PerGame&Period=0&PlayerID="+str(player_id)+\
                    "&PlusMinus=N&Rank=N&Season="+season+\
                    "&SeasonSegment=&SeasonType=Regular+Season&ShotClockRange=&Split=general&VsConference=&VsDivision="

        try:
            time.sleep(3)
            test_response = requests.get(splits_url, headers=headers)
            JSON_test = json.loads(test_response.text)
            TRADITIONAL_test = np.array(JSON_test['resultSets'][0]['rowSet'])
        except Timeout as ex:
            sys.exit('Traditional splits request failed: ', ex)


        if TRADITIONAL_test.size==0:
            #no SPLITS data then older player so capture average career stats 

 
            if verbose:
                print('Refining...')
                counter = counter +1
            process_STATS(stats_cursor, player_id, season)


    database.conn.commit()
    


def scrape_season_Database(season, path):

    #3 cases. 1) Either the table exists and is complete, 2) Exist but is incomplete, or 3) Does not exist, so new (incomplete)
     
    

    dbfile= path+"/NBA_season"+season+".db"
    dbfile_inc=path+"/NBA_season"+season+"_INCOMPLETE.db"


    if os.path.isfile(dbfile):
        #Case 1). Table exists and is complete
        complete=True
        dbms=open_database(dbfile)
        #check if the season average stats have already been calculated for the teams and players
        pl_avg_stats_flag= dbms.check_for_Player_avg_tats_calculation() 


    elif os.path.isfile(dbfile_inc):
        
        #Case 2) Exist but is incomplete
        complete=False
        dbms=open_database(dbfile_inc)

        num_games=dbms.get_num_games()
    
        print("Incomplete database found with %i games. Re-scraping game %i and continuing" % (num_games,num_games))
        scrape_NBA_past_games(dbms,max(0,num_games-1), season, overide_games_list)  #overwrite the last game since we do not know where the scrape was interrupted
        scrape_NBA_teamsNplayers(dbms,season) 

 
        dbms.closeConnection()
        
        #Rename incomplete file then reload dbms file as well
        dbfile_new = path+"/NBA_season"+season+".db"
        os.rename(dbfile_inc, dbfile_new)
        dbms=open_database(dbfile_new)

        pl_avg_stats_flag=False

    else:

        #Case 3) Does not exist, so new (incomplete)
        complete=False
        dbms,num_games, pl_avg_stats_flag=create_season_Database(dbfile_inc)      

        print("Database not found. Creating new season database")
        scrape_NBA_past_games(dbms,max(0,num_games-1), season, overide_games_list)  #overwrite the last game since we do not know where the scrape was interrupted
        scrape_NBA_teamsNplayers(dbms,season) 


        dbms.closeConnection()

        #Rename incomplete file then reload dbms file as well
        dbfile_new = path+"/NBA_season"+season+".db"
        os.rename(dbfile_inc, dbfile_new)
        dbms=open_database(dbfile_new)
      


    #Calculate season game-average stats for players if they do not exist
    if pl_avg_stats_flag==False or  not complete:
        calculateAvg_season_Player_Stats(dbms,None)
        #if it is an older dataset then also refine players
        if int(season[0:4])<2000:
            refine_Player_Stats(dbms,season)

    dbms.closeConnection()



 
#Scrape SCHEDULED games data from stats.nba.com for specific date
def scrape_NBA_scheduled_games(filename, injury_list_FILENAME, experts_FILENAME, Teams, D, M, Y):    
    
    
    if  os.path.isfile(filename):
        answer=input("file: "+filename+" already exists. Overwrite y/n? ")   
    else:
        answer="Y"
        
    injury_list_exist=False
    if os.path.isfile(injury_list_FILENAME):
        injury_list_exist=True

    experts_filename_exist=False
    if os.path.isfile(experts_FILENAME):
        experts_filename_exist=True  


    if answer.capitalize()=="Y":
        #Create a workbook and add a worksheet.
        workbook = openpyxl.Workbook()
        worksheet = workbook.worksheets[0]
            
        #Header
        worksheet.cell(1,1).value =  "Away" # Cell indices start from 1
        worksheet.cell(1,2).value =  "Home" # Cell indices start from 1
        worksheet.cell(1,3).value =  "Date" # Cell indices start from 1

        
        #Also generate the injury file if it doesn't exist
        if injury_list_exist==False:

            injury_workbook = openpyxl.Workbook()
            injury_worksheet = injury_workbook.worksheets[0]

            injury_worksheet.cell(1,1).value ="Away"                 # Cell indices start from 1
            injury_worksheet.cell(1,2).value ="Home"              # Cell indices start from 1
            injury_worksheet.cell(1,3).value ="Date"                # Cell indices start from 1
            injury_worksheet.cell(1,4).value ="Away Injuries"         # Cell indices start from 1
            injury_worksheet.cell(1,5).value ="Home Injuries"         # Cell indices start from 1

        

        
        #And experts file
        if experts_filename_exist==False:

            experts_workbook = openpyxl.Workbook()
            experts_worksheet = experts_workbook.worksheets[0]

            experts_worksheet.cell(1,1).value ="Away"                 # Cell indices start from 1
            experts_worksheet.cell(1,2).value ="Home"                 # Cell indices start from 1
            experts_worksheet.cell(1,3).value ="Date"                 # Cell indices start from 1

            

           
        worksheet.cell(1,4).value ="Our prediction"                # Cell indices start from 1
        worksheet.cell(1,5).value ="Our probs"                # Cell indices start from 1

        
        worksheet.cell(1,7).value ="Meta prediction"                # Cell indices start from 1        
        worksheet.cell(1,8).value ="Meta probs"                # Cell indices start from 1        
        worksheet.cell(1,10).value ="Stake"                # Cell indices start from 1        
        worksheet.cell(1,11).value ="Bets"                # Cell indices start from 1        


        #The market 
        worksheet.cell(1,12).value ="MARKET odds"                # Cell indices start from 1        
        worksheet.cell(1,14).value ="MARKET predictions"                # Cell indices start from 1        
        worksheet.cell(1,15).value ="MARKET probs"                # Cell indices start from 1        


        
        
        url='http://stats.nba.com/stats/scoreboardV2?DayOffset=0&LeagueID=00&gameDate='+M+'%2F'+D+'%2F'+Y
        

                       
        time.sleep(2)    
        response = requests.get(url, headers=headers)
    
           
        if response.status_code==200: 
                print("Scraping scheduled games for "+D+"/"+M+"/"+Y)
                 
                schedule_data = json.loads(response.text)
                schedule = schedule_data['resultSets'][0]['rowSet']
                
                #loop for scheduled games
                for i in range (len(schedule)): 
                    HomeTeam_ID=schedule[i][6]
                    AwayTeam_ID=schedule[i][7]
                    Date=schedule[i][5][:8]
                    for j in range(len(Teams)):
                        if AwayTeam_ID==Teams[j].ID:
                            TeamAway_short=Teams[j].short
                        if HomeTeam_ID==Teams[j].ID:
                            TeamHome_short=Teams[j].short
                    
                    #Write teams data
                    worksheet.cell(3+i,1).value =TeamAway_short               # Cell indices start from 1        
                    worksheet.cell(3+i,2).value =TeamHome_short               # Cell indices start from 1        

                    
                    #Write date data
                    date_string= str(Date[4:6]+"/"+Date[6:8]+"/"+Date[2:4])
                    worksheet.cell(3+i,3).value =date_string               # Cell indices start from 1        


                    if injury_list_exist==False:
                        injury_worksheet.cell(3+i,1).value =TeamAway_short               # Cell indices start from 1        
                        injury_worksheet.cell(3+i,2).value =TeamHome_short               # Cell indices start from 1        
                        injury_worksheet.cell(3+i,3).value =date_string     #month, day, year format. Cell indices start from 1        
                        
         
                    
                    if experts_filename_exist==False:
                        experts_worksheet.cell(3+i,1).value =TeamAway_short               # Cell indices start from 1        
                        experts_worksheet.cell(3+i,2).value =TeamHome_short               # Cell indices start from 1        
                        experts_worksheet.cell(3+i,3).value =date_string      #month, day, year format. Cell indices start from 1        


                   
        else:
                print("Web query failed: ", response.status_code)
                   
          
        workbook.save(filename)

        if injury_list_exist==False:
            injury_workbook.save(injury_list_FILENAME)
            #Populate injury list
            print("Populating injury list...")
            captureInjuryList(injury_list_FILENAME)



        if experts_filename_exist==False:
            experts_workbook.save(experts_FILENAME)





        

        #Add Betfair data
        answer = input("Append BETFAIR data? (y/n) ") 
        if  answer=="Y" or answer=="y": 
            token = input("Enter Session Token: ")
            print('Appending BETFAIR data')
            captureBetfairOdds(filename, token, starting_column=11)

      
        return 1
    
    else: 
            
        return 0
 
