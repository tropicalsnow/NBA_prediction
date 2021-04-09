import os.path
import openpyxl
import numpy as np
import sys
import requests
import pandas as pd


from TeamsList import getTeam_by_partial_ANY, getTeam_by_partial_Name

from featureUtilities import Spread2Prob_Convert


def capturePredictionTrackerData(game_file,  starting_column=38):

    download_url = "https://www.thepredictiontracker.com/nbapreds.csv"
    csv_file="nbapreds.csv"


    #Parse history files
    game_workbook = openpyxl.load_workbook(game_file)
    game_worksheet = game_workbook.worksheets[0]
    
    n_games=game_worksheet.max_row-2


    predictionEntries = []


    #Read full csv file first
    if not os.path.isfile(csv_file):
        print("CSV file does not exist. Attempting to download")

        r = requests.get(download_url)
        with open(csv_file, 'wb') as outfile:
            outfile.write(r.content)

     

    #open the csv and import it to pandas dataframe
    expert_DF = pd.read_csv(csv_file) 


    #replace team names to standard stats.nba.com abbreviations
    home_col_index = expert_DF.columns.get_loc("home")
    road_col_index = expert_DF.columns.get_loc("road")

    for i in range(len(expert_DF)):
        
        try:
            tH=getTeam_by_partial_Name(expert_DF.iloc[i,home_col_index]).short
        except:
            try:
                tH=getTeam_by_partial_ANY(expert_DF.iloc[i,home_col_index]).short
            except:
                tH=None

        expert_DF.iloc[i,home_col_index] = tH

        try:
            tA=getTeam_by_partial_Name(expert_DF.iloc[i,road_col_index]).short
        except:
            try:
                tA=getTeam_by_partial_ANY(expert_DF.iloc[i,road_col_index]).short
            except:
                tA=None
        
        expert_DF.iloc[i,road_col_index] = tA




    #grab excel file header
    col_head = []
    hindx=1
    while True:
        header= game_worksheet.cell(1, hindx).value
        if header is None:
            break
        col_head.append(header)
        hindx=hindx+1



    #add header to end
    for cols in expert_DF.columns:
        if (cols not in  ['home','road','std']) and (cols not in col_head):
            game_worksheet.cell(1, len(col_head)+1).value= cols
            col_head.append(cols)





    #Compare with games data. No dates so just compare on team names and possibly number of predictions = number of games
    prediction_length=len(expert_DF)
    if n_games != prediction_length:
        print('WARNING: Number of predictions are not the same with number of games')


    for i in  range(n_games):
        #Get game info
        away_team_Short =           game_worksheet.cell(i+2+1, 0+1).value #cell indices start from 1
        home_team_Short =           game_worksheet.cell(i+2+1, 1+1).value #cell indices start from 1


        row_out = expert_DF.loc[(expert_DF['road'] == away_team_Short) & (expert_DF['home'] == home_team_Short)]

        #drop road, home and std columns
        row_out=row_out.drop(columns=['home','road','std'])
        row_out=row_out.fillna(0)
        for c_num in range(1,len(col_head)+1):
            worksheet_col_name = game_worksheet.cell(1, c_num).value

            #get the line and convert it to probs
            try:
                line = row_out[worksheet_col_name].iloc[0]  #wrt to Home team. i.e. positive line means the home team has advantage

                if line == 0:
                    probs_AWAY=0.5 #set missing lines to 0.5 prob
                else:
                    probs_AWAY= Spread2Prob_Convert(line, [-0.04927,  0.1403, 0.5472, -0.04684])

        
                game_worksheet.cell(i+3, c_num).value = probs_AWAY
            
            except:
                pass

    
    game_workbook.save(game_file)  
    
